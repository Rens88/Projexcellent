#!/usr/bin/env python3
"""
Rapportage/make_report.py

What this script does
---------------------
1) Loads runtime settings from ../projexcellent_config.json.
2) Scans configured projects_dir and treats EACH direct subfolder as one project.
3) Derives project_id from the folder name: YYYY_NNNN_<description> -> YYYY_NNNN
4) Reads project metadata from project_info.xlsx (sheet: ProjectInfo, columns: Key/Value).
5) Validates hygiene rules:
   - Folder name format is valid
   - project_info.xlsx exists
   - project_id in project_info.xlsx matches the derived project_id
   - If status == "Closed" then actual_end_date must be filled
   - time_log.xlsx metadata project_id (cell B1) matches derived project_id (if present)
6) Reads time spent from time_log.xlsx (sheet: TimeLog, rows under the header),
   aggregates hours per project and per programma/requester.
7) Creates a single HTML report (Plotly) with:
   - Tabs: Counts / Hours
   - Period switcher: 1-day / 1-week / 2-weeks / month / year
   - (Optionally) single-period reports via `--report-type`
8) Exports to configured reports_dir:
   - project_report_with_hours.html (combined; default, includes Hours tab)
   - project_report.html (lite, no Hours tab)
   - Archive/*_with_hours_generated_YYYY-MM-DD.html (full)
   - Archive/*_generated_YYYY-MM-DD.html (lite)
   - Single-period exports also write a PNG (requires `pip install kaleido`)

Dependencies
------------
pip install pandas openpyxl plotly kaleido
"""

from __future__ import annotations

import argparse
import base64
import html
import os
import re
import shutil
import sys
import warnings
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from typing import Any, Dict, List, Optional, Set, Tuple

import pandas as pd
import plotly.graph_objects as go
import plotly.io as pio
from plotly.subplots import make_subplots

try:
    import holidays as holidays_lib  # Optional; used for national-holiday-aware workday counts.
except Exception:
    holidays_lib = None


# ----------------------------
# Paths and runtime configuration
# ----------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.normpath(os.path.join(SCRIPT_DIR, ".."))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

from projexcellent_config import DEFAULT_CONFIG_PATH, load_config, resolve_path, resolve_path_list

CONFIG_PATH = DEFAULT_CONFIG_PATH
CONFIG: Dict[str, Any] = {}
PROJECTEN_DIR = ""
DUMMY_PROJECTEN_DIR = ""
REPORT_DIR = ""
REPORTS_ARCHIVE_DIR = ""
ASSETS_DIR = ""
PROFILE_PHOTO_PATH = ""
TEAMNL_LOGO_PATH = ""
REPORT_TITLE = "Project Portfolio Overview"
PERSON_NAME = "john doe"
COMPANY_NAME = "NOC*NSF"
COMPANY_ABBREVIATION = "NN"
HOURS_REMAINING_EXCEL_PATHS: List[str] = []
HOURS_REMAINING_SHEET_NAME = "NN_maandelijks"
HOURS_REMAINING_HEADER_ROW = 2
WORKABLE_HOURS_PER_YEAR_OVERRIDE: Optional[float] = None
WORKABLE_HOURS_PER_WEEK_REFERENCE_VALUE: Optional[float] = None
USING_DUMMY_FALLBACK = False
REPORT_TYPE_CHOICES = ["combined", "yearly", "monthly", "biweekly", "weekly", "daily", "all"]
PROJECT_ROLE_STANDARD = "standard"
PROJECT_ROLE_COMPLETE_MISSING_HOURS = "complete_missing_hours"
DEFAULT_COMPLETE_MISSING_HOURS_FOLDER = "2026_0000_complete_missing_hours"
PROJECT_ROLE_KEY_TOKENS = ("projectrole", "reportingrole", "specialrole", "role")
COMPLETE_MISSING_HOURS_ROLE_TOKENS = (
    "completemissinghours",
    "completemissinghour",
    "missinghours",
    "missinghour",
)

def has_subfolders(path: str) -> bool:
    if not os.path.isdir(path):
        return False
    return any(
        os.path.isdir(os.path.join(path, entry))
        for entry in os.listdir(path)
    )

def _apply_runtime_config(config_path: Optional[str] = None) -> None:
    def _as_positive_float(raw: Any) -> Optional[float]:
        if raw is None:
            return None
        try:
            val = float(raw)
        except (TypeError, ValueError):
            return None
        if val <= 0:
            return None
        return val

    def _coerce_hex(raw: Any, fallback: str) -> str:
        txt = str(raw or "").strip()
        if not txt:
            return fallback
        if re.fullmatch(r"#?[0-9A-Fa-f]{6}", txt):
            return txt if txt.startswith("#") else f"#{txt}"
        return fallback

    def _coerce_cfg_color(primary_key: str, fallback: str, legacy_key: Optional[str] = None) -> str:
        raw = color_cfg.get(primary_key)
        if (raw is None or str(raw).strip() == "") and legacy_key:
            raw = color_cfg.get(legacy_key)
        return _coerce_hex(raw, fallback)

    global CONFIG_PATH
    global CONFIG
    global PROJECTEN_DIR
    global DUMMY_PROJECTEN_DIR
    global REPORT_DIR
    global REPORTS_ARCHIVE_DIR
    global ASSETS_DIR
    global PROFILE_PHOTO_PATH
    global TEAMNL_LOGO_PATH
    global REPORT_TITLE
    global PERSON_NAME
    global COMPANY_NAME
    global COMPANY_ABBREVIATION
    global HOURS_REMAINING_EXCEL_PATHS
    global HOURS_REMAINING_SHEET_NAME
    global HOURS_REMAINING_HEADER_ROW
    global WORKABLE_HOURS_PER_YEAR_OVERRIDE
    global WORKABLE_HOURS_PER_WEEK_REFERENCE_VALUE
    global USING_DUMMY_FALLBACK
    global BASE_BLUE
    global BASE_RED
    global BASE_ORANGE
    global BASE_YELLOW
    global BASE_GREEN
    global BASE_BLACK
    global TEAMNL_BASE_COLORS
    global YEAR_PLAN_COMPLETED_COLOR
    global YEAR_PLAN_CURRENT_BILLED_COLOR
    global YEAR_PLAN_CURRENT_COMBINED_COLOR
    global YEAR_PLAN_CURRENT_EXPECTED_COLOR
    global YEAR_PLAN_EXPECTED_COLOR

    CONFIG_PATH = config_path or DEFAULT_CONFIG_PATH
    CONFIG = load_config(CONFIG_PATH)

    paths_cfg = CONFIG.get("paths", {})
    branding_cfg = CONFIG.get("branding", {})
    runtime_cfg = CONFIG.get("runtime", {})
    company_cfg = CONFIG.get("company", {})
    hours_cfg = CONFIG.get("hours", {})
    color_cfg = CONFIG.get("color_scheme", {})
    hours_remaining_cfg = paths_cfg.get("hours_remaining", {})
    if not isinstance(hours_remaining_cfg, dict):
        hours_remaining_cfg = {}

    REPORT_TITLE = str(CONFIG.get("report_title", "Project Portfolio Overview") or "").strip() or "Project Portfolio Overview"
    PERSON_NAME = str(CONFIG.get("person_name", "john doe") or "").strip() or "john doe"
    COMPANY_NAME = str(company_cfg.get("name", "NOC*NSF") or "").strip() or "NOC*NSF"
    COMPANY_ABBREVIATION = str(company_cfg.get("abbreviation", "NN") or "").strip() or "NN"
    PROJECTEN_DIR = resolve_path(CONFIG, paths_cfg.get("projects_dir", "Projecten"))
    DUMMY_PROJECTEN_DIR = resolve_path(CONFIG, paths_cfg.get("dummy_projects_dir", "DummyProjecten"))
    REPORT_DIR = resolve_path(CONFIG, paths_cfg.get("reports_dir", "Reports"))
    REPORTS_ARCHIVE_DIR = os.path.join(REPORT_DIR, "Archive")
    ASSETS_DIR = resolve_path(CONFIG, paths_cfg.get("assets_dir", "assets"))
    PROFILE_PHOTO_PATH = resolve_path(
        CONFIG,
        branding_cfg.get("profile_photo", "assets/profile_photo.jpg"),
    )
    TEAMNL_LOGO_PATH = resolve_path(
        CONFIG,
        branding_cfg.get("logo", "assets/logo.png"),
    )

    paths_from_nested = resolve_path_list(CONFIG, hours_remaining_cfg.get("excel_paths", []))
    if paths_from_nested:
        HOURS_REMAINING_EXCEL_PATHS = paths_from_nested
    else:
        HOURS_REMAINING_EXCEL_PATHS = resolve_path_list(CONFIG, paths_cfg.get("hours_remaining_excel_paths", []))
    if not HOURS_REMAINING_EXCEL_PATHS:
        HOURS_REMAINING_EXCEL_PATHS = resolve_path_list(
            CONFIG,
            ["Data/hours_remaining.xlsx", "Data/nn_maandelijks.xlsx", "Data/NN_maandelijks.xlsx"],
        )

    HOURS_REMAINING_SHEET_NAME = (
        str(hours_remaining_cfg.get("sheet_name", "NN_maandelijks") or "").strip() or "NN_maandelijks"
    )
    try:
        HOURS_REMAINING_HEADER_ROW = max(1, int(hours_remaining_cfg.get("header_row", 2) or 2))
    except (TypeError, ValueError):
        HOURS_REMAINING_HEADER_ROW = 2

    WORKABLE_HOURS_PER_YEAR_OVERRIDE = _as_positive_float(hours_cfg.get("workable_hours_per_year"))
    WORKABLE_HOURS_PER_WEEK_REFERENCE_VALUE = _as_positive_float(hours_cfg.get("workable_hours_per_week_reference_value"))

    # Canonical keys: base_one, base_2..base_6.
    # Legacy keys remain supported as fallback for older configs.
    BASE_BLUE = _coerce_cfg_color("base_one", "#01378A", legacy_key="base_blue")
    BASE_RED = _coerce_cfg_color("base_2", "#E1011A", legacy_key="base_red")
    BASE_ORANGE = _coerce_cfg_color("base_3", "#EA6D08", legacy_key="base_orange")
    BASE_YELLOW = _coerce_cfg_color("base_4", "#F4C300", legacy_key="base_yellow")
    BASE_GREEN = _coerce_cfg_color("base_5", "#009F3D", legacy_key="base_green")
    BASE_BLACK = _coerce_cfg_color("base_6", "#111111", legacy_key="base_black")

    TEAMNL_BASE_COLORS = [BASE_BLUE, BASE_RED, BASE_ORANGE, BASE_YELLOW, BASE_GREEN, BASE_BLACK]

    YEAR_PLAN_COMPLETED_COLOR = _coerce_hex(color_cfg.get("year_plan_completed"), BASE_BLUE)
    YEAR_PLAN_CURRENT_BILLED_COLOR = _coerce_hex(color_cfg.get("year_plan_current_billed"), BASE_BLUE)
    YEAR_PLAN_CURRENT_COMBINED_COLOR = _coerce_hex(color_cfg.get("year_plan_current_combined"), BASE_BLUE)
    YEAR_PLAN_CURRENT_EXPECTED_COLOR = _coerce_hex(color_cfg.get("year_plan_current_expected"), BASE_RED)
    YEAR_PLAN_EXPECTED_COLOR = _coerce_hex(color_cfg.get("year_plan_expected"), BASE_YELLOW)

    use_dummy = bool(runtime_cfg.get("use_dummy_projects_when_projects_empty", True))
    USING_DUMMY_FALLBACK = False
    if use_dummy and not has_subfolders(PROJECTEN_DIR):
        print(f'WARNING: No project folders found in "{PROJECTEN_DIR}". Using "{DUMMY_PROJECTEN_DIR}" for testing purposes.')
        PROJECTEN_DIR = DUMMY_PROJECTEN_DIR
        USING_DUMMY_FALLBACK = True


_apply_runtime_config()


# ----------------------------
# Warnings configuration
# ----------------------------
warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    message="Data Validation extension is not supported and will be removed",
)
# warnings.filterwarnings(
#     "ignore",
#     category=FutureWarning,
#     message="The behavior of array concatenation with empty entries is deprecated",
# )


# ----------------------------
# Color defaults
# ----------------------------
BASE_BLUE = BASE_BLUE if "BASE_BLUE" in globals() else "#01378A"
BASE_RED = BASE_RED if "BASE_RED" in globals() else "#E1011A"
BASE_ORANGE = BASE_ORANGE if "BASE_ORANGE" in globals() else "#EA6D08"
BASE_YELLOW = BASE_YELLOW if "BASE_YELLOW" in globals() else "#F4C300"
BASE_GREEN = BASE_GREEN if "BASE_GREEN" in globals() else "#009F3D"
BASE_BLACK = BASE_BLACK if "BASE_BLACK" in globals() else "#111111"
TEAMNL_BASE_COLORS = TEAMNL_BASE_COLORS if "TEAMNL_BASE_COLORS" in globals() else [
    BASE_BLUE,
    BASE_RED,
    BASE_ORANGE,
    BASE_YELLOW,
    BASE_GREEN,
    BASE_BLACK,
]
YEAR_PLAN_COMPLETED_COLOR = YEAR_PLAN_COMPLETED_COLOR if "YEAR_PLAN_COMPLETED_COLOR" in globals() else BASE_BLUE
YEAR_PLAN_CURRENT_BILLED_COLOR = YEAR_PLAN_CURRENT_BILLED_COLOR if "YEAR_PLAN_CURRENT_BILLED_COLOR" in globals() else BASE_BLUE
YEAR_PLAN_CURRENT_COMBINED_COLOR = YEAR_PLAN_CURRENT_COMBINED_COLOR if "YEAR_PLAN_CURRENT_COMBINED_COLOR" in globals() else BASE_BLUE
YEAR_PLAN_CURRENT_EXPECTED_COLOR = YEAR_PLAN_CURRENT_EXPECTED_COLOR if "YEAR_PLAN_CURRENT_EXPECTED_COLOR" in globals() else BASE_RED
YEAR_PLAN_EXPECTED_COLOR = YEAR_PLAN_EXPECTED_COLOR if "YEAR_PLAN_EXPECTED_COLOR" in globals() else BASE_YELLOW

SHADE_STEPS = [0.0, -0.25, 0.25, -0.50, 0.5, 0.75, -0.75]


# ----------------------------
# Color helpers
# ----------------------------
def _clamp_channel(value: float) -> int:
    return int(max(0, min(255, round(value))))


def adjust_color_luminance(hex_color: str, factor: float) -> str:
    """
    Lightens (factor>0) or darkens (factor<0) a hex color by the given factor.
    """
    color = hex_color.lstrip("#")
    r, g, b = int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)

    if factor >= 0:
        r = _clamp_channel(r + (255 - r) * factor)
        g = _clamp_channel(g + (255 - g) * factor)
        b = _clamp_channel(b + (255 - b) * factor)
    else:
        r = _clamp_channel(r * (1 + factor))
        g = _clamp_channel(g * (1 + factor))
        b = _clamp_channel(b * (1 + factor))

    return f"#{r:02X}{g:02X}{b:02X}"


def hex_to_rgba(hex_color: str, alpha: float) -> str:
    color = hex_color.lstrip("#")
    if len(color) != 6:
        return f"rgba(0,0,0,{alpha})"
    r, g, b = int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)
    return f"rgba({r},{g},{b},{alpha})"


def _hex_to_rgb(hex_color: str) -> Optional[Tuple[int, int, int]]:
    if not isinstance(hex_color, str):
        return None
    color = hex_color.strip().lstrip("#")
    if len(color) != 6:
        return None
    try:
        r, g, b = int(color[0:2], 16), int(color[2:4], 16), int(color[4:6], 16)
    except ValueError:
        return None
    return r, g, b


def _rgb_to_hex(rgb: Tuple[int, int, int]) -> str:
    r, g, b = rgb
    return f"#{_clamp_channel(r):02X}{_clamp_channel(g):02X}{_clamp_channel(b):02X}"


def desaturate_hex_towards_gray(hex_color: str, blend_factor: float = 0.30, gray_rgb: Tuple[int, int, int] = (140, 140, 140)) -> str:
    rgb = _hex_to_rgb(hex_color)
    if rgb is None:
        return hex_color
    f = max(0.0, min(1.0, float(blend_factor)))
    r = _clamp_channel(rgb[0] * (1 - f) + gray_rgb[0] * f)
    g = _clamp_channel(rgb[1] * (1 - f) + gray_rgb[1] * f)
    b = _clamp_channel(rgb[2] * (1 - f) + gray_rgb[2] * f)
    return _rgb_to_hex((r, g, b))


def is_active_status(status: Any) -> bool:
    status_str = str(status).strip() if status is not None else ""
    return status_str.lower() == "active"


def marker_style_for_status(base_color_hex: str, status: Any) -> Tuple[str, float]:
    """
    Returns (marker_color_hex, opacity) based on status:
      - Active if status.lower() == "active" -> base color, opacity=1.0
      - Closed otherwise -> slightly desaturated color, opacity=0.4
    """
    base = base_color_hex or BASE_BLACK
    if is_active_status(status):
        return base, 1.0
    return desaturate_hex_towards_gray(base, blend_factor=0.30), 0.4


def build_color_maps(projects_df: pd.DataFrame) -> Tuple[Dict[str, str], Dict[str, str]]:
    """
    Returns:
      - program_color_map: programma -> base color
      - project_color_map: project_id -> shaded color (based on the programma base color)
    """
    program_color_map: Dict[str, str] = {}
    project_color_map: Dict[str, str] = {}

    programs = sorted(projects_df["programma"].fillna("Unknown").replace("", "Unknown").unique().tolist())
    for idx, programma in enumerate(programs):
        program_color_map[programma] = TEAMNL_BASE_COLORS[idx % len(TEAMNL_BASE_COLORS)]

    for programma in programs:
        base = program_color_map[programma]
        mask = projects_df["programma"].fillna("Unknown").replace("", "Unknown") == programma
        projects_in_program = projects_df.loc[mask].sort_values("project_id")

        for shade_idx, (_, project_row) in enumerate(projects_in_program.iterrows()):
            shade = SHADE_STEPS[shade_idx % len(SHADE_STEPS)]
            project_id = str(project_row.get("project_id", "")).strip()
            if project_id:
                project_color_map[project_id] = adjust_color_luminance(base, shade)

    return program_color_map, project_color_map


# ----------------------------
# Project folder naming rules
# ----------------------------
def derive_project_id_from_folder(folder_name: str) -> str:
    """
    Expected folder name format:
        YYYY_NNNN_<description>
    Example:
        2026_0001_SSC_Fysiologie_ondersteuning

    Derived project_id is:
        YYYY_NNNN
    """
    parts = folder_name.split("_")
    if len(parts) < 3:
        raise ValueError(
            f"Invalid project folder name '{folder_name}'. "
            "Expected format: YYYY_NNNN_<description>"
        )

    year, counter = parts[0], parts[1]
    if not (year.isdigit() and len(year) == 4):
        raise ValueError(f"Invalid year in project folder '{folder_name}' (expected 4 digits).")
    if not counter.isdigit():
        raise ValueError(f"Invalid counter in project folder '{folder_name}' (expected digits).")

    return f"{year}_{counter}"


def discover_project_folders(projecten_dir: str) -> List[str]:
    """Each direct subfolder under Projecten/ is treated as a project."""
    if not os.path.isdir(projecten_dir):
        raise FileNotFoundError(f"Projecten folder not found: {projecten_dir}")

    folders: List[str] = []
    for name in sorted(os.listdir(projecten_dir)):
        path = os.path.join(projecten_dir, name)
        if os.path.isdir(path):
            folders.append(path)
    return folders


# ----------------------------
# Reading project_info.xlsx (key/value)
# ----------------------------
def read_project_info_kv_from_xlsx(path: str) -> Dict[str, Any]:
    """
    Reads Excel with:
      sheet: ProjectInfo
      row 1 headers: Key | Value
      rows 2..n: key/value
    """
    df = pd.read_excel(path, sheet_name="ProjectInfo", header=0, usecols=[0, 1])
    df.columns = ["key", "value"]
    df = df.dropna(subset=["key"]).copy()

    df["key"] = df["key"].astype(str).str.strip()
    df["value"] = df["value"].apply(lambda v: v.strip() if isinstance(v, str) else v)

    return dict(zip(df["key"], df["value"]))


def parse_date(value: Any) -> Optional[pd.Timestamp]:
    """Best-effort date parsing; returns pandas Timestamp or None."""
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    if isinstance(value, str) and value.strip() == "":
        return None
    ts = pd.to_datetime(value, errors="coerce")
    if pd.isna(ts):
        return None
    return ts


def _normalize_role_token(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().casefold())


def _canonical_project_role(raw_role: Any) -> str:
    role_token = _normalize_role_token(raw_role)
    if role_token in COMPLETE_MISSING_HOURS_ROLE_TOKENS:
        return PROJECT_ROLE_COMPLETE_MISSING_HOURS
    return PROJECT_ROLE_STANDARD


def resolve_project_reporting_role(info: Dict[str, Any], folder_name: str) -> str:
    normalized_info: Dict[str, Any] = {}
    for key, value in info.items():
        key_token = _normalize_role_token(key)
        if key_token and key_token not in normalized_info:
            normalized_info[key_token] = value

    explicit_role_found = False
    for key_token in PROJECT_ROLE_KEY_TOKENS:
        if key_token in normalized_info:
            explicit_role_found = True
            explicit_role = _canonical_project_role(normalized_info.get(key_token))
            if explicit_role != PROJECT_ROLE_STANDARD:
                return explicit_role

    folder_token = _normalize_role_token(folder_name)
    project_name_token = _normalize_role_token(info.get("project_name"))
    if (
        folder_token == _normalize_role_token(DEFAULT_COMPLETE_MISSING_HOURS_FOLDER)
        or project_name_token == _normalize_role_token(DEFAULT_COMPLETE_MISSING_HOURS_FOLDER)
    ):
        return PROJECT_ROLE_COMPLETE_MISSING_HOURS

    if explicit_role_found:
        return PROJECT_ROLE_STANDARD

    return PROJECT_ROLE_STANDARD




def _clean_group_value(val: Any) -> Optional[str]:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    s = str(val).strip()
    if not s:
        return None
    return "Unknown" if s.lower() == "unknown" else s


def extract_group_values(row: pd.Series, base_col: str) -> List[str]:
    """
    Returns ordered unique values for a base column that may have numbered variants (e.g., programma, programma02).
    Skips empty entries and prefers real values over 'Unknown'.
    """
    matches: List[Tuple[int, str]] = []
    base_len = len(base_col)
    for col in row.index:
        if col == base_col:
            matches.append((1, col))
        elif col.startswith(base_col) and col[base_len:].isdigit():
            matches.append((int(col[base_len:]), col))

    values: List[str] = []
    saw_unknown = False
    for _, col in sorted(matches, key=lambda x: x[0]):
        val = _clean_group_value(row.get(col))
        if val is None:
            continue
        if val == "Unknown":
            saw_unknown = True
            continue
        if val not in values:
            values.append(val)

    if values:
        return values
    return ["Unknown"] if saw_unknown else []


def _split_pipe_values(val: Any) -> List[str]:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return []
    parts = str(val).split("|")
    return [p.strip() for p in parts if p and p.strip()]


# ----------------------------
# Reading time_log.xlsx
# ----------------------------
TIMELOG_SHEET_NAME = "TimeLog"
TIMELOG_HEADER_ROW_1BASED = 6  # row 6 contains column headers in your template

TIMELOG_COLUMNS = [
    "Date*",
    "StartTime",
    "EndTime",
    "DurationMinutes*",
    "ActivityType*",
    "WhatIDid*",
    "OutputLink",
    "NextStep",
    "Tags",
    "Location",
]


def read_time_log_entries(time_log_path: str) -> pd.DataFrame:
    """
    Reads time_log.xlsx (sheet 'TimeLog') and returns a dataframe with computed duration_minutes.

    Rules:
    - Prefer DurationMinutes as the source of truth.
    - If DurationMinutes is empty, but StartTime and EndTime are present, compute minutes.
    - Ignore fully empty rows.
    """
    # Header row is 6 -> pandas header index 5
    df = pd.read_excel(
        time_log_path,
        sheet_name=TIMELOG_SHEET_NAME,
        header=TIMELOG_HEADER_ROW_1BASED - 1,
        usecols="A:J",
        engine="openpyxl",
    )

    missing = [c for c in TIMELOG_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(
            f"time_log.xlsx has unexpected columns in {time_log_path}. "
            f"Missing expected: {missing}. Found: {list(df.columns)}"
        )

    df = df.dropna(how="all").copy()
    if df.empty:
        df["duration_minutes"] = pd.Series(dtype="float")
        df["date"] = pd.Series(dtype="datetime64[ns]")
        return df

    df["date"] = pd.to_datetime(df["Date*"], errors="coerce")
    df["duration_minutes"] = pd.to_numeric(df["DurationMinutes*"], errors="coerce")

    def to_minutes_from_times(start, end) -> Optional[float]:
        if pd.isna(start) or pd.isna(end):
            return None
        start_dt = pd.to_datetime(str(start), errors="coerce")
        end_dt = pd.to_datetime(str(end), errors="coerce")
        if pd.isna(start_dt) or pd.isna(end_dt):
            return None
        minutes = (end_dt - start_dt).total_seconds() / 60.0
        if minutes <= 0:
            return None
        return minutes

    needs_compute = df["duration_minutes"].isna()
    if needs_compute.any():
        computed: List[Optional[float]] = []
        for s, e, need in zip(df["StartTime"], df["EndTime"], needs_compute):
            computed.append(to_minutes_from_times(s, e) if need else None)
        df.loc[needs_compute, "duration_minutes"] = computed

    # Keep rows with at least a date or a duration
    df = df[(~df["date"].isna()) | (~df["duration_minutes"].isna())].copy()
    return df


def read_time_log_project_metadata(time_log_path: str) -> Dict[str, str]:
    """
    Reads metadata at the top of the TimeLog sheet:
      B1: Project ID
      B2: Project Name
      B3: Programma
    """
    meta = {"project_id": "", "project_name": "", "programma": ""}
    try:
        raw = pd.read_excel(time_log_path, sheet_name=TIMELOG_SHEET_NAME, header=None, nrows=3, usecols="A:B")
        meta["project_id"] = str(raw.iloc[0, 1]).strip() if raw.shape[0] > 0 and not pd.isna(raw.iloc[0, 1]) else ""
        meta["project_name"] = str(raw.iloc[1, 1]).strip() if raw.shape[0] > 1 and not pd.isna(raw.iloc[1, 1]) else ""
        meta["programma"] = str(raw.iloc[2, 1]).strip() if raw.shape[0] > 2 and not pd.isna(raw.iloc[2, 1]) else ""
    except Exception:
        raise ValueError(f"Failed to read time_log.xlsx metadata from {time_log_path}")
    return meta


# ----------------------------
# Human-readable record objects
# ----------------------------
@dataclass
class ProjectRecord:
    folder_name: str
    folder_path: str
    project_id: str
    project_info_path: str
    time_log_path: str
    info: Dict[str, Any]


# ----------------------------
# Load + validate projects
# ----------------------------


# ----------------------------
# Hover helper
# ----------------------------
HOVER_KEYS = [
    "project_id",
    "project_name",
    "programma(s)",
    "requester(s)",
    "owner",
    "status",
    "priority",
    "theme(s)",
    "start_date",
    "target_end_date",
    "actual_end_date",
]


def normalize_status_for_display(val: Any) -> str:
    s = str(val).strip() if val is not None else ""
    return s if s else "Unknown"


def format_date_yyyymmdd(val: Any) -> Optional[str]:
    ts = parse_date(val)
    if ts is None:
        return None
    return ts.date().isoformat()


def resolve_end_date_for_hover(row: pd.Series) -> str:
    actual = format_date_yyyymmdd(row.get("actual_end_date"))
    if actual:
        return actual
    target = format_date_yyyymmdd(row.get("target_end_date"))
    if target:
        return target
    start = parse_date(row.get("start_date"))
    if start is not None:
        return f"{start.year:04d}-12-31"
    return "(no end date)"


def build_hover_text(project_row: pd.Series, extra: Optional[Dict[str, Any]] = None) -> str:
    extra = dict(extra) if extra else {}

    status_display = normalize_status_for_display(project_row.get("status") if "status" in project_row else None)
    resolved_end_date = extra.pop("resolved_end_date", None)
    if resolved_end_date is None or str(resolved_end_date).strip() == "":
        resolved_end_date = resolve_end_date_for_hover(project_row)

    parts = [f"<b>Status</b>: {status_display}", f"<b>resolved_end_date</b>: {resolved_end_date}"]

    for k in HOVER_KEYS:
        if k == "status":
            continue
        if k in project_row and pd.notna(project_row[k]) and str(project_row[k]).strip() != "":
            val = project_row[k]
            if k in ("start_date", "target_end_date", "actual_end_date"):
                formatted = format_date_yyyymmdd(val)
                if formatted is not None:
                    val = formatted
            parts.append(f"<b>{k}</b>: {val}")
    if "__project_folder" in project_row and pd.notna(project_row["__project_folder"]):
        parts.append(f"<b>folder</b>: {project_row['__project_folder']}")
    if extra:
        for k, v in extra.items():
            if v is not None and str(v).strip() != "":
                parts.append(f"<b>{k}</b>: {v}")
    return "<br>".join(parts)


# ----------------------------
# Period helpers
# ----------------------------
def _last_completed_month(asof_date: date) -> Tuple[date, date, str]:
    first_of_month = date(asof_date.year, asof_date.month, 1)
    last_day_prev = first_of_month - timedelta(days=1)
    period_start = date(last_day_prev.year, last_day_prev.month, 1)
    period_end = last_day_prev
    period_key = f"{period_end.year:04d}-{period_end.month:02d}"
    return period_start, period_end, period_key


def _last_completed_iso_week(asof_date: date) -> Tuple[date, date, str]:
    last_sunday = asof_date - timedelta(days=asof_date.weekday() + 1)
    period_start = last_sunday - timedelta(days=6)
    period_end = last_sunday
    iso = period_end.isocalendar()
    period_key = f"{iso.year}-W{iso.week:02d}"
    return period_start, period_end, period_key


def _last_completed_biweekly(asof_date: date) -> Tuple[date, date, str]:
    # Two full ISO weeks (Mon-Sun), ending on the last completed Sunday.
    last_sunday = asof_date - timedelta(days=asof_date.weekday() + 1)
    period_end = last_sunday
    period_start = period_end - timedelta(days=13)
    iso_start = period_start.isocalendar()
    iso_end = period_end.isocalendar()
    period_key = f"{iso_start.year}-W{iso_start.week:02d}_to_{iso_end.year}-W{iso_end.week:02d}"
    return period_start, period_end, period_key


def compute_report_periods(asof_date: date) -> Dict[str, Dict[str, Any]]:
    daily_start = asof_date
    daily_end = asof_date
    day_key = asof_date.isoformat()
    monthly_start = date(asof_date.year, asof_date.month, 1)
    monthly_end = asof_date
    month_key = f"{asof_date.year:04d}-{asof_date.month:02d}"
    weekly_start, weekly_end, week_key = _last_completed_iso_week(asof_date)
    biweekly_start, biweekly_end, biweek_key = _last_completed_biweekly(asof_date)
    yearly_start = date(asof_date.year, 1, 1)
    yearly_end = asof_date
    year_key = f"{asof_date.year:04d}"
    return {
        "daily": dict(label="1-day", start=daily_start, end=daily_end, key=day_key),
        "weekly": dict(label="1-week", start=weekly_start, end=weekly_end, key=week_key),
        "biweekly": dict(label="2-weeks", start=biweekly_start, end=biweekly_end, key=biweek_key),
        "monthly": dict(label="Month (to-date)", start=monthly_start, end=monthly_end, key=month_key),
        "yearly": dict(label="Year (to-date)", start=yearly_start, end=yearly_end, key=year_key),
    }


_MONTH_ABBR_LOWER = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]


def _fmt_day_month_lower(d: date) -> str:
    return f"{int(d.day)}-{_MONTH_ABBR_LOWER[d.month - 1]}"


def format_period_range_compact(period_start: date, period_end: date) -> str:
    if period_start == period_end:
        return f"{_fmt_day_month_lower(period_start)} ({period_start.year:04d})"
    if period_start.year == period_end.year:
        return (
            f"{_fmt_day_month_lower(period_start)} to {_fmt_day_month_lower(period_end)} "
            f"({period_start.year:04d})"
        )
    return (
        f"{_fmt_day_month_lower(period_start)} ({period_start.year:04d}) "
        f"to {_fmt_day_month_lower(period_end)} ({period_end.year:04d})"
    )


def list_completed_month_periods(asof_date: date, time_entries_df: Optional[pd.DataFrame] = None) -> List[Dict[str, Any]]:
    """
    Returns month periods (start/end/key/label), newest-first, including the current month through `asof_date`.

    If time_entries_df is provided, the earliest month is derived from the earliest available time entry date;
    otherwise only the current month is returned.
    """
    current_start = date(asof_date.year, asof_date.month, 1)
    start_month = current_start

    if time_entries_df is not None and not time_entries_df.empty and "date" in time_entries_df.columns:
        dates = pd.to_datetime(time_entries_df["date"], errors="coerce").dropna()
        if not dates.empty:
            min_date = dates.min().date()
            if min_date <= asof_date:
                start_month = date(min_date.year, min_date.month, 1)

    month_starts = pd.date_range(pd.Timestamp(start_month), pd.Timestamp(current_start), freq="MS")
    periods: List[Dict[str, Any]] = []
    for month_start in month_starts:
        ms = month_start.date()
        month_end = (month_start + pd.offsets.MonthEnd(0)).date()
        is_current_month = (ms.year == asof_date.year) and (ms.month == asof_date.month)
        me = asof_date if is_current_month else month_end
        key = f"{ms.year:04d}-{ms.month:02d}"
        label = month_start.strftime("%b %Y")
        if is_current_month and asof_date < month_end:
            label += " (to-date)"
        periods.append(dict(start=ms, end=me, key=key, label=label))

    periods.sort(key=lambda p: p["start"], reverse=True)
    return periods


def list_available_day_periods(asof_date: date, time_entries_df: Optional[pd.DataFrame] = None) -> List[Dict[str, Any]]:
    """
    Returns day periods (start/end/key/label), newest-first, from available time-entry dates up to `asof_date`.

    If there are no valid entry dates, falls back to a single day period for `asof_date`.
    """
    fallback = dict(start=asof_date, end=asof_date, key=asof_date.isoformat(), label=asof_date.strftime("%a %d %b %Y"))
    if time_entries_df is None or time_entries_df.empty or "date" not in time_entries_df.columns:
        return [fallback]

    dates = pd.to_datetime(time_entries_df["date"], errors="coerce").dropna()
    if dates.empty:
        return [fallback]

    available = dates.dt.normalize()
    available = available[available <= pd.Timestamp(asof_date)]
    if available.empty:
        return [fallback]

    unique_days = sorted({d.date() for d in available}, reverse=True)
    periods: List[Dict[str, Any]] = []
    for day_val in unique_days:
        periods.append(
            dict(
                start=day_val,
                end=day_val,
                key=day_val.isoformat(),
                label=day_val.strftime("%a %d %b %Y"),
            )
        )
    return periods


def filter_time_entries_by_period(time_entries_df: pd.DataFrame, period_start: date, period_end: date) -> pd.DataFrame:
    if time_entries_df.empty or "date" not in time_entries_df.columns:
        return time_entries_df.copy()
    mask = (time_entries_df["date"] >= pd.Timestamp(period_start)) & (time_entries_df["date"] <= pd.Timestamp(period_end))
    return time_entries_df.loc[mask].copy()


def filter_projects_with_hours(
    projects_df: pd.DataFrame, time_entries_df_filtered: pd.DataFrame
) -> pd.DataFrame:
    if time_entries_df_filtered.empty or "project_id" not in time_entries_df_filtered.columns:
        return projects_df.iloc[0:0].copy()
    hours_by_project = (
        time_entries_df_filtered.groupby("project_id")["duration_hours"]
        .sum(min_count=1)
        .fillna(0.0)
    )
    valid_ids = {str(pid) for pid, hours in hours_by_project.items() if hours > 0}
    if not valid_ids:
        return projects_df.iloc[0:0].copy()
    mask = projects_df["project_id"].astype(str).isin(valid_ids)
    return projects_df.loc[mask].copy()


def collect_project_ids_by_role(projects_df: pd.DataFrame, role: str) -> Set[str]:
    if (
        projects_df is None
        or projects_df.empty
        or "project_id" not in projects_df.columns
        or "__reporting_role" not in projects_df.columns
    ):
        return set()
    role_token = str(role).strip().casefold()
    matches = projects_df.loc[
        projects_df["__reporting_role"].fillna("").astype(str).str.strip().str.casefold() == role_token,
        "project_id",
    ]
    return {str(pid).strip() for pid in matches if str(pid).strip()}


def filter_projects_excluding_project_ids(
    projects_df: pd.DataFrame, excluded_project_ids: Optional[Set[str]]
) -> pd.DataFrame:
    if projects_df is None:
        return pd.DataFrame()
    excluded = {str(pid).strip() for pid in (excluded_project_ids or set()) if str(pid).strip()}
    if not excluded or "project_id" not in projects_df.columns:
        return projects_df.copy()
    mask = ~projects_df["project_id"].astype(str).str.strip().isin(excluded)
    return projects_df.loc[mask].copy()


def filter_time_entries_excluding_project_ids(
    time_entries_df: pd.DataFrame, excluded_project_ids: Optional[Set[str]]
) -> pd.DataFrame:
    if time_entries_df is None:
        return pd.DataFrame()
    excluded = {str(pid).strip() for pid in (excluded_project_ids or set()) if str(pid).strip()}
    if not excluded or "project_id" not in time_entries_df.columns:
        return time_entries_df.copy()
    mask = ~time_entries_df["project_id"].astype(str).str.strip().isin(excluded)
    return time_entries_df.loc[mask].copy()


def build_year_week_grid(target_year: int) -> Tuple[pd.DatetimeIndex, pd.DatetimeIndex, pd.DatetimeIndex, float]:
    year_start = date(target_year, 1, 1)
    year_end = date(target_year, 12, 31)
    week_starts = pd.date_range(pd.Timestamp(year_start), pd.Timestamp(year_end), freq="W-MON")
    week_ends = week_starts + pd.Timedelta(days=6)
    bar_width = pd.Timedelta(days=7)
    half_bar = bar_width / 2
    week_positions = week_starts + half_bar
    bar_width_ms = bar_width / pd.Timedelta(milliseconds=1)
    return week_starts, week_ends, week_positions, bar_width_ms


def estimate_magnitude_weight(value: Any) -> int:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return 1
    if isinstance(value, (int, float)):
        hours = float(value)
    else:
        s = str(value).strip().lower()
        if not s:
            return 1
        if any(k in s for k in ["small", "klein"]):
            return 1
        if "medium" in s:
            return 3
        if "substantial" in s:
            return 8
        if "large" in s or "groot" in s:
            return 16
        if "enormous" in s or "enorm" in s:
            return 30
        try:
            hours = float(re.sub(r"[^\d.]+", "", s))
        except ValueError:
            return 1

    if hours < 6:
        return 1
    if hours < 24:
        return 3
    if hours < 160:
        return 8
    if hours < 320:
        return 16
    return 30


def build_period_week_grid(period_start: date, period_end: date) -> Tuple[pd.DatetimeIndex, pd.DatetimeIndex, pd.DatetimeIndex, float]:
    start_week = period_start - timedelta(days=period_start.weekday())
    end_week = period_end - timedelta(days=period_end.weekday())
    week_starts = pd.date_range(pd.Timestamp(start_week), pd.Timestamp(end_week), freq="W-MON")
    week_ends = week_starts + pd.Timedelta(days=6)
    bar_width = pd.Timedelta(days=7)
    half_bar = bar_width / 2
    week_positions = week_starts + half_bar
    bar_width_ms = bar_width / pd.Timedelta(milliseconds=1)
    return week_starts, week_ends, week_positions, bar_width_ms


# ----------------------------
# Yearly-capacity helpers
# ----------------------------
_MONTH_MAP = {
    "jan": 1,
    "feb": 2,
    "mar": 3,
    "apr": 4,
    "may": 5,
    "jun": 6,
    "jul": 7,
    "aug": 8,
    "sep": 9,
    "oct": 10,
    "nov": 11,
    "dec": 12,
}


def _clean_colname(name: str) -> str:
    return re.sub(r"\s+", " ", str(name)).strip().lower()


def _find_col(cols: List[str], include_tokens: List[str], exclude_tokens: Optional[List[str]] = None) -> Optional[str]:
    exclude_tokens = exclude_tokens or []
    for col in cols:
        norm = _clean_colname(col)
        if all(tok in norm for tok in include_tokens) and not any(tok in norm for tok in exclude_tokens):
            return col
    return None


def _to_float(val: Any) -> Optional[float]:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _to_float_relaxed(val: Any) -> Optional[float]:
    """Parse numeric values from mixed strings (e.g. '1,200', '1200 h', '1.200,5')."""
    direct = _to_float(val)
    if direct is not None:
        return direct
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None

    # Keep only likely numeric characters for locale-aware cleanup.
    s = re.sub(r"[^0-9,.\-]+", "", s)
    if not s:
        return None
    if "," in s and "." in s:
        # If comma appears after dot, assume EU decimal style.
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _priority_rank(priority: Any) -> int:
    p = str(priority or "").strip().lower()
    mapping = {"low": 1, "medium": 2, "high": 3, "critical": 4}
    return mapping.get(p, 0)


def _get_nl_holiday_dates(year: int) -> List[date]:
    """Return Dutch public holiday dates for `year` when holidays package is available."""
    if holidays_lib is None:
        return []
    try:
        holiday_obj = holidays_lib.country_holidays("NL", years=[year])
    except Exception:
        try:
            holiday_obj = holidays_lib.NL(years=[year])
        except Exception:
            return []

    holiday_dates: List[date] = []
    for h in holiday_obj.keys():
        if isinstance(h, datetime):
            holiday_dates.append(h.date())
        elif isinstance(h, pd.Timestamp):
            holiday_dates.append(h.date())
        elif isinstance(h, date):
            holiday_dates.append(h)
        else:
            ts = pd.to_datetime(h, errors="coerce")
            if not pd.isna(ts):
                holiday_dates.append(ts.date())
    return sorted(set(holiday_dates))


def _count_workdays_inclusive(start_day: date, end_day: date, holiday_dates: Optional[List[date]] = None) -> float:
    """Count Mon-Fri days in [start_day, end_day], excluding `holiday_dates` when provided."""
    if start_day is None or end_day is None or start_day > end_day:
        return 0.0
    days = pd.date_range(pd.Timestamp(start_day), pd.Timestamp(end_day), freq="D")
    if days.empty:
        return 0.0
    is_workday = days.weekday < 5
    if holiday_dates:
        holiday_idx = pd.to_datetime(pd.Series(holiday_dates, dtype="datetime64[ns]"), errors="coerce").dropna().dt.normalize()
        if not holiday_idx.empty:
            is_holiday = days.normalize().isin(set(holiday_idx.tolist()))
            is_workday = is_workday & ~is_holiday
    return float(is_workday.sum())


def _parse_month_label(val: Any) -> Optional[pd.Timestamp]:
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, (pd.Timestamp, datetime, date)):
        return pd.Timestamp(val.year, val.month, 1)
    s = str(val).strip()
    if not s:
        return None
    ts = pd.to_datetime(s, errors="coerce")
    if not pd.isna(ts):
        return pd.Timestamp(ts.year, ts.month, 1)
    cleaned = re.sub(r"[^A-Za-z0-9]", "", s).lower()
    m = re.match(r"([a-z]{3,9})(\d{4})", cleaned)
    if m and m.group(1)[:3] in _MONTH_MAP:
        return pd.Timestamp(int(m.group(2)), _MONTH_MAP[m.group(1)[:3]], 1)
    m = re.match(r"(\d{4})(\d{1,2})", cleaned)
    if m:
        year = int(m.group(1))
        month = int(m.group(2))
        if 1 <= month <= 12:
            return pd.Timestamp(year, month, 1)
    return None


def _company_label_short() -> str:
    return COMPANY_ABBREVIATION or COMPANY_NAME or "Company"


def _company_label_long() -> str:
    return COMPANY_NAME or COMPANY_ABBREVIATION or "Company"


def _build_hours_remaining_df_from_config(
    target_year: int,
    all_time_entries_df: Optional[pd.DataFrame],
) -> pd.DataFrame:
    """
    Build a synthetic hours-remaining table from config.hours.workable_hours_per_year.
    This replaces external billing workbooks when configured.
    """
    if not WORKABLE_HOURS_PER_YEAR_OVERRIDE or WORKABLE_HOURS_PER_YEAR_OVERRIDE <= 0:
        return pd.DataFrame()

    annual_hours = float(WORKABLE_HOURS_PER_YEAR_OVERRIDE)
    month_starts = pd.date_range(pd.Timestamp(date(target_year, 1, 1)), pd.Timestamp(date(target_year, 12, 1)), freq="MS")
    nl_holidays = _get_nl_holiday_dates(target_year)

    month_workdays: List[float] = []
    for month_ts in month_starts:
        month_start = date(month_ts.year, month_ts.month, 1)
        month_end = (month_ts + pd.offsets.MonthEnd(0)).date()
        month_workdays.append(_count_workdays_inclusive(month_start, month_end, nl_holidays))
    total_workdays = float(sum(month_workdays))
    if total_workdays <= 0:
        total_workdays = float(len(month_starts))  # fallback equal split
        month_workdays = [1.0] * len(month_starts)

    billed_by_month: Dict[pd.Timestamp, float] = {}
    if all_time_entries_df is not None and not all_time_entries_df.empty and {"date", "duration_hours"}.issubset(all_time_entries_df.columns):
        entries = all_time_entries_df.copy()
        entries["date"] = pd.to_datetime(entries["date"], errors="coerce")
        entries["duration_hours"] = pd.to_numeric(entries["duration_hours"], errors="coerce").fillna(0.0)
        entries = entries.dropna(subset=["date"])
        entries = entries.loc[entries["date"].dt.year == target_year]
        if not entries.empty:
            entries["__month"] = entries["date"].dt.to_period("M").dt.to_timestamp()
            grouped = entries.groupby("__month")["duration_hours"].sum(min_count=1).fillna(0.0)
            billed_by_month = {pd.Timestamp(m): max(float(h), 0.0) for m, h in grouped.items()}

    rows: List[Dict[str, Any]] = []
    cumulative = 0.0
    for month_ts, month_workdays_val in zip(month_starts, month_workdays):
        monthly_workable_hours = annual_hours * (month_workdays_val / total_workdays)
        month_billed = billed_by_month.get(pd.Timestamp(month_ts), 0.0)
        cumulative += month_billed
        remaining = max(annual_hours - cumulative, 0.0)
        rows.append(
            {
                "Tabblad": month_ts,
                "cumulatief": cumulative,
                "uren per maand": month_billed,
                "resterend": remaining,
                "werkbare uren per maand": monthly_workable_hours,
                "werkbare dagen per maand": month_workdays_val,
            }
        )
    return pd.DataFrame(rows)


def _find_nn_maandelijks_path() -> Optional[str]:
    for path in HOURS_REMAINING_EXCEL_PATHS:
        if os.path.exists(path):
            return path
    return None


def load_nn_maandelijks_df(
    asof_date: Optional[date] = None,
    all_time_entries_df: Optional[pd.DataFrame] = None,
) -> Tuple[Optional[pd.DataFrame], Optional[str], str]:
    target_year = (asof_date or date.today()).year
    if WORKABLE_HOURS_PER_YEAR_OVERRIDE and WORKABLE_HOURS_PER_YEAR_OVERRIDE > 0:
        df = _build_hours_remaining_df_from_config(target_year, all_time_entries_df)
        if df.empty:
            return None, None, "Config yearly workable-hours is set, but synthetic hours data could not be built."
        return (
            df,
            None,
            (
                "Using config.hours.workable_hours_per_year="
                f"{WORKABLE_HOURS_PER_YEAR_OVERRIDE:.0f}h as yearly capacity source."
            ),
        )

    path = _find_nn_maandelijks_path()
    if not path:
        return None, None, "Hours-remaining source file not found; skipping yearly-capacity summaries."
    try:
        header_idx = max(HOURS_REMAINING_HEADER_ROW - 1, 0)
        df = pd.read_excel(path, sheet_name=HOURS_REMAINING_SHEET_NAME, header=header_idx)
    except Exception as exc:
        return None, path, f"Failed to read hours-remaining source sheet '{HOURS_REMAINING_SHEET_NAME}': {exc}"
    return df, path, f"Hours-remaining source loaded from {path} (sheet: {HOURS_REMAINING_SHEET_NAME})."


def compute_nn_summary(
    nn_df: Optional[pd.DataFrame],
    period_type: str,
    period_end: date,
    time_entries_df_filtered: pd.DataFrame,
    all_time_entries_df: Optional[pd.DataFrame] = None,
    asof_date: Optional[date] = None,
) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    if nn_df is None or nn_df.empty:
        return None, "Yearly-capacity data not available."

    df = nn_df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    if df.empty:
        return None, "Hours-remaining sheet is empty."

    month_col = "Tabblad"
    if month_col not in df.columns:
        month_col = df.columns[0]

    df["__month"] = df[month_col].apply(_parse_month_label)
    df["__month"] = pd.to_datetime(df["__month"], errors="coerce")
    df = df.dropna(subset=["__month"]).copy()
    if df.empty:
        return None, "No month rows found in the hours-remaining sheet."

    df = df.sort_values("__month").drop_duplicates(subset=["__month"], keep="last")
    target_year = period_end.year
    target_month = pd.Timestamp(period_end.year, period_end.month, 1)
    reference_date = asof_date or period_end
    reference_month = pd.Timestamp(reference_date.year, reference_date.month, 1)

    year_rows = df.loc[df["__month"].dt.year == target_year].copy()
    if year_rows.empty:
        return None, f"No hours-remaining rows found for {target_year}."
    year_rows = year_rows.sort_values("__month").drop_duplicates(subset=["__month"], keep="last").copy()

    cols = list(year_rows.columns)
    billed_year_col = _find_col(cols, ["cumulatief"])
    billed_month_col = _find_col(cols, ["uren per maand"])
    remaining_col = _find_col(cols, ["resterend"])
    workable_col = _find_col(cols, ["werkbare", "uren", "per", "maand"]) or _find_col(cols, ["werkbare", "uren"])
    workable_days_col = _find_col(cols, ["werkbare", "dagen", "per", "maand"]) or _find_col(cols, ["werkbare", "dagen"])

    if billed_month_col:
        year_rows["__billed_month"] = pd.to_numeric(year_rows[billed_month_col], errors="coerce")
    elif billed_year_col:
        cumulative = pd.to_numeric(year_rows[billed_year_col], errors="coerce")
        monthly_delta = cumulative.diff()
        if not monthly_delta.empty:
            monthly_delta.iloc[0] = cumulative.iloc[0]
        year_rows["__billed_month"] = monthly_delta
    else:
        year_rows["__billed_month"] = pd.Series([float("nan")] * len(year_rows), index=year_rows.index, dtype="float64")

    year_rows["__workable"] = (
        pd.to_numeric(year_rows[workable_col], errors="coerce").fillna(0.0) if workable_col else 0.0
    )
    if workable_days_col:
        year_rows["__workable_days"] = pd.to_numeric(year_rows[workable_days_col], errors="coerce").fillna(0.0)
    else:
        # Fallback: approximate workdays from workable hours with 8h/day.
        year_rows["__workable_days"] = pd.to_numeric(year_rows["__workable"], errors="coerce").fillna(0.0) / 8.0
    year_rows["__billed_month"] = pd.to_numeric(year_rows["__billed_month"], errors="coerce")
    year_rows["__billed_month"] = year_rows["__billed_month"].clip(lower=0.0)
    year_rows["__workable"] = pd.to_numeric(year_rows["__workable"], errors="coerce").fillna(0.0).clip(lower=0.0)
    year_rows["__workable_days"] = pd.to_numeric(year_rows["__workable_days"], errors="coerce").fillna(0.0).clip(lower=0.0)

    asof_row_df = year_rows.loc[year_rows["__month"] == target_month]
    if asof_row_df.empty:
        if period_type == "monthly":
            return None, f"No hours-remaining row found for {target_month.date().isoformat()}."
        asof_row_df = year_rows.loc[year_rows["__month"] <= target_month]
        if asof_row_df.empty:
            return None, f"No hours-remaining rows found for {target_year} up to {target_month.date().isoformat()}."

    asof_row = asof_row_df.iloc[-1]
    asof_month = pd.Timestamp(asof_row["__month"])
    asof_month_end = (asof_month + pd.offsets.MonthEnd(0)).date()
    ongoing_month_split = period_end < asof_month_end
    reference_row_df = year_rows.loc[year_rows["__month"] == reference_month]
    reference_row = reference_row_df.iloc[-1] if not reference_row_df.empty else None

    nl_holidays = _get_nl_holiday_dates(target_year)
    holidays_excluded = bool(nl_holidays)
    year_calendar_workdays = _count_workdays_inclusive(date(target_year, 1, 1), date(target_year, 12, 31), nl_holidays)
    month_calendar_workdays: Dict[pd.Timestamp, float] = {}
    for month_ts in year_rows["__month"]:
        month_ts = pd.Timestamp(month_ts)
        month_start_date = date(month_ts.year, month_ts.month, 1)
        month_end_date = (month_ts + pd.offsets.MonthEnd(0)).date()
        month_calendar_workdays[month_ts] = _count_workdays_inclusive(month_start_date, month_end_date, nl_holidays)

    billed = None
    if period_type == "monthly":
        billed_month_val = _to_float(asof_row.get("__billed_month"))
        if billed_month_val is not None:
            billed = billed_month_val
        elif billed_year_col:
            billed = _to_float(asof_row.get(billed_year_col))
        scope_label = "month-to-date" if ongoing_month_split else "month"
    else:
        billed_to_date_rows = year_rows.loc[year_rows["__month"] <= asof_month]
        billed_to_date_sum = pd.to_numeric(billed_to_date_rows["__billed_month"], errors="coerce").sum(min_count=1)
        if not pd.isna(billed_to_date_sum):
            billed = float(billed_to_date_sum)
        elif billed_year_col:
            billed = _to_float(asof_row.get(billed_year_col))
        scope_label = "to-date"

    remaining = _to_float(asof_row.get(remaining_col)) if remaining_col else None
    billed_cumulative_at_asof = _to_float(asof_row.get(billed_year_col)) if billed_year_col else None

    project_logged_hours = (
        float(pd.to_numeric(time_entries_df_filtered["duration_hours"], errors="coerce").fillna(0.0).sum())
        if not time_entries_df_filtered.empty and "duration_hours" in time_entries_df_filtered.columns
        else 0.0
    )
    completeness_ratio = None
    if billed is not None and billed > 0:
        completeness_ratio = project_logged_hours / billed

    remaining_for_distribution = max(float(remaining), 0.0) if remaining is not None else 0.0
    if reference_month > asof_month and reference_row is not None and remaining_col:
        reference_remaining = _to_float(reference_row.get(remaining_col))
        if reference_remaining is not None:
            remaining_for_distribution = max(float(reference_remaining), 0.0)
    year_available_hours = float(pd.to_numeric(year_rows["__workable"], errors="coerce").fillna(0.0).sum())
    year_available_workdays = float(pd.to_numeric(year_rows["__workable_days"], errors="coerce").fillna(0.0).sum())

    include_current_in_distribution = ongoing_month_split

    distribution_rows = year_rows.loc[year_rows["__month"] > asof_month].copy()
    if include_current_in_distribution:
        distribution_rows = pd.concat(
            [year_rows.loc[year_rows["__month"] == asof_month], distribution_rows],
            ignore_index=True,
        )
    if distribution_rows.empty and remaining_for_distribution > 0:
        distribution_rows = year_rows.loc[year_rows["__month"] == asof_month].copy()

    expected_alloc_by_month: Dict[pd.Timestamp, float] = {}
    if remaining_for_distribution > 0 and not distribution_rows.empty:
        distribution_rows = distribution_rows.copy()
        distribution_rows["__workable"] = pd.to_numeric(distribution_rows["__workable"], errors="coerce").fillna(0.0)
        weight_sum = float(distribution_rows["__workable"].sum())
        if weight_sum <= 0:
            distribution_rows["__weight"] = 1.0
            weight_sum = float(distribution_rows["__weight"].sum())
        else:
            distribution_rows["__weight"] = distribution_rows["__workable"]

        alloc_values = remaining_for_distribution * (distribution_rows["__weight"] / weight_sum)
        for _, dist_row in distribution_rows.iterrows():
            month_ts = pd.Timestamp(dist_row["__month"])
            alloc_val = float(alloc_values.loc[dist_row.name]) if dist_row.name in alloc_values.index else 0.0
            expected_alloc_by_month[month_ts] = max(alloc_val, 0.0)

    month_segments: List[Dict[str, Any]] = []
    month_overview_past_selection = period_type == "monthly" and reference_month > asof_month
    for _, month_row in year_rows.iterrows():
        month_ts = pd.Timestamp(month_row["__month"])
        month_num = int(month_ts.month)
        month_abbr = month_ts.strftime("%b")
        month_name = month_ts.strftime("%B %Y")
        month_start_date = date(month_ts.year, month_ts.month, 1)
        month_end_date = (month_ts + pd.offsets.MonthEnd(0)).date()
        calendar_workdays_month = float(month_calendar_workdays.get(month_ts, 0.0))
        billed_month_val = float(pd.to_numeric(month_row.get("__billed_month"), errors="coerce")) if pd.notna(month_row.get("__billed_month")) else 0.0
        workable_hours = float(pd.to_numeric(month_row.get("__workable"), errors="coerce")) if pd.notna(month_row.get("__workable")) else 0.0
        workable_days = float(pd.to_numeric(month_row.get("__workable_days"), errors="coerce")) if pd.notna(month_row.get("__workable_days")) else 0.0
        billed_month_val = max(billed_month_val, 0.0)
        workable_hours = max(workable_hours, 0.0)
        workable_days = max(workable_days, 0.0)
        expected_val = max(float(expected_alloc_by_month.get(month_ts, 0.0)), 0.0)
        billed_for_block = billed_month_val

        expected_pct = (expected_val / workable_hours * 100.0) if workable_hours > 0 else 0.0

        calendar_workdays_billed = calendar_workdays_month
        calendar_workdays_expected = 0.0
        if month_ts == asof_month and ongoing_month_split:
            billed_part_end = min(period_end, month_end_date)
            calendar_workdays_billed = _count_workdays_inclusive(month_start_date, billed_part_end, nl_holidays)
            calendar_workdays_expected = _count_workdays_inclusive(
                billed_part_end + timedelta(days=1),
                month_end_date,
                nl_holidays,
            )

        if month_ts < asof_month:
            month_segments.append(
                dict(
                    month_abbr=month_abbr,
                    month_name=month_name,
                    segment_label="Completed",
                    hours=max(billed_for_block, 0.0),
                    billed_so_far=billed_for_block,
                    workable_hours=workable_hours,
                    workable_days=workable_days,
                    calendar_workdays=calendar_workdays_month,
                    expected_remaining_hours=0.0,
                    expected_pct=0.0,
                    phase="completed",
                    month_num=month_num,
                )
            )
            continue

        if month_overview_past_selection and month_ts == asof_month:
            month_segments.append(
                dict(
                    month_abbr=month_abbr,
                    month_name=month_name,
                    segment_label="Completed",
                    hours=max(billed_for_block, 0.0),
                    billed_so_far=billed_for_block,
                    workable_hours=workable_hours,
                    workable_days=workable_days,
                    calendar_workdays=calendar_workdays_month,
                    expected_remaining_hours=0.0,
                    expected_pct=0.0,
                    phase="completed",
                    month_num=month_num,
                )
            )
            continue

        if month_ts == asof_month and ongoing_month_split:
            month_segments.append(
                dict(
                    month_abbr=month_abbr,
                    month_name=month_name,
                    segment_label="Billed so far",
                    hours=max(billed_for_block, 0.0),
                    billed_so_far=billed_for_block,
                    workable_hours=workable_hours,
                    workable_days=workable_days,
                    calendar_workdays=calendar_workdays_billed,
                    expected_remaining_hours=expected_val,
                    expected_pct=expected_pct,
                    phase="current_billed",
                    month_num=month_num,
                )
            )
            month_segments.append(
                dict(
                    month_abbr=month_abbr,
                    month_name=month_name,
                    segment_label="Expected remaining",
                    hours=expected_val,
                    billed_so_far=billed_for_block,
                    workable_hours=workable_hours,
                    workable_days=workable_days,
                    calendar_workdays=calendar_workdays_expected,
                    expected_remaining_hours=expected_val,
                    expected_pct=expected_pct,
                    phase="current_expected",
                    month_num=month_num,
                )
            )
            continue

        if month_ts == asof_month and period_type == "monthly":
            combined_hours = max(billed_for_block + expected_val, 0.0)
            month_segments.append(
                dict(
                    month_abbr=month_abbr,
                    month_name=month_name,
                    segment_label="Billed + expected",
                    hours=combined_hours,
                    billed_so_far=billed_for_block,
                    workable_hours=workable_hours,
                    workable_days=workable_days,
                    calendar_workdays=calendar_workdays_month,
                    expected_remaining_hours=expected_val,
                    expected_pct=expected_pct,
                    phase="current_combined",
                    month_num=month_num,
                )
            )
            continue

        if month_ts == asof_month:
            month_segments.append(
                dict(
                    month_abbr=month_abbr,
                    month_name=month_name,
                    segment_label="Completed",
                    hours=max(billed_for_block, 0.0),
                    billed_so_far=billed_for_block,
                    workable_hours=workable_hours,
                    workable_days=workable_days,
                    calendar_workdays=calendar_workdays_month,
                    expected_remaining_hours=0.0,
                    expected_pct=0.0,
                    phase="completed",
                    month_num=month_num,
                )
            )
            continue

        if month_overview_past_selection and month_ts > asof_month and month_ts <= reference_month:
            combined_hours = max(billed_for_block + expected_val, 0.0)
            month_segments.append(
                dict(
                    month_abbr=month_abbr,
                    month_name=month_name,
                    segment_label="Billed + expected",
                    hours=combined_hours,
                    billed_so_far=billed_for_block,
                    workable_hours=workable_hours,
                    workable_days=workable_days,
                    calendar_workdays=calendar_workdays_month,
                    expected_remaining_hours=expected_val,
                    expected_pct=expected_pct,
                    phase="current_combined",
                    month_num=month_num,
                )
            )
            continue

        if month_ts > asof_month and month_ts == reference_month:
            combined_hours = max(billed_for_block + expected_val, 0.0)
            month_segments.append(
                dict(
                    month_abbr=month_abbr,
                    month_name=month_name,
                    segment_label="Billed + expected",
                    hours=combined_hours,
                    billed_so_far=billed_for_block,
                    workable_hours=workable_hours,
                    workable_days=workable_days,
                    calendar_workdays=calendar_workdays_month,
                    expected_remaining_hours=expected_val,
                    expected_pct=expected_pct,
                    phase="current_combined",
                    month_num=month_num,
                )
            )
            continue

        month_segments.append(
            dict(
                month_abbr=month_abbr,
                month_name=month_name,
                segment_label="Expected",
                hours=expected_val,
                billed_so_far=0.0,
                workable_hours=workable_hours,
                workable_days=workable_days,
                calendar_workdays=calendar_workdays_month,
                expected_remaining_hours=expected_val,
                expected_pct=expected_pct,
                phase="future_expected",
                month_num=month_num,
            )
        )

    summary = dict(
        period_type=period_type,
        period_month=asof_month,
        billed=billed,
        remaining=remaining,
        project_logged_hours=project_logged_hours,
        completeness_ratio=completeness_ratio,
        scope_label=scope_label,
        year_available_hours=year_available_hours,
        year_available_workdays=year_available_workdays,
        year_calendar_workdays=year_calendar_workdays,
        holidays_excluded=holidays_excluded,
        nn_total_hours=(
            (billed_cumulative_at_asof + remaining)
            if (billed_cumulative_at_asof is not None and remaining is not None)
            else ((billed + remaining) if (billed is not None and remaining is not None) else None)
        ),
        month_segments=month_segments,
    )
    return summary, None


def _build_nn_sideways_bar_subtitle_text(nn_summary: Optional[Dict[str, Any]]) -> str:
    if not nn_summary:
        return ""

    nn_total_hours = _to_float(nn_summary.get("nn_total_hours")) or 0.0
    capacity_hours = _to_float(nn_summary.get("year_available_hours")) or 0.0
    month_segments = nn_summary.get("month_segments") or []
    if nn_total_hours <= 0:
        nn_total_hours = sum(
            max(_to_float(segment.get("hours")) or 0.0, 0.0)
            for segment in month_segments
        )
    if nn_total_hours <= 0:
        return ""

    subtitle_text = f"100% = {_company_label_long()} total: {nn_total_hours:.0f} h"
    if capacity_hours > 0:
        subtitle_text += f" | Capacity plan: {capacity_hours:.0f} h"
    return subtitle_text


def build_nn_sideways_bar_chart_html(nn_summary: Optional[Dict[str, Any]], div_id: str = "nn-sideways-bar") -> str:
    if not nn_summary:
        return ""
    segments = nn_summary.get("month_segments") or []
    nn_total_hours_raw = nn_summary.get("nn_total_hours")
    nn_total_hours = _to_float(nn_total_hours_raw) or 0.0
    capacity_hours_raw = nn_summary.get("year_available_hours")
    capacity_hours = _to_float(capacity_hours_raw) or 0.0
    total_calendar_workdays_raw = nn_summary.get("year_calendar_workdays")
    total_calendar_workdays = _to_float(total_calendar_workdays_raw)
    total_workable_days_raw = nn_summary.get("year_available_workdays")
    total_workable_days = _to_float(total_workable_days_raw)
    holidays_excluded = bool(nn_summary.get("holidays_excluded"))
    if total_calendar_workdays is None or total_calendar_workdays <= 0:
        total_calendar_workdays = total_workable_days
    if total_workable_days is None or total_workable_days <= 0:
        total_workable_days = (capacity_hours / 8.0) if capacity_hours > 0 else 0.0
    if total_calendar_workdays is None or total_calendar_workdays <= 0:
        total_calendar_workdays = total_workable_days

    total_segment_hours = 0.0
    cumulative_hours = 0.0
    cumulative_workdays = 0.0
    fig = go.Figure()

    hover_rows = [
        ("Block", "%{customdata[5]}"),
        ("Billed so far (current month)", "%{customdata[1]:.0f} h"),
        (f"Expected remaining for {_company_label_short()}", "%{customdata[3]:.0f} h"),
        ("Remaining workable hours", "%{customdata[2]:.0f} h"),
        (f"Expected {_company_label_short()} share this month", "%{customdata[4]:.0f}%"),
        ("Block hours", "%{customdata[6]:.0f} h"),
        ("Cumsum hours (year)", "%{customdata[7]:.0f} h"),
        (f"Block pct of {_company_label_long()} total", "%{customdata[9]:.0f}% (base %{customdata[8]:.0f} h)"),
        (f"Cumsum pct of {_company_label_long()} total", "%{customdata[10]:.0f}% (base %{customdata[8]:.0f} h)"),
        ("Cumsum working days (Mon-Fri)", "%{customdata[11]:.0f}% (%{customdata[12]:.0f}/%{customdata[13]:.0f} d)"),
    ]
    hover_label_width = max(len(label) for label, _ in hover_rows)
    hover_lines = ["<b>%{customdata[0]}</b>"]
    for label, value_expr in hover_rows:
        pad = "&nbsp;" * (hover_label_width - len(label) + 1)
        hover_lines.append(f"{label}:{pad}{value_expr}")
    hovertemplate = "<br>".join(hover_lines) + "<extra></extra>"

    valid_segments: List[Dict[str, Any]] = []
    month_total_hours: Dict[str, float] = {}
    month_workable_days: Dict[str, float] = {}
    month_calendar_workdays: Dict[str, float] = {}

    for segment in segments:
        seg_hours = _to_float(segment.get("hours")) or 0.0
        seg_hours = max(seg_hours, 0.0)
        if seg_hours <= 0:
            continue
        valid_segments.append(segment)
        month_key = str(segment.get("month_name", "")).strip() or str(segment.get("month_abbr", "")).strip()
        month_total_hours[month_key] = month_total_hours.get(month_key, 0.0) + seg_hours

        seg_calendar_workdays = _to_float(segment.get("calendar_workdays"))
        if seg_calendar_workdays is not None:
            month_calendar_workdays[month_key] = max(
                month_calendar_workdays.get(month_key, 0.0),
                max(seg_calendar_workdays, 0.0),
            )

        seg_workable_days = _to_float(segment.get("workable_days"))
        if seg_workable_days is None:
            seg_workable_hours = _to_float(segment.get("workable_hours")) or 0.0
            if capacity_hours > 0 and total_workable_days > 0:
                seg_workable_days = seg_workable_hours * (total_workable_days / capacity_hours)
            else:
                seg_workable_days = seg_workable_hours / 8.0
        month_workable_days[month_key] = max(month_workable_days.get(month_key, 0.0), max(seg_workable_days, 0.0))

    total_segment_hours = float(sum(month_total_hours.values()))
    if total_segment_hours <= 0:
        return ""

    if nn_total_hours <= 0:
        nn_total_hours = total_segment_hours
    nn_total_hours = max(nn_total_hours, 1.0)

    for segment in valid_segments:
        seg_hours = _to_float(segment.get("hours")) or 0.0
        seg_hours = max(seg_hours, 0.0)
        cumulative_hours += seg_hours
        month_key = str(segment.get("month_name", "")).strip() or str(segment.get("month_abbr", "")).strip()
        month_total_for_split = max(month_total_hours.get(month_key, 0.0), 0.0)
        segment_calendar_workdays = _to_float(segment.get("calendar_workdays"))
        if segment_calendar_workdays is None:
            month_calendar_workdays_total = max(month_calendar_workdays.get(month_key, 0.0), 0.0)
            if month_calendar_workdays_total > 0 and month_total_for_split > 0:
                segment_calendar_workdays = month_calendar_workdays_total * (seg_hours / month_total_for_split)
            else:
                month_workable_days_total = max(month_workable_days.get(month_key, 0.0), 0.0)
                segment_calendar_workdays = (
                    month_workable_days_total * (seg_hours / month_total_for_split) if month_total_for_split > 0 else 0.0
                )
        segment_calendar_workdays = max(float(segment_calendar_workdays), 0.0)
        cumulative_workdays += segment_calendar_workdays
        block_pct_of_nn_total = (seg_hours / nn_total_hours * 100.0) if nn_total_hours > 0 else 0.0
        cumsum_pct_of_nn_total = (cumulative_hours / nn_total_hours * 100.0) if nn_total_hours > 0 else 0.0
        cumsum_workdays_pct = (
            (cumulative_workdays / total_calendar_workdays * 100.0) if total_calendar_workdays > 0 else 0.0
        )

        phase = str(segment.get("phase", "")).strip().lower()
        if phase == "completed":
            color = YEAR_PLAN_COMPLETED_COLOR
        elif phase == "current_billed":
            color = YEAR_PLAN_CURRENT_BILLED_COLOR
        elif phase == "current_combined":
            color = YEAR_PLAN_CURRENT_COMBINED_COLOR
        elif phase == "current_expected":
            color = YEAR_PLAN_CURRENT_EXPECTED_COLOR
        else:
            color = YEAR_PLAN_EXPECTED_COLOR

        fig.add_trace(
            go.Bar(
                x=[seg_hours],
                y=[_company_label_short()],
                orientation="h",
                marker=dict(color=color, line=dict(color="#FFFFFF", width=1)),
                width=[0.62],
                text=[str(segment.get("month_abbr", ""))],
                textposition="inside",
                insidetextanchor="middle",
                textfont=dict(size=11, color="#FFFFFF"),
                customdata=[
                    [
                        str(segment.get("month_name", "")),
                        float(_to_float(segment.get("billed_so_far")) or 0.0),
                        float(_to_float(segment.get("workable_hours")) or 0.0),
                        float(_to_float(segment.get("expected_remaining_hours")) or 0.0),
                        float(_to_float(segment.get("expected_pct")) or 0.0),
                        str(segment.get("segment_label", "")),
                        seg_hours,
                        cumulative_hours,
                        nn_total_hours,
                        block_pct_of_nn_total,
                        cumsum_pct_of_nn_total,
                        cumsum_workdays_pct,
                        cumulative_workdays,
                        total_calendar_workdays,
                    ]
                ],
                hovertemplate=hovertemplate,
                showlegend=False,
            )
        )

    tick_fracs = [0.0, 0.25, 0.50, 0.75, 1.0]
    tickvals = [nn_total_hours * f for f in tick_fracs]
    ticktext = [f"{int(f * 100)}%" for f in tick_fracs]

    fig.update_layout(
        barmode="stack",
        margin=dict(l=2, r=4, t=8, b=14),
        height=128,
        hovermode="closest",
        hoverdistance=12,
        hoverlabel=dict(font=dict(size=11, family="Consolas, 'Courier New', monospace"), align="left"),
        showlegend=False,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        xaxis=dict(
            range=[0, nn_total_hours],
            tickvals=tickvals,
            ticktext=ticktext,
            showgrid=True,
            gridcolor="rgba(0,0,0,0.14)",
            zeroline=False,
            fixedrange=True,
        ),
        yaxis=dict(showticklabels=False, showgrid=False, zeroline=False, fixedrange=True),
    )
    return pio.to_html(fig, include_plotlyjs=False, full_html=False, div_id=div_id)


def build_nn_metrics_html(nn_summary: Optional[Dict[str, Any]], note: Optional[str]) -> str:
    if not nn_summary:
        note_text = note or f"{_company_label_short()} summary not available."
        return f"<div class='nn-note'>{html.escape(note_text)}</div>"

    billed = nn_summary.get("billed")
    remaining = nn_summary.get("remaining")
    logged = nn_summary.get("project_logged_hours")
    ratio = nn_summary.get("completeness_ratio")
    period_type = str(nn_summary.get("period_type", "")).strip().lower()
    scope_label = str(nn_summary.get("scope_label", "")).strip()
    if not scope_label:
        scope_label = "month" if period_type == "monthly" else "to-date"

    if period_type == "monthly":
        billed_label = f"Billed hours ({scope_label})"
        logged_label = f"Reported hours ({scope_label})"
        completeness_label = f"Tracking completeness ({scope_label})"
    else:
        billed_label = "Billed hours (to-date)"
        logged_label = "Reported hours (to-date)"
        completeness_label = "Tracking completeness (to-date)"

    def fmt(val: Optional[float]) -> str:
        if val is None:
            return "n/a"
        return f"{val:.0f}"

    ratio_text = f"{ratio * 100:.0f}%" if ratio is not None else "n/a"

    return (
        "<div class='nn-metrics'>"
        "<div><b>" + html.escape(billed_label) + "</b>: " + html.escape(fmt(billed)) + "</div>"
        "<div><b>Remaining</b>: " + html.escape(fmt(remaining)) + "</div>"
        "<div><b>" + html.escape(logged_label) + "</b>: " + html.escape(fmt(logged)) + "</div>"
        "<div><b>" + html.escape(completeness_label) + "</b>: " + html.escape(ratio_text) + "</div>"
        "</div>"
    )


def build_nn_sideways_bar_title_html(nn_summary: Optional[Dict[str, Any]] = None) -> str:
    help_lines = [
        f"100% reflects the total {_company_label_long()} hours planned for the selected year.",
        (
            "Workable hours or capacity can change over time and includes all planned holidays "
            "plus all planned other projects."
        ),
        (
            "It does not yet include illness, other unplanned absence, or any new projects that "
            "are not known yet."
        ),
        (
            f"How to read: if expected {_company_label_short()} share this month or expected remaining {_company_label_short()} hours drops "
            f"below your threshold, it can be a signal to reduce {_company_label_short()} hours short term so enough "
            "hours remain for later in the year."
        ),
        "Past months always show actual billed hours.",
    ]
    tooltip_html = "<br>".join(html.escape(line) for line in help_lines)
    subtitle_text = _build_nn_sideways_bar_subtitle_text(nn_summary)
    title_text = f"{_company_label_short()} Year Plan (Actual + Expected)"
    subtitle_html = f" | {html.escape(subtitle_text)}" if subtitle_text else ""
    return (
        "<div class='nn-sideways-bar-title-row'>"
        "<div class='nn-sideways-bar-title'>"
        f"{html.escape(title_text)}"
        f"<span class='nn-sideways-bar-subtitle'>{subtitle_html}</span>"
        "</div>"
        f"<span class='nn-help-icon' tabindex='0' role='button' aria-label='{html.escape(_company_label_short())} year plan explanation'>?"
        f"<span class='nn-help-tooltip'>{tooltip_html}</span>"
        "</span>"
        "</div>"
    )


def _plotly_cdn_src() -> str:
    # Pin plotly.js to the plotly.py-bundled version when possible.
    # Some older/newer plotly.py versions expose this as `plotlyjs_version` (public)
    # instead of `_plotlyjs_version` (private). Falling back to `plotly-latest` can
    # cause subtle rendering regressions between `.show()` and exported HTML.
    for attr in ("plotlyjs_version", "_plotlyjs_version"):
        version = getattr(pio, attr, None)
        if isinstance(version, str) and version.strip():
            return f"https://cdn.plot.ly/plotly-{version}.min.js"
    return "https://cdn.plot.ly/plotly-latest.min.js"


def build_project_info_tables_html(
    projects_df: pd.DataFrame, project_info_map: Dict[str, Dict[str, Any]]
) -> str:
    cards: List[str] = []
    for _, row in projects_df.iterrows():
        project_id = str(row.get("project_id", "")).strip()
        project_name = str(row.get("project_name", project_id)).strip()
        info = project_info_map.get(project_id, {})

        rows: List[str] = []
        for key in sorted(info.keys()):
            val = info.get(key)
            if val is None or (isinstance(val, float) and pd.isna(val)) or str(val).strip() == "":
                continue
            rows.append(
                "<tr><td>" + html.escape(str(key)) + "</td><td>" + html.escape(str(val)) + "</td></tr>"
            )

        if not rows:
            rows_html = "<tr><td colspan='2'>No data</td></tr>"
        else:
            rows_html = "".join(rows)

        cards.append(
            "<div class='project-card'>"
            "<div class='project-card-header'>" + html.escape(project_id) + " — " + html.escape(project_name) + "</div>"
            "<table>" + rows_html + "</table>"
            "</div>"
        )

    return "<div class='project-cards'>" + "".join(cards) + "</div>"


def _escape_html_multiline(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value)
    escaped = html.escape(text)
    escaped = escaped.replace("\r\n", "\n").replace("\r", "\n")
    return escaped.replace("\n", "<br/>")


def _format_minutes_hhmm(total_minutes: Any) -> str:
    if total_minutes is None or (isinstance(total_minutes, float) and pd.isna(total_minutes)):
        return "00:00"
    try:
        minutes_int = int(round(float(total_minutes)))
    except (TypeError, ValueError):
        return "00:00"
    sign = "-" if minutes_int < 0 else ""
    minutes_int = abs(minutes_int)
    hours, minutes = divmod(minutes_int, 60)
    return f"{sign}{hours:02d}:{minutes:02d}"


def build_logged_hours_breakdown_html(
    time_entries_df_filtered: pd.DataFrame,
    title: str = "Logged hours (by project)",
    show_percentage: bool = False,
    include_total_in_note: bool = False,
    foldable: bool = True,
) -> str:
    def _wrap(content_html: str) -> str:
        if foldable:
            return (
                "<details class='hours-breakdown'>"
                f"<summary>{html.escape(title)}</summary>"
                f"{content_html}"
                "</details>"
            )
        return (
            "<section class='hours-breakdown'>"
            "<div class='hours-breakdown-header'>"
            f"<h3>{html.escape(title)}</h3>"
            "</div>"
            f"{content_html}"
            "</section>"
        )

    if time_entries_df_filtered is None or time_entries_df_filtered.empty:
        return _wrap("<div class='hours-breakdown-note'>No time log entries in this period.</div>")

    required_cols = {"project_id", "duration_hours"}
    missing = [c for c in required_cols if c not in time_entries_df_filtered.columns]
    if missing:
        return _wrap(
            "<div class='hours-breakdown-note'>"
            "Time log entries are missing expected columns: "
            + html.escape(", ".join(missing))
            + "</div>"
        )

    entries = time_entries_df_filtered.copy()
    entries["project_id"] = entries["project_id"].astype(str).str.strip()

    if "project_name" not in entries.columns:
        entries["project_name"] = entries["project_id"]
    entries["project_name"] = entries["project_name"].fillna("").astype(str).str.strip()
    entries.loc[entries["project_name"] == "", "project_name"] = entries["project_id"]

    what_i_did_col: Optional[str] = None
    for candidate in ("WhatIDid*", "WhatIDid"):
        if candidate in entries.columns:
            what_i_did_col = candidate
            break

    entries["duration_hours"] = pd.to_numeric(entries["duration_hours"], errors="coerce")
    entries = entries.dropna(subset=["duration_hours"]).copy()
    entries = entries[entries["duration_hours"] > 0].copy()
    if "duration_minutes" in entries.columns:
        entries["duration_minutes"] = pd.to_numeric(entries["duration_minutes"], errors="coerce")
    else:
        entries["duration_minutes"] = entries["duration_hours"] * 60.0
    if entries.empty:
        return _wrap("<div class='hours-breakdown-note'>No logged hours in this period.</div>")

    totals = (
        entries.groupby(["project_id", "project_name"], as_index=False)["duration_minutes"]
        .sum(min_count=1)
        .rename(columns={"duration_minutes": "total_minutes"})
    )
    totals["total_minutes"] = pd.to_numeric(totals["total_minutes"], errors="coerce").fillna(0.0)
    totals = totals.sort_values(["total_minutes", "project_name", "project_id"], ascending=[False, True, True])

    total_period_minutes = float(entries["duration_minutes"].sum()) if show_percentage else 0.0
    if show_percentage and total_period_minutes <= 0:
        show_percentage = False

    def fmt_pct(minutes: Any) -> str:
        if not show_percentage:
            return ""
        try:
            m = float(minutes or 0.0)
        except (TypeError, ValueError):
            m = 0.0
        return f"{(m * 100.0 / total_period_minutes):.1f}%"

    projects_html: List[str] = []
    for _, row in totals.iterrows():
        project_id = str(row.get("project_id", "")).strip()
        project_name = str(row.get("project_name", project_id)).strip() or project_id
        total_minutes = float(row.get("total_minutes", 0.0) or 0.0)

        project_entries = entries.loc[entries["project_id"] == project_id].copy()
        project_entries = project_entries.sort_values(["duration_minutes"], ascending=False)

        entry_rows: List[str] = []
        for _, entry in project_entries.iterrows():
            dur_text = _format_minutes_hhmm(entry.get("duration_minutes"))
            pct_text = fmt_pct(entry.get("duration_minutes"))
            desc_val = entry.get(what_i_did_col) if what_i_did_col else ""
            desc_html = _escape_html_multiline(desc_val).strip()
            if not desc_html:
                desc_html = "<span class='hours-entry-empty'>(no details)</span>"
            pct_cell = (
                f"<td class='hours-entry-percent'>{html.escape(pct_text)}</td>"
                if show_percentage
                else ""
            )
            entry_rows.append(
                "<tr>"
                f"<td class='hours-entry-duration'>{html.escape(dur_text)}</td>"
                f"{pct_cell}"
                f"<td>{desc_html}</td>"
                "</tr>"
            )

        pct_summary = f" <span class='hours-project-percent'>({html.escape(fmt_pct(total_minutes))})</span>" if show_percentage else ""
        summary_html = (
            f"<span class='hours-project-total'>{html.escape(_format_minutes_hhmm(total_minutes))}</span>, "
            f"<span class='hours-project-name'>{html.escape(project_name)}</span>"
            f"{pct_summary}"
        )

        header_pct = "<th>Percent</th>" if show_percentage else ""
        projects_html.append(
            "<details class='hours-project'>"
            f"<summary>{summary_html}</summary>"
            "<div class='hours-project-entries'>"
            "<table class='hours-entry-table'>"
            f"<thead><tr><th>Duration</th>{header_pct}<th>Details</th></tr></thead>"
            "<tbody>"
            + "".join(entry_rows)
            + "</tbody>"
            "</table>"
            "</div>"
            "</details>"
        )

    if show_percentage:
        note_text = "Percentages are of the total logged time in this period."
        if include_total_in_note:
            note_text += f" Total logged time: {_format_minutes_hhmm(total_period_minutes)}."
    else:
        note_text = "Click a project to expand."
    return _wrap(
        f"<div class='hours-breakdown-note'>{html.escape(note_text)}</div>"
        "<div class='hours-breakdown-list'>"
        + "".join(projects_html)
        + "</div>"
    )


def _activitytype_column_name(df: pd.DataFrame) -> Optional[str]:
    for candidate in ("ActivityType*", "ActivityType"):
        if candidate in df.columns:
            return candidate
    return None


def compute_activitytype_weighted_distribution(time_entries_df: pd.DataFrame) -> pd.DataFrame:
    if time_entries_df is None or time_entries_df.empty:
        return pd.DataFrame(columns=["activity_type", "hours", "pct"])
    col = _activitytype_column_name(time_entries_df)
    if not col:
        return pd.DataFrame(columns=["activity_type", "hours", "pct"])

    tmp = time_entries_df.copy()
    tmp["__hours"] = pd.to_numeric(tmp.get("duration_hours"), errors="coerce").fillna(0.0)
    tmp["__activity_type"] = tmp[col].fillna("").astype(str).map(lambda s: s.strip())
    tmp.loc[tmp["__activity_type"] == "", "__activity_type"] = "Unknown"
    tmp = tmp.loc[tmp["__hours"] > 0].copy()
    if tmp.empty:
        return pd.DataFrame(columns=["activity_type", "hours", "pct"])

    grouped = (
        tmp.groupby("__activity_type", as_index=False)["__hours"]
        .sum(min_count=1)
        .rename(columns={"__activity_type": "activity_type", "__hours": "hours"})
    )
    grouped["hours"] = pd.to_numeric(grouped["hours"], errors="coerce").fillna(0.0)
    grouped = grouped.loc[grouped["hours"] > 0].copy()
    if grouped.empty:
        return pd.DataFrame(columns=["activity_type", "hours", "pct"])
    grouped = grouped.sort_values("hours", ascending=False).reset_index(drop=True)
    total = float(grouped["hours"].sum())
    grouped["pct"] = grouped["hours"] / total * 100.0 if total > 0 else 0.0
    return grouped


def add_activitytype_weighted_pie(
    fig: go.Figure,
    time_entries_df: pd.DataFrame,
    subplot_row: int,
    title: str,
) -> None:
    dist = compute_activitytype_weighted_distribution(time_entries_df)
    if dist.empty:
        fig.add_trace(
            go.Pie(
                labels=["(no ActivityType* data)"],
                values=[1],
                hole=0.45,
                textinfo="label",
                marker=dict(colors=["#E3E3E3"]),
                hovertemplate="No ActivityType* data available.<extra></extra>",
                showlegend=False,
            ),
            row=subplot_row,
            col=1,
        )
    else:
        colors = [TEAMNL_BASE_COLORS[i % len(TEAMNL_BASE_COLORS)] for i in range(len(dist))]
        fig.add_trace(
            go.Pie(
                labels=dist["activity_type"].tolist(),
                values=dist["hours"].tolist(),
                hole=0.45,
                textinfo="percent+label",
                textposition="outside",
                marker=dict(colors=colors),
                hovertemplate="%{label}<br>Hours=%{value:.2f}<br>Share=%{percent}<extra></extra>",
                showlegend=False,
                sort=False,
            ),
            row=subplot_row,
            col=1,
        )

    # Pie charts are "domain" subplots, where add_annotation(row=..., col=...)
    # raises in Plotly. Anchor the title to the pie domain in paper coords.
    try:
        last_trace = fig.data[-1] if fig.data else None
        domain = getattr(last_trace, "domain", None) if last_trace is not None else None
        x_domain = list(getattr(domain, "x", []) or [])
        y_domain = list(getattr(domain, "y", []) or [])
        if len(x_domain) == 2 and len(y_domain) == 2:
            fig.add_annotation(
                text=f"<b>{title}</b>",
                x=float(x_domain[0]),
                xref="paper",
                xanchor="left",
                y=min(float(y_domain[1]) + 0.03, 1.0),
                yref="paper",
                yanchor="bottom",
                showarrow=False,
                align="left",
            )
            return
    except Exception:
        pass

    # Fallback for non-domain subplots (kept for safety/future reuse).
    try:
        fig.add_annotation(
            text=f"<b>{title}</b>",
            x=0,
            xref=axis_domain_ref("x", subplot_row),
            y=1.18,
            yref=axis_domain_ref("y", subplot_row),
            showarrow=False,
            align="left",
            row=subplot_row,
            col=1,
        )
    except Exception:
        # Never fail report generation on decorative title annotation.
        return


def build_project_activitytype_pie_html(project_entries_df: pd.DataFrame) -> str:
    dist = compute_activitytype_weighted_distribution(project_entries_df)
    if dist.empty:
        return (
            "<div class='project-activitytype'>"
            "<div class='project-timelog-title'>ActivityType* (weighted by hours)</div>"
            "<div class='hours-entry-empty'>No ActivityType* data.</div>"
            "</div>"
        )

    segments: List[str] = []
    legend_rows: List[str] = []
    start_pct = 0.0
    for idx, row in dist.iterrows():
        label = str(row.get("activity_type", "Unknown")).strip() or "Unknown"
        pct = float(row.get("pct", 0.0) or 0.0)
        end_pct = min(start_pct + pct, 100.0)
        color = TEAMNL_BASE_COLORS[int(idx) % len(TEAMNL_BASE_COLORS)]
        segments.append(f"{color} {start_pct:.4f}% {end_pct:.4f}%")
        start_pct = end_pct
        legend_rows.append(
            "<div class='project-activitytype-row'>"
            f"<span class='project-activitytype-dot' style='background:{html.escape(color)}'></span>"
            f"<span class='project-activitytype-label'>{html.escape(label)} ({pct:.1f}%)</span>"
            "</div>"
        )

    gradient = ", ".join(segments) if segments else "#E3E3E3 0% 100%"
    return (
        "<div class='project-activitytype'>"
        "<div class='project-timelog-title'>ActivityType* (weighted by hours)</div>"
        f"<div class='project-activitytype-pie' style='background:conic-gradient({gradient});'></div>"
        "<div class='project-activitytype-legend'>"
        + "".join(legend_rows)
        + "</div>"
        "</div>"
    )


# ----------------------------
# Plot builders
# ----------------------------
def axis_domain_ref(axis_letter: str, subplot_row: int) -> str:
    # Plotly uses "x domain"/"y domain" for the first subplot, not "x1 domain"/"y1 domain"
    return f"{axis_letter} domain" if subplot_row == 1 else f"{axis_letter}{subplot_row} domain"


def apply_axis_style(fig: go.Figure, total_rows: int) -> None:
    axis_style = dict(
        showgrid=True,
        gridwidth=1,
        gridcolor="rgba(0,0,0,0.15)",
        showline=True,
        linewidth=1,
        linecolor="rgba(0,0,0,0.8)",
        mirror=True,
        ticks="outside",
        zeroline=True,
        zerolinewidth=1,
        zerolinecolor="rgba(0,0,0,0.15)",
        automargin=True,
    )
    for row in range(1, total_rows + 1):
        try:
            fig.update_xaxes(**axis_style, row=row, col=1)
            fig.update_yaxes(**axis_style, row=row, col=1, tickmode="auto")
        except Exception:
            # Domain traces (e.g., pie charts) do not expose cartesian axes.
            continue


def _find_asset(prefix: str, exts: Tuple[str, ...]) -> Optional[str]:
    asset_dir = ASSETS_DIR
    if not os.path.isdir(asset_dir):
        return None
    for filename in os.listdir(asset_dir):
        lower = filename.lower()
        if lower.startswith(prefix) and lower.endswith(exts):
            return os.path.join(asset_dir, filename)
    return None


def _encode_image_to_data_uri(img_path: str) -> Optional[str]:
    if not img_path or not os.path.exists(img_path):
        return None
    mime = "image/png"
    lower = img_path.lower()
    if lower.endswith(".svg"):
        mime = "image/svg+xml"
    elif lower.endswith((".jpg", ".jpeg")):
        mime = "image/jpeg"
    elif lower.endswith(".webp"):
        mime = "image/webp"
    with open(img_path, "rb") as f:
        encoded = base64.b64encode(f.read()).decode("ascii")
    return f"data:{mime};base64,{encoded}"


def _preferred_asset(configured_path: str, fallback_prefix: str) -> Optional[str]:
    if configured_path and os.path.exists(configured_path):
        return configured_path
    return _find_asset(fallback_prefix, (".png", ".jpg", ".jpeg", ".svg", ".webp"))


def add_company_logo(fig: go.Figure) -> None:
    logo_path = _preferred_asset(TEAMNL_LOGO_PATH, "logo")
    data_uri = _encode_image_to_data_uri(logo_path) if logo_path else None
    if not data_uri:
        print(
            "Logo image not found. Update 'branding.logo' in projexcellent_config.json "
            f"or add a matching file to {ASSETS_DIR}."
        )
        return

    fig.add_layout_image(
        dict(
            source=data_uri,
            xref="paper",
            yref="paper",
            x=1.02,
            y=1.14,
            sizex=0.24,
            sizey=0.12,
            xanchor="right",
            yanchor="top",
            layer="above",
        )
    )


# Backward-compatible alias for older code paths.
add_teamnl_logo = add_company_logo


def build_header_assets() -> Dict[str, Optional[str]]:
    profile_path = _preferred_asset(PROFILE_PHOTO_PATH, "profile_photo")
    company_logo_path = _preferred_asset(TEAMNL_LOGO_PATH, "logo")
    return {
        "profile_data_uri": _encode_image_to_data_uri(profile_path) if profile_path else None,
        "company_logo_data_uri": _encode_image_to_data_uri(company_logo_path) if company_logo_path else None,
    }


def add_profile_picture(fig: go.Figure) -> None:
    profile_path = _preferred_asset(PROFILE_PHOTO_PATH, "profile_photo")
    data_uri = _encode_image_to_data_uri(profile_path) if profile_path else None
    if not data_uri:
        print(
            "Profile picture not found. Update 'branding.profile_photo' in projexcellent_config.json "
            f"or add a matching file to {ASSETS_DIR}."
        )
        return

    fig.add_layout_image(
        dict(
            source=data_uri,
            xref="paper",
            yref="paper",
            x=0.82,
            y=1.14,
            sizex=0.12,
            sizey=0.12,
            xanchor="right",
            yanchor="top",
            layer="above",
        )
    )


def add_stacked_project_count_bars(
    fig: go.Figure,
    projects_df: pd.DataFrame,
    group_col: str,
    subplot_row: int,
    title: str,
    project_color_map: Dict[str, str],
) -> None:
    all_groups: set[str] = set()
    project_groups: List[Tuple[pd.Series, List[str]]] = []
    group_counts: Dict[str, int] = {}

    for _, project in projects_df.iterrows():
        values = extract_group_values(project, group_col)
        if not values:
            values = ["Unknown"]
        values = list(dict.fromkeys(values))  # preserve order, drop dupes
        project_groups.append((project, values))
        all_groups.update(values)
        for group_val in values:
            group_counts[group_val] = group_counts.get(group_val, 0) + 1

    groups = sorted(all_groups, key=lambda g: (-group_counts.get(g, 0), str(g)))
    fig.update_xaxes(categoryorder="array", categoryarray=groups, row=subplot_row, col=1)

    # Stacking order: first traces are at the bottom; ensure closed projects sit below active ones.
    project_groups.sort(key=lambda item: 1 if is_active_status(item[0].get("status")) else 0)

    for project, values in project_groups:
        project_id = str(project.get("project_id", "")).strip()
        base_color = project_color_map.get(project_id, BASE_BLACK)
        marker_color, opacity = marker_style_for_status(base_color, project.get("status"))
        hover = build_hover_text(project, extra={"resolved_end_date": resolve_end_date_for_hover(project)})
        bar_name = str(project.get("project_name", project.get("project_id", "project")))

        for group_val in values:
            fig.add_trace(
                go.Bar(
                    x=[group_val],
                    y=[1],
                    name=bar_name,
                    hovertemplate=hover + "<extra></extra>",
                    marker_color=marker_color,
                    opacity=opacity,
                    showlegend=False,
                ),
                row=subplot_row,
                col=1,
            )

    fig.update_yaxes(title_text="Project count", row=subplot_row, col=1)
    fig.add_annotation(
        text=f"<b>{title}</b>",
        x=0,
        xref="x domain",
        y=1.18,
        yref=axis_domain_ref("y", subplot_row),
        showarrow=False,
        align="left",
        row=subplot_row,
        col=1,
    )    


def add_stacked_hours_bars(
    fig: go.Figure,
    projects_df: pd.DataFrame,
    time_entries_df: pd.DataFrame,
    group_col: str,
    subplot_row: int,
    title: str,
    project_color_map: Dict[str, str],
    section_title: Optional[str] = None,
) -> None:
    def _add_section_title() -> None:
        if not section_title:
            return
        fig.add_annotation(
            text=f"<b>{section_title}</b>",
            x=0,
            xref="x domain",
            y=1.34,
            yref=axis_domain_ref("y", subplot_row),
            xanchor="left",
            yanchor="bottom",
            showarrow=False,
            align="left",
            font=dict(size=24, color=BASE_BLUE),
            row=subplot_row,
            col=1,
        )

    if time_entries_df.empty:
        fig.add_trace(
            go.Bar(x=["(no time_log data found)"], y=[0], hovertemplate="No time entries found.<extra></extra>", showlegend=False),
            row=subplot_row, col=1
        )
        fig.update_yaxes(title_text="Hours", row=subplot_row, col=1)
        fig.add_annotation(
            text=f"<b>{title}</b>",
            x=0,
            xref="x domain",
            y=1.18,
            yref=axis_domain_ref("y", subplot_row),
            xanchor="left",
            yanchor="bottom",
            showarrow=False,
            align="left",
            row=subplot_row,
            col=1,
        )        
        _add_section_title()
        return

    project_hours = (
        time_entries_df.groupby("project_id", as_index=False)["duration_hours"]
        .sum()
        .rename(columns={"duration_hours": "total_hours"})
    )
    merged = projects_df.merge(project_hours, on="project_id", how="left")
    merged["total_hours"] = merged["total_hours"].fillna(0)

    all_groups: set[str] = set()
    project_groups: List[Tuple[pd.Series, List[str]]] = []
    group_hours: Dict[str, float] = {}

    for _, project in merged.iterrows():
        values = extract_group_values(project, group_col)
        if not values:
            values = ["Unknown"]
        values = list(dict.fromkeys(values))
        project_groups.append((project, values))
        all_groups.update(values)
        hours = float(project.get("total_hours", 0.0))
        for group_val in values:
            group_hours[group_val] = group_hours.get(group_val, 0.0) + hours

    groups = sorted(all_groups, key=lambda g: (-group_hours.get(g, 0.0), str(g)))
    fig.update_xaxes(categoryorder="array", categoryarray=groups, row=subplot_row, col=1)

    # Stacking order: first traces are at the bottom; ensure closed projects sit below active ones.
    project_groups.sort(key=lambda item: 1 if is_active_status(item[0].get("status")) else 0)

    for project, values in project_groups:
        hours = float(project.get("total_hours", 0.0))
        project_id = str(project.get("project_id", "")).strip()
        base_color = project_color_map.get(project_id, BASE_BLACK)
        marker_color, opacity = marker_style_for_status(base_color, project.get("status"))
        hover = build_hover_text(
            project,
            extra={"total_hours": f"{hours:.2f}", "resolved_end_date": resolve_end_date_for_hover(project)},
        )

        for group_val in values:
            fig.add_trace(
                go.Bar(
                    x=[group_val],
                    y=[hours],
                    name=str(project.get("project_name", project.get("project_id", "project"))),
                    hovertemplate=hover + "<extra></extra>",
                    marker_color=marker_color,
                    opacity=opacity,
                    showlegend=False,
                ),
                row=subplot_row,
                col=1,
            )

    fig.update_yaxes(title_text="Hours", row=subplot_row, col=1)
    fig.add_annotation(
        text=f"<b>{title}</b>",
        x=0,
        xref="x domain",
        y=1.18,
        yref=axis_domain_ref("y", subplot_row),
        xanchor="left",
        yanchor="bottom",
        showarrow=False,
        align="left",
        row=subplot_row,
        col=1,
    )
    _add_section_title()

def add_trend_started_closed(
    fig: go.Figure,
    projects_df: pd.DataFrame,
    subplot_row: int,
    title: str,
    project_color_map: Optional[Dict[str, str]] = None,
    target_year: Optional[int] = None,
) -> None:
    if target_year is None:
        target_year = date.today().year

    year_start = date(target_year, 1, 1)
    year_end = date(target_year, 12, 31)
    week_starts, week_ends, week_positions, bar_width_ms = build_year_week_grid(target_year)
    x_positions = week_positions.to_pydatetime().tolist() if len(week_positions) else []
    if week_starts.empty:
        fig.add_trace(
            go.Bar(x=[], y=[], hovertemplate="No weekly activity data.<extra></extra>"),
            row=subplot_row,
            col=1,
        )
        fig.add_annotation(
            text=f"<b>{title}</b> (no weeks found)",
            x=0,
            xref="x domain",
            y=1.18,
            yref=axis_domain_ref("y", subplot_row),
            showarrow=False,
            align="left",
            row=subplot_row,
            col=1,
        )
        fig.update_yaxes(title_text="Active projects", row=subplot_row, col=1)
        return

    project_rows = projects_df.dropna(subset=["start_date"]).copy()
    if project_rows.empty:
        fig.add_trace(
            go.Bar(x=[], y=[], hovertemplate="No weekly activity data.<extra></extra>"),
            row=subplot_row,
            col=1,
        )
        fig.add_annotation(
            text=f"<b>{title}</b> (no project dates found)",
            x=0,
            xref="x domain",
            y=1.18,
            yref=axis_domain_ref("y", subplot_row),
            showarrow=False,
            align="left",
            row=subplot_row,
            col=1,
        )
        fig.update_yaxes(title_text="Active projects", row=subplot_row, col=1)
        return

    project_rows["project_id"] = project_rows["project_id"].astype(str)
    project_rows = project_rows.set_index("project_id")

    def _resolve_expected_end(row: pd.Series) -> Optional[pd.Timestamp]:
        expected = row.get("expected_end_date")
        expected = parse_date(expected) if expected is not None else None
        if expected is not None:
            return expected
        target_end = row.get("target_end_date")
        if pd.notna(target_end):
            return pd.Timestamp(target_end)
        return None

    def _project_sort_key(project_id: str) -> Tuple[int, str]:
        try:
            return int(project_id.replace("_", "")), project_id
        except (TypeError, ValueError):
            return float("inf"), project_id

    ordered_projects = sorted(project_rows.index.tolist(), key=_project_sort_key)
    # Stacking order: first traces are at the bottom; ensure closed projects sit below active ones.
    ordered_projects.sort(key=lambda pid: 1 if is_active_status(project_rows.loc[pid].get("status")) else 0)
    color_map = project_color_map or {}

    for project_id in ordered_projects:
        row = project_rows.loc[project_id]
        project_start = row.get("start_date")
        if pd.isna(project_start):
            continue
        project_start = pd.Timestamp(project_start)

        status_val = normalize_status_for_display(row.get("status"))
        if status_val.lower() == "closed":
            project_end = row.get("actual_end_date")
            if pd.isna(project_end):
                continue
            project_end = pd.Timestamp(project_end)
        else:
            expected_end = _resolve_expected_end(row)
            project_end = expected_end if expected_end is not None else pd.Timestamp(year_end)

        active = (week_ends >= project_start) & (week_starts <= project_end)
        if not active.any():
            continue
        y_vals = active.astype(int).tolist()

        base_color = color_map.get(project_id, BASE_BLACK)
        marker_color, opacity = marker_style_for_status(base_color, status_val)
        hover_text = build_hover_text(row, extra={"resolved_end_date": resolve_end_date_for_hover(row)})

        fig.add_trace(
            go.Bar(
                x=x_positions,
                y=y_vals,
                name=project_id,
                width=[bar_width_ms] * len(y_vals),
                marker_color=marker_color,
                opacity=opacity,
                hovertext=[hover_text] * len(y_vals),
                hovertemplate="%{hovertext}<br>Week starting %{x|%Y-%m-%d}<br>Active=%{y}<extra></extra>",
                showlegend=False,
            ),
            row=subplot_row,
            col=1,
        )

    fig.update_yaxes(title_text="Active projects", row=subplot_row, col=1)
    fig.update_xaxes(title_text="Week", row=subplot_row, col=1, type="date")
    fig.add_annotation(
        text=f"<b>{title}</b>",
        x=0,
        xref="x domain",
        y=1.18,
        yref=axis_domain_ref("y", subplot_row),
        showarrow=False,
        align="left",
        row=subplot_row,
        col=1,
    )


def add_hours_per_week(
    fig: go.Figure,
    projects_df: pd.DataFrame,
    time_entries_df: pd.DataFrame,
    subplot_row: int,
    title: str,
    project_color_map: Optional[Dict[str, str]] = None,
    display_start: Optional[date] = None,
    display_end: Optional[date] = None,
    data_start: Optional[date] = None,
    data_end: Optional[date] = None,
) -> None:
    if display_start is None or display_end is None:
        display_start = date.today().replace(month=1, day=1)
        display_end = date.today()
    if data_start is None:
        data_start = display_start
    if data_end is None:
        data_end = display_end

    week_starts, week_ends, week_positions, bar_width_ms = build_period_week_grid(display_start, display_end)
    x_positions = week_positions.to_pydatetime().tolist() if len(week_positions) else []
    if week_starts.empty:
        fig.add_trace(
            go.Bar(x=[], y=[], hovertemplate="No weekly hours data.<extra></extra>"),
            row=subplot_row,
            col=1,
        )
        fig.add_annotation(
            text=f"<b>{title}</b> (no weeks found)",
            x=0,
            xref="x domain",
            y=1.18,
            yref=axis_domain_ref("y", subplot_row),
            showarrow=False,
            align="left",
            row=subplot_row,
            col=1,
        )
        fig.update_yaxes(title_text="Hours", row=subplot_row, col=1)
        fig.update_xaxes(title_text="Week", row=subplot_row, col=1, type="date")
        return

    required_cols = {"project_id", "date", "duration_hours"}
    if time_entries_df.empty or not required_cols.issubset(time_entries_df.columns):
        fig.add_trace(
            go.Scatter(x=[x_positions[0]] if x_positions else [], y=[0], mode="markers", marker_opacity=0, showlegend=False),
            row=subplot_row,
            col=1,
        )
        fig.add_annotation(
            text=f"<b>{title}</b> (no time entries)",
            x=0,
            xref="x domain",
            y=1.18,
            yref=axis_domain_ref("y", subplot_row),
            showarrow=False,
            align="left",
            row=subplot_row,
            col=1,
        )
        fig.update_yaxes(title_text="Hours", row=subplot_row, col=1)
        fig.update_xaxes(title_text="Week", row=subplot_row, col=1, type="date")
        return

    entries = time_entries_df.dropna(subset=["project_id", "date", "duration_hours"]).copy()
    entries["project_id"] = entries["project_id"].astype(str)
    entries["duration_hours"] = pd.to_numeric(entries["duration_hours"], errors="coerce")
    entries = entries.dropna(subset=["duration_hours"])
    entries = entries[(entries["date"] >= pd.Timestamp(data_start)) & (entries["date"] <= pd.Timestamp(data_end))]
    if entries.empty:
        fig.add_trace(
            go.Scatter(x=[x_positions[0]] if x_positions else [], y=[0], mode="markers", marker_opacity=0, showlegend=False),
            row=subplot_row,
            col=1,
        )
        fig.add_annotation(
            text=f"<b>{title}</b> (no time entries)",
            x=0,
            xref="x domain",
            y=1.18,
            yref=axis_domain_ref("y", subplot_row),
            showarrow=False,
            align="left",
            row=subplot_row,
            col=1,
        )
        fig.update_yaxes(title_text="Hours", row=subplot_row, col=1)
        fig.update_xaxes(title_text="Week", row=subplot_row, col=1, type="date")
        return

    entries["__week_start"] = entries["date"].dt.normalize() - pd.to_timedelta(entries["date"].dt.weekday, unit="D")
    weekly_hours = (
        entries.groupby(["project_id", "__week_start"], as_index=False)["duration_hours"]
        .sum(min_count=1)
        .rename(columns={"duration_hours": "week_hours"})
    )
    weekly_hours["week_hours"] = pd.to_numeric(weekly_hours["week_hours"], errors="coerce").fillna(0.0)
    totals_by_project = (
        weekly_hours.groupby("project_id")["week_hours"]
        .sum(min_count=1)
        .fillna(0.0)
    )

    project_rows = projects_df.dropna(subset=["start_date"]).copy()

    project_rows["project_id"] = project_rows["project_id"].astype(str)
    project_rows = project_rows.set_index("project_id")

    def _resolve_expected_end(row: pd.Series) -> Optional[pd.Timestamp]:
        expected = row.get("expected_end_date")
        expected = parse_date(expected) if expected is not None else None
        if expected is not None:
            return expected
        target_end = row.get("target_end_date")
        if pd.notna(target_end):
            return pd.Timestamp(target_end)
        return None

    def _project_sort_key(project_id: str) -> Tuple[int, str]:
        try:
            return int(project_id.replace("_", "")), project_id
        except (TypeError, ValueError):
            return float("inf"), project_id

    ordered_projects = sorted(totals_by_project.index.tolist(), key=_project_sort_key)
    # Stacking order: first traces are at the bottom; ensure closed projects sit below active ones.
    status_by_project: Dict[str, Any] = {}
    if not projects_df.empty and "project_id" in projects_df.columns:
        for _, prow in projects_df.iterrows():
            pid = str(prow.get("project_id", "")).strip()
            if pid and pid not in status_by_project:
                status_by_project[pid] = prow.get("status")
    ordered_projects.sort(key=lambda pid: 1 if is_active_status(status_by_project.get(pid)) else 0)
    color_map = project_color_map or {}

    week_index_map = {pd.Timestamp(ws): idx for idx, ws in enumerate(week_starts)}
    added_trace = False

    for project_id in ordered_projects:
        total_hours = float(totals_by_project.get(project_id, 0.0) or 0.0)
        if total_hours <= 0:
            continue

        if project_id in project_rows.index:
            row = project_rows.loc[project_id]
        else:
            row = pd.Series({"project_id": project_id, "project_name": project_id})

        y_vals = [0.0] * len(week_starts)
        project_weekly = weekly_hours.loc[weekly_hours["project_id"] == project_id]
        for _, wrow in project_weekly.iterrows():
            week_start = wrow.get("__week_start")
            if pd.isna(week_start):
                continue
            idx = week_index_map.get(pd.Timestamp(week_start))
            if idx is None:
                continue
            y_vals[idx] = float(wrow.get("week_hours", 0.0) or 0.0)

        if not any(v > 0 for v in y_vals):
            continue

        base_color = color_map.get(project_id, BASE_BLACK)
        marker_color, opacity = marker_style_for_status(base_color, row.get("status"))
        hover_text = build_hover_text(
            row,
            extra={"period_hours": f"{total_hours:.2f}", "resolved_end_date": resolve_end_date_for_hover(row)},
        )

        fig.add_trace(
            go.Bar(
                x=x_positions,
                y=y_vals,
                name=project_id,
                width=[bar_width_ms] * len(y_vals),
                marker_color=marker_color,
                opacity=opacity,
                hovertext=[hover_text] * len(y_vals),
                hovertemplate="%{hovertext}<br>Week starting %{x|%Y-%m-%d}<br>Hours=%{y:.2f}<extra></extra>",
                showlegend=False,
            ),
            row=subplot_row,
            col=1,
        )
        added_trace = True

    if not added_trace:
        fig.add_trace(
            go.Scatter(x=[x_positions[0]] if x_positions else [], y=[0], mode="markers", marker_opacity=0, showlegend=False),
            row=subplot_row,
            col=1,
        )
        fig.add_annotation(
            text="No reported hours data.",
            x=0.5,
            xref="x domain",
            y=0.5,
            yref=axis_domain_ref("y", subplot_row),
            showarrow=False,
            align="center",
            row=subplot_row,
            col=1,
        )

    fig.update_yaxes(title_text="Hours", row=subplot_row, col=1)
    fig.update_xaxes(title_text="Week", row=subplot_row, col=1, type="date")
    fig.add_annotation(
        text=f"<b>{title}</b>",
        x=0,
        xref="x domain",
        y=1.18,
        yref=axis_domain_ref("y", subplot_row),
        showarrow=False,
        align="left",
        row=subplot_row,
        col=1,
    )


def add_estimated_magnitude_per_week(
    fig: go.Figure,
    projects_df: pd.DataFrame,
    subplot_row: int,
    title: str,
    project_color_map: Optional[Dict[str, str]] = None,
    period_start: Optional[date] = None,
    period_end: Optional[date] = None,
) -> None:
    if period_start is None or period_end is None:
        period_start = date.today().replace(month=1, day=1)
        period_end = date.today()
    year_end = period_end
    week_starts, week_ends, week_positions, bar_width_ms = build_period_week_grid(period_start, period_end)
    x_positions = week_positions.to_pydatetime().tolist() if len(week_positions) else []
    if week_starts.empty:
        fig.add_trace(
            go.Bar(x=[], y=[], hovertemplate="No weekly magnitude data.<extra></extra>"),
            row=subplot_row,
            col=1,
        )
        fig.add_annotation(
            text=f"<b>{title}</b> (no weeks found)",
            x=0,
            xref="x domain",
            y=1.18,
            yref=axis_domain_ref("y", subplot_row),
            showarrow=False,
            align="left",
            row=subplot_row,
            col=1,
        )
        fig.update_yaxes(title_text="Estimated magnitude", row=subplot_row, col=1)
        fig.update_xaxes(title_text="Week", row=subplot_row, col=1, type="date")
        return

    project_rows = projects_df.dropna(subset=["start_date"]).copy()
    if project_rows.empty:
        fig.add_trace(
            go.Bar(x=[], y=[], hovertemplate="No weekly magnitude data.<extra></extra>"),
            row=subplot_row,
            col=1,
        )
        fig.add_annotation(
            text=f"<b>{title}</b> (no project dates found)",
            x=0,
            xref="x domain",
            y=1.18,
            yref=axis_domain_ref("y", subplot_row),
            showarrow=False,
            align="left",
            row=subplot_row,
            col=1,
        )
        fig.update_yaxes(title_text="Estimated magnitude", row=subplot_row, col=1)
        fig.update_xaxes(title_text="Week", row=subplot_row, col=1, type="date")
        return

    project_rows["project_id"] = project_rows["project_id"].astype(str)
    project_rows = project_rows.set_index("project_id")

    def _resolve_hours_cap(row: pd.Series) -> Optional[float]:
        hours_cap = _to_float_relaxed(row.get("hours_cap"))
        if hours_cap is not None:
            return hours_cap
        cap_key = _find_col([str(c) for c in row.index], ["hours", "cap"])
        if cap_key:
            return _to_float_relaxed(row.get(cap_key))
        return None

    def _resolve_expected_end(row: pd.Series) -> Optional[pd.Timestamp]:
        expected = row.get("expected_end_date")
        expected = parse_date(expected) if expected is not None else None
        if expected is not None:
            return expected
        target_end = row.get("target_end_date")
        if pd.notna(target_end):
            return pd.Timestamp(target_end)
        return None

    def _project_sort_key(project_id: str) -> Tuple[int, str]:
        try:
            return int(project_id.replace("_", "")), project_id
        except (TypeError, ValueError):
            return float("inf"), project_id

    ordered_projects = sorted(project_rows.index.tolist(), key=_project_sort_key)
    # Stacking order: first traces are at the bottom; ensure closed projects sit below active ones.
    ordered_projects.sort(key=lambda pid: 1 if is_active_status(project_rows.loc[pid].get("status")) else 0)
    color_map = project_color_map or {}

    for project_id in ordered_projects:
        row = project_rows.loc[project_id]
        project_start = row.get("start_date")
        if pd.isna(project_start):
            continue
        project_start = pd.Timestamp(project_start)

        status_val = normalize_status_for_display(row.get("status"))
        if status_val.lower() == "closed":
            project_end = row.get("actual_end_date")
            if pd.isna(project_end):
                continue
            project_end = pd.Timestamp(project_end)
        else:
            expected_end = _resolve_expected_end(row)
            project_end = expected_end if expected_end is not None else pd.Timestamp(year_end)

        active = (week_ends >= project_start) & (week_starts <= project_end)
        total_week_start = project_start - pd.Timedelta(days=project_start.weekday())
        total_week_end = project_end - pd.Timedelta(days=project_end.weekday())
        total_weeks = pd.date_range(total_week_start, total_week_end, freq="W-MON")
        total_active_weeks = len(total_weeks)
        if total_active_weeks == 0:
            continue

        hours_cap = _resolve_hours_cap(row)
        magnitude_value = row.get("estimated_magnitude")
        magnitude_weight = float(estimate_magnitude_weight(magnitude_value))
        use_hours_cap = hours_cap is not None and hours_cap > 0
        weight = float(hours_cap) if use_hours_cap else magnitude_weight
        per_week = weight / total_active_weeks
        y_vals = [per_week if is_active else 0.0 for is_active in active.tolist()]

        base_color = color_map.get(project_id, BASE_BLACK)
        marker_color, opacity = marker_style_for_status(base_color, status_val)
        hover_text = build_hover_text(
            row,
            extra={
                "weight_source": "hours_cap" if use_hours_cap else "estimated_magnitude",
                "hours_cap": hours_cap if hours_cap is not None else "",
                "estimated_magnitude": magnitude_value,
                "estimated_magnitude_weight": f"{magnitude_weight:.2f}",
                "weight": f"{weight:.2f}",
                "resolved_end_date": resolve_end_date_for_hover(row),
            },
        )

        fig.add_trace(
            go.Bar(
                x=x_positions,
                y=y_vals,
                name=project_id,
                width=[bar_width_ms] * len(y_vals),
                marker_color=marker_color,
                opacity=opacity,
                hovertext=[hover_text] * len(y_vals),
                hovertemplate="%{hovertext}<br>Week starting %{x|%Y-%m-%d}<br>Weight=%{y:.2f}<extra></extra>",
                showlegend=False,
            ),
            row=subplot_row,
            col=1,
        )

    fig.update_yaxes(title_text="Estimated magnitude", row=subplot_row, col=1)
    fig.update_xaxes(title_text="Week", row=subplot_row, col=1, type="date")
    fig.add_annotation(
        text=f"<b>{title}</b>",
        x=0,
        xref="x domain",
        y=1.18,
        yref=axis_domain_ref("y", subplot_row),
        showarrow=False,
        align="left",
        row=subplot_row,
        col=1,
    )


def add_reported_hours_per_project(
    fig: go.Figure,
    projects_df: pd.DataFrame,
    time_entries_df: pd.DataFrame,
    subplot_row: int,
    title: str,
    project_color_map: Optional[Dict[str, str]] = None,
) -> None:
    if time_entries_df.empty or "duration_hours" not in time_entries_df.columns:
        fig.add_trace(
            go.Bar(x=["(no reported hours)"], y=[0], hovertemplate="No reported hours data.<extra></extra>", showlegend=False),
            row=subplot_row,
            col=1,
        )
        fig.update_yaxes(title_text="Hours", row=subplot_row, col=1)
        fig.add_annotation(
            text=f"<b>{title}</b>",
            x=0,
            xref="x domain",
            y=1.18,
            yref=axis_domain_ref("y", subplot_row),
            showarrow=False,
            align="left",
            row=subplot_row,
            col=1,
        )
        return

    hours_by_project = (
        time_entries_df.groupby("project_id", as_index=False)["duration_hours"]
        .sum()
        .rename(columns={"duration_hours": "total_hours"})
    )
    if hours_by_project.empty:
        fig.add_trace(
            go.Bar(x=["(no reported hours)"], y=[0], hovertemplate="No reported hours data.<extra></extra>", showlegend=False),
            row=subplot_row,
            col=1,
        )
        fig.update_yaxes(title_text="Hours", row=subplot_row, col=1)
        fig.add_annotation(
            text=f"<b>{title}</b>",
            x=0,
            xref="x domain",
            y=1.18,
            yref=axis_domain_ref("y", subplot_row),
            showarrow=False,
            align="left",
            row=subplot_row,
            col=1,
        )
        return

    hours_by_project["project_id"] = hours_by_project["project_id"].astype(str)
    project_rows = projects_df.copy()
    project_rows["project_id"] = project_rows["project_id"].astype(str)
    merged = project_rows.merge(hours_by_project, on="project_id", how="right")
    merged = merged[merged["total_hours"] > 0].copy()

    if merged.empty:
        fig.add_trace(
            go.Bar(x=["(no reported hours)"], y=[0], hovertemplate="No reported hours data.<extra></extra>", showlegend=False),
            row=subplot_row,
            col=1,
        )
        fig.update_yaxes(title_text="Hours", row=subplot_row, col=1)
        fig.add_annotation(
            text=f"<b>{title}</b>",
            x=0,
            xref="x domain",
            y=1.18,
            yref=axis_domain_ref("y", subplot_row),
            showarrow=False,
            align="left",
            row=subplot_row,
            col=1,
        )
        return

    merged = merged.sort_values("total_hours", ascending=False)
    labels: List[str] = []
    colors: List[str] = []
    hovers: List[str] = []
    hours: List[float] = []
    color_map = project_color_map or {}

    for _, row in merged.iterrows():
        project_id = str(row.get("project_id", "")).strip()
        project_name = str(row.get("project_name", project_id)).strip()
        label = f"{project_name} ({project_id})" if project_name else project_id
        labels.append(label)
        hours.append(float(row.get("total_hours", 0.0)))
        base_color = color_map.get(project_id, BASE_BLACK)
        marker_color, opacity = marker_style_for_status(base_color, row.get("status"))
        colors.append(hex_to_rgba(marker_color, opacity) if opacity < 1.0 else marker_color)
        hovers.append(
            build_hover_text(
                row,
                extra={
                    "total_hours": f"{row.get('total_hours', 0.0):.2f}",
                    "resolved_end_date": resolve_end_date_for_hover(row),
                },
            )
        )

    fig.add_trace(
        go.Bar(
            x=labels,
            y=hours,
            marker_color=colors,
            hovertext=hovers,
            hovertemplate="%{hovertext}<br>Total hours=%{y:.2f}<extra></extra>",
            showlegend=False,
        ),
        row=subplot_row,
        col=1,
    )
    fig.update_xaxes(categoryorder="array", categoryarray=labels, row=subplot_row, col=1)
    fig.update_yaxes(title_text="Hours", row=subplot_row, col=1)
    fig.add_annotation(
        text=f"<b>{title}</b>",
        x=0,
        xref="x domain",
        y=1.18,
        yref=axis_domain_ref("y", subplot_row),
        showarrow=False,
        align="left",
        row=subplot_row,
        col=1,
    )


def add_nn_summary_bars(
    fig: go.Figure,
    nn_summary: Optional[Dict[str, Any]],
    subplot_row: int,
    title: str,
) -> None:
    if not nn_summary or nn_summary.get("billed") is None or nn_summary.get("remaining") is None:
        fig.add_trace(
            go.Bar(
                x=[f"{_company_label_short()} summary unavailable"],
                y=[0],
                hovertemplate=f"No {_company_label_short()} summary available.<extra></extra>",
                showlegend=False,
            ),
            row=subplot_row,
            col=1,
        )
        fig.update_yaxes(title_text="Hours", row=subplot_row, col=1)
        fig.add_annotation(
            text=f"<b>{title}</b>",
            x=0,
            xref="x domain",
            y=1.18,
            yref=axis_domain_ref("y", subplot_row),
            showarrow=False,
            align="left",
            row=subplot_row,
            col=1,
        )
        return

    billed = float(nn_summary.get("billed", 0.0))
    remaining = float(nn_summary.get("remaining", 0.0))
    fig.add_trace(
        go.Bar(x=["Gefactureerd"], y=[billed], marker_color=BASE_BLUE, showlegend=False),
        row=subplot_row,
        col=1,
    )
    fig.add_trace(
        go.Bar(x=["Resterend"], y=[remaining], marker_color=BASE_YELLOW, showlegend=False),
        row=subplot_row,
        col=1,
    )
    fig.update_yaxes(title_text="Hours", row=subplot_row, col=1)
    fig.add_annotation(
        text=f"<b>{title}</b>",
        x=0,
        xref="x domain",
        y=1.18,
        yref=axis_domain_ref("y", subplot_row),
        showarrow=False,
        align="left",
        row=subplot_row,
        col=1,
    )


def build_counts_figure(
    projects_df: pd.DataFrame,
    export_date: str,
    period_start: date,
    period_end: date,
    period_label: str,
    project_color_map: Optional[Dict[str, str]] = None,
    timeline_projects_df: Optional[pd.DataFrame] = None,
    timeline_year: Optional[int] = None,
) -> go.Figure:
    total_rows = 4
    if project_color_map is None:
        _, project_color_map = build_color_maps(projects_df)
    if timeline_projects_df is None:
        timeline_projects_df = projects_df
    fig = make_subplots(rows=total_rows, cols=1, shared_xaxes=False, vertical_spacing=0.10)

    add_stacked_project_count_bars(fig, projects_df, "programma", 1,
                                   "Projects per programma (stacked: each project = 1 block)",
                                   project_color_map)
    add_stacked_project_count_bars(fig, projects_df, "theme", 2,
                                   "Projects per theme (stacked: each project = 1 block)",
                                   project_color_map)
    add_stacked_project_count_bars(fig, projects_df, "requester", 3,
                                   "Projects per requester (stacked: each project = 1 block)",
                                   project_color_map)
    add_trend_started_closed(
        fig,
        timeline_projects_df,
        4,
        "Active projects per week (stacked by project)",
        project_color_map,
        target_year=timeline_year,
    )

    apply_axis_style(fig, total_rows)
    fig.update_layout(
        barmode="stack",
        height=1600,
        margin=dict(l=60, r=60, t=40, b=60),
        plot_bgcolor="rgba(255,255,255,1)",
        paper_bgcolor="rgba(250,250,250,1)",
        hoverlabel=dict(namelength=-1),
        showlegend=False,
    )
    return fig


def build_hours_figure(
    projects_df: pd.DataFrame,
    time_entries_df_filtered: pd.DataFrame,
    export_date: str,
    period_start: date,
    period_end: date,
    period_label: str,
    report_type: str,
    exclude_project_ids: Optional[Set[str]] = None,
) -> go.Figure:
    if report_type in ("daily", "weekly", "biweekly"):
        total_rows = 6
        separator_row = 2
        row_heights = [0.21, 0.05, 0.19, 0.19, 0.19, 0.17]
        deep_dive_start_row = 3
        activitytype_row = 6
    else:
        total_rows = 8
        separator_row = 4
        row_heights = [0.13, 0.13, 0.13, 0.05, 0.13, 0.13, 0.13, 0.17]
        deep_dive_start_row = 5
        activitytype_row = 8
    _, project_color_map = build_color_maps(projects_df)
    projects_df_for_breakdown = filter_projects_excluding_project_ids(projects_df, exclude_project_ids)
    time_entries_df_for_breakdown = filter_time_entries_excluding_project_ids(
        time_entries_df_filtered,
        exclude_project_ids,
    )
    vertical_spacing = 0.10 if report_type in ("daily", "weekly", "biweekly") else 0.09
    specs = [[{"type": "xy"}] for _ in range(total_rows)]
    specs[activitytype_row - 1] = [{"type": "domain"}]
    fig = make_subplots(
        rows=total_rows,
        cols=1,
        shared_xaxes=False,
        vertical_spacing=vertical_spacing,
        row_heights=row_heights,
        specs=specs,
    )

    display_start = period_start
    display_end = period_end
    if report_type == "yearly":
        display_end = date(period_start.year, 12, 31)

    if report_type in ("daily", "weekly", "biweekly"):
        add_reported_hours_per_project(
            fig,
            projects_df_for_breakdown,
            time_entries_df_for_breakdown,
            1,
            "Billed hours per project",
            project_color_map,
        )
        add_stacked_hours_bars(
            fig,
            projects_df_for_breakdown,
            time_entries_df_for_breakdown,
            "programma",
            deep_dive_start_row,
            "Hours per programma (stacked: each project contributes its hours)",
            project_color_map,
            section_title="Deep-dive",
        )
        add_stacked_hours_bars(
            fig,
            projects_df_for_breakdown,
            time_entries_df_for_breakdown,
            "theme",
            4,
            "Hours per theme (stacked: each project contributes its hours)",
            project_color_map,
        )
        add_stacked_hours_bars(
            fig,
            projects_df_for_breakdown,
            time_entries_df_for_breakdown,
            "requester",
            5,
            "Hours per requester (stacked: each project contributes its hours)",
            project_color_map,
        )
        add_activitytype_weighted_pie(
            fig,
            time_entries_df_for_breakdown,
            activitytype_row,
            "ActivityType* share (weighted by hours)",
        )
    else:
        add_reported_hours_per_project(
            fig,
            projects_df_for_breakdown,
            time_entries_df_for_breakdown,
            1,
            "Billed hours per project",
            project_color_map,
        )
        add_hours_per_week(
            fig,
            projects_df,
            time_entries_df_filtered,
            2,
            "Billed hours per week (stacked by project)",
            project_color_map,
            display_start=display_start,
            display_end=display_end,
            data_start=period_start,
            data_end=period_end,
        )
        add_estimated_magnitude_per_week(
            fig,
            projects_df,
            3,
            "Estimated magnitude per week (stacked by project)",
            project_color_map,
            period_start=display_start,
            period_end=display_end,
        )
        add_stacked_hours_bars(
            fig,
            projects_df_for_breakdown,
            time_entries_df_for_breakdown,
            "programma",
            deep_dive_start_row,
            "Hours per programma (stacked: each project contributes its hours)",
            project_color_map,
            section_title="Deep-dive",
        )
        add_stacked_hours_bars(
            fig,
            projects_df_for_breakdown,
            time_entries_df_for_breakdown,
            "theme",
            6,
            "Hours per theme (stacked: each project contributes its hours)",
            project_color_map,
        )
        add_stacked_hours_bars(
            fig,
            projects_df_for_breakdown,
            time_entries_df_for_breakdown,
            "requester",
            7,
            "Hours per requester (stacked: each project contributes its hours)",
            project_color_map,
        )
        add_activitytype_weighted_pie(
            fig,
            time_entries_df_for_breakdown,
            activitytype_row,
            "ActivityType* share (weighted by hours)",
        )

    apply_axis_style(fig, total_rows)

    # fig.add_annotation(
    #     text="<b>Deep-dive</b>",
    #     x=0,
    #     xref=axis_domain_ref("x", separator_row),
    #     y=0.5,
    #     yref=axis_domain_ref("y", separator_row),
    #     showarrow=False,
    #     align="left",
    #     font=dict(size=22, color=BASE_BLUE),
    #     row=separator_row,
    #     col=1,
    # )
    fig.update_xaxes(visible=False, row=separator_row, col=1)
    fig.update_yaxes(visible=False, row=separator_row, col=1)

    fig.update_layout(
        barmode="stack",
        height=1900 if report_type in ("weekly", "biweekly") else 2600,
        margin=dict(l=60, r=60, t=40, b=60),
        plot_bgcolor="rgba(255,255,255,1)",
        paper_bgcolor="rgba(250,250,250,1)",
        hoverlabel=dict(namelength=-1),
        showlegend=False,
    )
    return fig


def _apply_standard_hours_layout(fig: go.Figure, height: int) -> None:
    fig.update_layout(
        barmode="stack",
        height=height,
        margin=dict(l=60, r=60, t=40, b=60),
        plot_bgcolor="rgba(255,255,255,1)",
        paper_bgcolor="rgba(250,250,250,1)",
        hoverlabel=dict(namelength=-1),
        showlegend=False,
    )


def build_hours_section_figures(
    projects_df: pd.DataFrame,
    time_entries_df_filtered: pd.DataFrame,
    period_start: date,
    period_end: date,
    report_type: str,
    exclude_project_ids: Optional[Set[str]] = None,
) -> Dict[str, go.Figure]:
    _, project_color_map = build_color_maps(projects_df)
    projects_df_for_breakdown = filter_projects_excluding_project_ids(projects_df, exclude_project_ids)
    time_entries_df_for_breakdown = filter_time_entries_excluding_project_ids(
        time_entries_df_filtered,
        exclude_project_ids,
    )
    display_start = period_start
    display_end = period_end
    if report_type == "yearly":
        display_end = date(period_start.year, 12, 31)

    section_figs: Dict[str, go.Figure] = {}

    primary_fig = make_subplots(rows=1, cols=1, shared_xaxes=False)
    add_reported_hours_per_project(
        primary_fig,
        projects_df_for_breakdown,
        time_entries_df_for_breakdown,
        1,
        "Billed hours per project",
        project_color_map,
    )
    apply_axis_style(primary_fig, 1)
    _apply_standard_hours_layout(primary_fig, height=520)
    section_figs["primary"] = primary_fig

    if report_type in ("monthly", "yearly"):
        timeline_fig = make_subplots(rows=2, cols=1, shared_xaxes=False, vertical_spacing=0.12)
        add_hours_per_week(
            timeline_fig,
            projects_df,
            time_entries_df_filtered,
            1,
            "Billed hours per week (stacked by project)",
            project_color_map,
            display_start=display_start,
            display_end=display_end,
            data_start=period_start,
            data_end=period_end,
        )
        add_estimated_magnitude_per_week(
            timeline_fig,
            projects_df,
            2,
            "Estimated magnitude per week (stacked by project)",
            project_color_map,
            period_start=display_start,
            period_end=display_end,
        )
        apply_axis_style(timeline_fig, 2)
        _apply_standard_hours_layout(timeline_fig, height=980)
        section_figs["timeline"] = timeline_fig

    deep_dive_fig = make_subplots(
        rows=4,
        cols=1,
        shared_xaxes=False,
        vertical_spacing=0.10,
        row_heights=[0.26, 0.26, 0.26, 0.22],
        specs=[[{"type": "xy"}], [{"type": "xy"}], [{"type": "xy"}], [{"type": "domain"}]],
    )
    add_stacked_hours_bars(
        deep_dive_fig,
        projects_df_for_breakdown,
        time_entries_df_for_breakdown,
        "programma",
        1,
        "Hours per programma (stacked: each project contributes its hours)",
        project_color_map,
        section_title="Deep-dive",
    )
    add_stacked_hours_bars(
        deep_dive_fig,
        projects_df_for_breakdown,
        time_entries_df_for_breakdown,
        "theme",
        2,
        "Hours per theme (stacked: each project contributes its hours)",
        project_color_map,
    )
    add_stacked_hours_bars(
        deep_dive_fig,
        projects_df_for_breakdown,
        time_entries_df_for_breakdown,
        "requester",
        3,
        "Hours per requester (stacked: each project contributes its hours)",
        project_color_map,
    )
    add_activitytype_weighted_pie(
        deep_dive_fig,
        time_entries_df_for_breakdown,
        4,
        "ActivityType* share (weighted by hours)",
    )
    apply_axis_style(deep_dive_fig, 4)
    _apply_standard_hours_layout(deep_dive_fig, height=1720)
    section_figs["deep_dive"] = deep_dive_fig

    return section_figs


def build_percentage_section_payloads(
    section_figs: Dict[str, go.Figure],
    total_period_hours: float,
    weekly_reference_hours: Optional[float],
    weekly_progress_guidance: Optional[Dict[str, Any]],
    timeline_explanation_html: str,
) -> List[Dict[str, Any]]:
    sections: List[Dict[str, Any]] = []

    primary_source = section_figs.get("primary")
    if primary_source is not None:
        primary_pct = build_percentage_figure_from_hours(
            primary_source,
            total_period_hours=total_period_hours,
            weekly_reference_hours=weekly_reference_hours,
            weekly_progress_guidance=weekly_progress_guidance,
        )
        sections.append(
            {
                "section_id": "reported",
                "title": "Time-distribution across projects",
                "figure": primary_pct,
                "foldable": False,
                "open": True,
                "extra_html": "",
            }
        )

    timeline_source = section_figs.get("timeline")
    if timeline_source is not None:
        timeline_pct = build_percentage_figure_from_hours(
            timeline_source,
            total_period_hours=total_period_hours,
            weekly_reference_hours=weekly_reference_hours,
            weekly_progress_guidance=weekly_progress_guidance,
        )
        _ensure_subplot_title_annotation(
            timeline_pct,
            1,
            "Billed hours per week (stacked by project)",
        )
        sections.append(
            {
                "section_id": "timeline",
                "title": "Timeline",
                "figure": timeline_pct,
                "foldable": True,
                "open": False,
                "extra_html": timeline_explanation_html or "",
            }
        )

    deep_dive_source = section_figs.get("deep_dive")
    if deep_dive_source is not None:
        deep_dive_pct = build_percentage_figure_from_hours(
            deep_dive_source,
            total_period_hours=total_period_hours,
            weekly_reference_hours=weekly_reference_hours,
            weekly_progress_guidance=weekly_progress_guidance,
        )
        _ensure_subplot_title_annotation(
            deep_dive_pct,
            1,
            "Hours per programma (stacked: each project contributes its hours)",
        )
        sections.append(
            {
                "section_id": "deep-dive",
                "title": "Deep-dive",
                "figure": deep_dive_pct,
                "foldable": True,
                "open": False,
                "extra_html": "",
            }
        )

    return sections


def _to_float_list(values: Any) -> List[float]:
    out: List[float] = []
    if values is None:
        return out
    for v in list(values):
        try:
            out.append(float(v))
        except (TypeError, ValueError):
            out.append(0.0)
    return out


def _layout_axis_key(axis_id: str, axis_letter: str) -> str:
    if not axis_id:
        return f"{axis_letter}axis"
    suffix = axis_id[1:]  # "y" -> "", "y2" -> "2"
    return f"{axis_letter}axis{suffix}"


def _axis_title_text(axis_obj: Any) -> str:
    title = getattr(axis_obj, "title", None)
    if title is None:
        return ""
    if isinstance(title, dict):
        text = title.get("text")
    else:
        text = getattr(title, "text", None)
    return str(text).strip() if text is not None else ""


def _is_date_xaxis(fig: go.Figure, xaxis_id: str) -> bool:
    layout_key = _layout_axis_key(xaxis_id, "x")
    axis_obj = getattr(fig.layout, layout_key, None)
    axis_type = getattr(axis_obj, "type", None) if axis_obj is not None else None
    return str(axis_type).lower() == "date"


def _max_stacked_y_for_axis(fig: go.Figure, yaxis_id: str) -> float:
    sums_by_x: Dict[Any, float] = {}
    for trace in fig.data:
        if getattr(trace, "type", None) != "bar":
            continue
        trace_yaxis = getattr(trace, "yaxis", None) or "y"
        if trace_yaxis != yaxis_id:
            continue
        xs = list(getattr(trace, "x", None) or [])
        ys = _to_float_list(getattr(trace, "y", None))
        for x, y in zip(xs, ys):
            sums_by_x[x] = sums_by_x.get(x, 0.0) + float(y or 0.0)
    return max(sums_by_x.values()) if sums_by_x else 0.0


def compute_weekly_reference_hours(
    all_time_entries_df: pd.DataFrame,
    period_end: date,
    nn_summary: Optional[Dict[str, Any]] = None,
) -> Tuple[Optional[float], str, Optional[str]]:
    """
    Preferred order:
      1) config.hours.workable_hours_per_week_reference_value
      2) yearly total hours (nn_total_hours) / 46 (when available)
      3) average reported hours per week so far
    Returns (reference_hours, source_key, title_note_or_none).
    """
    if WORKABLE_HOURS_PER_WEEK_REFERENCE_VALUE and WORKABLE_HOURS_PER_WEEK_REFERENCE_VALUE > 0:
        ref = float(WORKABLE_HOURS_PER_WEEK_REFERENCE_VALUE)
        note = f"Percent base: workable_hours_per_week_reference_value = {ref:.1f} h/week"
        return ref, "config_week_reference", note

    yearly_total = _to_float((nn_summary or {}).get("nn_total_hours"))
    if yearly_total is not None and yearly_total > 0:
        ref = float(yearly_total / 46.0)
        note = f"Percent base: yearly total hours / 46 = {yearly_total:.0f}/46 = {ref:.1f} h/week"
        return ref, "year_total_div_46", note

    yearly_available = _to_float((nn_summary or {}).get("year_available_hours"))
    if yearly_available is not None and yearly_available > 0:
        ref = float(yearly_available / 46.0)
        note = (
            "Percent base: yearly available hours / 46 "
            f"(fallback) = {yearly_available:.0f}/46 = {ref:.1f} h/week"
        )
        return ref, "year_capacity_div_46", note

    if (
        all_time_entries_df is None
        or all_time_entries_df.empty
        or "date" not in all_time_entries_df.columns
        or "duration_hours" not in all_time_entries_df.columns
    ):
        return None, "none", None

    entries = all_time_entries_df.copy()
    entries["date"] = pd.to_datetime(entries["date"], errors="coerce")
    entries["duration_hours"] = pd.to_numeric(entries["duration_hours"], errors="coerce").fillna(0.0)
    entries = entries.dropna(subset=["date"])
    if entries.empty:
        return None, "none", None

    year_start = date(period_end.year, 1, 1)
    entries = entries.loc[(entries["date"] >= pd.Timestamp(year_start)) & (entries["date"] <= pd.Timestamp(period_end))]
    if entries.empty:
        return None, "none", None

    total_hours = float(entries["duration_hours"].sum())
    elapsed_days = max((period_end - year_start).days + 1, 1)
    elapsed_weeks = max(elapsed_days / 7.0, 1.0)
    ref = total_hours / elapsed_weeks
    if ref <= 0:
        return None, "none", None
    return ref, "avg_reported_so_far", None


def compute_weekly_progress_guidance(
    all_time_entries_df: pd.DataFrame,
    period_end: date,
    nn_summary: Optional[Dict[str, Any]],
    weekly_reference_hours: Optional[float],
) -> Optional[Dict[str, Any]]:
    """
    Guidance lines for "Reported time per week" percentage chart:
      - Avg percent worked so far (from year start through current week)
      - Avg percent needed for remaining year (from next week through year end)
    """
    ref = _to_float(weekly_reference_hours)
    if ref is None or ref <= 0:
        return None

    summary = nn_summary or {}
    yearly_total = _to_float(summary.get("nn_total_hours"))
    if yearly_total is None or yearly_total <= 0:
        return None

    yearly_available = _to_float(summary.get("year_available_hours"))
    remaining_from_summary = _to_float(summary.get("remaining"))

    year_start = date(period_end.year, 1, 1)
    year_end = date(period_end.year, 12, 31)
    year_start_week = year_start - timedelta(days=year_start.weekday())
    current_week_start = period_end - timedelta(days=period_end.weekday())
    current_week_end = min(current_week_start + timedelta(days=6), year_end)
    next_week_start = current_week_start + timedelta(days=7)
    year_end_week_start = year_end - timedelta(days=year_end.weekday())

    reported_so_far = 0.0
    if (
        all_time_entries_df is not None
        and not all_time_entries_df.empty
        and "date" in all_time_entries_df.columns
        and "duration_hours" in all_time_entries_df.columns
    ):
        entries = all_time_entries_df.copy()
        entries["date"] = pd.to_datetime(entries["date"], errors="coerce")
        entries["duration_hours"] = pd.to_numeric(entries["duration_hours"], errors="coerce").fillna(0.0)
        entries = entries.dropna(subset=["date"])
        if not entries.empty:
            entries = entries.loc[
                (entries["date"] >= pd.Timestamp(year_start))
                & (entries["date"] <= pd.Timestamp(period_end))
            ]
            reported_so_far = float(entries["duration_hours"].sum())

    billed_so_far: Optional[float] = None
    if remaining_from_summary is not None:
        billed_so_far = max(float(yearly_total) - float(remaining_from_summary), 0.0)

    tracked_source = "reported"
    tracked_so_far = reported_so_far
    if billed_so_far is not None:
        tracked_source = "billed"
        tracked_so_far = billed_so_far

    elapsed_weeks = max(((current_week_start - year_start_week).days // 7) + 1, 1)
    avg_so_far_hours = tracked_so_far / float(elapsed_weeks)
    worked_pct = avg_so_far_hours * 100.0 / ref

    remaining_hours = (
        max(float(remaining_from_summary), 0.0)
        if remaining_from_summary is not None
        else max(float(yearly_total) - tracked_so_far, 0.0)
    )
    weeks_remaining = 0
    remaining_pct: Optional[float] = None
    avg_remaining_hours: Optional[float] = None
    if next_week_start <= year_end_week_start:
        weeks_remaining = ((year_end_week_start - next_week_start).days // 7) + 1
        if weeks_remaining > 0:
            avg_remaining_hours = remaining_hours / float(weeks_remaining)
            remaining_pct = avg_remaining_hours * 100.0 / ref

    return {
        "reference_hours_per_week": ref,
        "yearly_total_hours": yearly_total,
        "yearly_available_hours": yearly_available,
        "reported_hours_so_far": reported_so_far,
        "billed_hours_so_far": billed_so_far,
        "tracked_source": tracked_source,
        "tracked_hours_so_far": tracked_so_far,
        "elapsed_weeks": elapsed_weeks,
        "worked_pct": worked_pct,
        "worked_x0": year_start_week,
        "worked_x1": current_week_end,
        "worked_hours_per_week": avg_so_far_hours,
        "remaining_hours": remaining_hours,
        "weeks_remaining": weeks_remaining,
        "remaining_pct": remaining_pct,
        "remaining_x0": next_week_start if weeks_remaining > 0 else None,
        "remaining_x1": year_end,
        "remaining_hours_per_week": avg_remaining_hours,
    }


def compute_monthly_average_guidance(
    period_time_entries_df: pd.DataFrame,
    period_start: date,
    period_end: date,
    weekly_reference_hours: Optional[float],
    nn_summary: Optional[Dict[str, Any]] = None,
) -> Optional[Dict[str, Any]]:
    """
    Month tab guidance:
      - Summary values for worked/billed pace in the selected period.
    """
    ref = _to_float(weekly_reference_hours)
    if ref is None or ref <= 0:
        return None

    total_reported_hours = 0.0
    if (
        period_time_entries_df is not None
        and not period_time_entries_df.empty
        and "duration_hours" in period_time_entries_df.columns
    ):
        total_reported_hours = float(
            pd.to_numeric(period_time_entries_df["duration_hours"], errors="coerce").fillna(0.0).sum()
        )

    billed_period_hours = _to_float((nn_summary or {}).get("billed"))
    tracked_source = "reported"
    tracked_hours = total_reported_hours
    if billed_period_hours is not None and billed_period_hours >= 0:
        tracked_source = "billed"
        tracked_hours = float(billed_period_hours)

    week_starts, _, _, _ = build_period_week_grid(period_start, period_end)
    weeks_in_period = len(week_starts)
    if weeks_in_period <= 0:
        return None

    avg_hpw = tracked_hours / float(weeks_in_period)
    worked_pct = avg_hpw * 100.0 / ref

    return {
        "reference_hours_per_week": ref,
        "tracked_source": tracked_source,
        "tracked_hours_so_far": tracked_hours,
        "reported_hours_so_far": total_reported_hours,
        "billed_hours_so_far": billed_period_hours,
        "elapsed_weeks": weeks_in_period,
        "worked_pct": worked_pct,
        "worked_x0": None,
        "worked_x1": period_end,
        "worked_hours_per_week": avg_hpw,
        "remaining_pct": None,
        "remaining_x0": None,
        "remaining_x1": None,
        "remaining_hours_per_week": None,
        "remaining_hours": None,
        "weeks_remaining": 0,
        "show_worked_line": False,
    }


def build_monthly_average_explanation_html(
    weekly_progress_guidance: Optional[Dict[str, Any]],
) -> str:
    if not weekly_progress_guidance:
        return ""

    ref = _to_float(weekly_progress_guidance.get("reference_hours_per_week"))
    tracked_hours = _to_float(weekly_progress_guidance.get("tracked_hours_so_far"))
    elapsed_weeks = int(weekly_progress_guidance.get("elapsed_weeks") or 0)
    worked_hpw = _to_float(weekly_progress_guidance.get("worked_hours_per_week"))
    worked_pct = _to_float(weekly_progress_guidance.get("worked_pct"))
    if ref is None or ref <= 0 or tracked_hours is None or elapsed_weeks <= 0 or worked_hpw is None or worked_pct is None:
        return ""

    summary_sentence = (
        f"During this period, {tracked_hours:.1f} hours have been billed in {elapsed_weeks} weeks, "
        f"averaging {worked_hpw:.1f} h/week ({worked_pct:.1f} % of the assumed {ref:.1f} h/week)."
    )

    return "<div class='weekly-guidance'><div>" + html.escape(summary_sentence) + "</div></div>"


def build_weekly_progress_explanation_html(
    weekly_progress_guidance: Optional[Dict[str, Any]],
    weekly_reference_source: str,
) -> str:
    if not weekly_progress_guidance:
        return ""

    ref = _to_float(weekly_progress_guidance.get("reference_hours_per_week"))
    if ref is None or ref <= 0:
        return ""

    yearly_total = _to_float(weekly_progress_guidance.get("yearly_total_hours")) or 0.0
    yearly_available = _to_float(weekly_progress_guidance.get("yearly_available_hours")) or 0.0
    reported_so_far = _to_float(weekly_progress_guidance.get("reported_hours_so_far")) or 0.0
    tracked_so_far = _to_float(weekly_progress_guidance.get("tracked_hours_so_far")) or reported_so_far
    tracked_source = str(weekly_progress_guidance.get("tracked_source", "reported")).strip().lower()
    elapsed_weeks = int(weekly_progress_guidance.get("elapsed_weeks") or 0)
    worked_hpw = _to_float(weekly_progress_guidance.get("worked_hours_per_week")) or 0.0
    worked_pct = _to_float(weekly_progress_guidance.get("worked_pct")) or 0.0
    remaining_hours = _to_float(weekly_progress_guidance.get("remaining_hours")) or 0.0
    weeks_remaining = int(weekly_progress_guidance.get("weeks_remaining") or 0)
    remaining_hpw = _to_float(weekly_progress_guidance.get("remaining_hours_per_week"))
    remaining_pct = _to_float(weekly_progress_guidance.get("remaining_pct"))

    if weekly_reference_source == "config_week_reference":
        basis_formula = (
            f"100% = config.hours.workable_hours_per_week_reference_value = {ref:.1f} h/week."
        )
    elif weekly_reference_source == "year_total_div_46":
        basis_formula = (
            f"100% = yearly total hours / 46 = {yearly_total:.0f}/46 = {ref:.1f} h/week."
        )
    elif weekly_reference_source == "year_capacity_div_46":
        basis_formula = (
            f"100% = yearly available hours / 46 = {yearly_available:.0f}/46 = {ref:.1f} h/week (fallback)."
        )
    else:
        elapsed_for_formula = max(elapsed_weeks, 1)
        basis_formula = (
            f"100% = average reported hours so far = {reported_so_far:.1f}/{elapsed_for_formula} = {ref:.1f} h/week."
        )

    if tracked_source == "billed":
        progress_sentence = (
            f"Up until today, {tracked_so_far:.1f} hours have been billed in {elapsed_weeks} weeks, "
            f"averaging {worked_hpw:.1f} h/week ({worked_pct:.1f}%) [horizontal dotted line]."
        )
    else:
        progress_sentence = (
            f"Up until today, {tracked_so_far:.1f} hours have been reported in {elapsed_weeks} weeks, "
            f"averaging {worked_hpw:.1f} h/week ({worked_pct:.1f}%) [horizontal dotted line]."
        )

    if remaining_pct is not None and remaining_hpw is not None and weeks_remaining > 0:
        remaining_sentence = (
            f"Spreading the remaining {remaining_hours:.1f} hours over the remaining {weeks_remaining} weeks, "
            f"the projection is {remaining_hpw:.1f} h/week ({remaining_pct:.1f}%) for the remainder of the year "
            "[horizontal dashed line]."
        )
        if worked_pct > remaining_pct:
            comparison_sentence = (
                f"Given that {worked_pct:.1f}% (up until now) is larger than {remaining_pct:.1f}% "
                "(spread over remaining weeks), the remainder of the year can run at a lower weekly pace "
                "than the average pace so far."
            )
        elif worked_pct < remaining_pct:
            comparison_sentence = (
                f"Given that {worked_pct:.1f}% (up until now) is smaller than {remaining_pct:.1f}% "
                "(spread over remaining weeks), the remainder of the year requires a higher weekly pace "
                "than the average pace so far."
            )
        else:
            comparison_sentence = (
                f"Given that {worked_pct:.1f}% (up until now) equals {remaining_pct:.1f}% "
                "(spread over remaining weeks), the remainder of the year can continue at the same average pace."
            )
    else:
        remaining_sentence = "No remaining weeks are available in the current year for a remainder projection."
        comparison_sentence = "Comparison between worked-so-far and remainder projection percentages is not available."

    return (
        "<div class='weekly-guidance'>"
        "<div><b>The percentage shown is based on the assumption that 100% equals "
        f"{ref:.1f} hours per week.</b></div>"
        f"<div class='weekly-guidance-formula'>{html.escape(basis_formula)}</div>"
        f"<div>{html.escape(progress_sentence)}</div>"
        f"<div>{html.escape(remaining_sentence)}</div>"
        f"<div>{html.escape(comparison_sentence)}</div>"
        "</div>"
    )


def _append_weekly_reference_note_to_titles(fig: go.Figure, note_text: str) -> None:
    if not note_text:
        return
    for ann in list(getattr(fig.layout, "annotations", []) or []):
        text = str(getattr(ann, "text", "") or "")
        if (
            "Billed hours per week" in text
            or "Reported time per week" in text
            or "Estimated magnitude per week" in text
        ):
            if "Percent base:" in text:
                continue
            ann.text = (
                text
                + "<br><span style='font-size:11px;font-weight:500'>"
                + html.escape(note_text)
                + "</span>"
            )


def _ensure_subplot_title_annotation(fig: go.Figure, subplot_row: int, title_text: str) -> None:
    if not title_text:
        return
    for ann in list(getattr(fig.layout, "annotations", []) or []):
        text = str(getattr(ann, "text", "") or "")
        if title_text in text:
            return
    fig.add_annotation(
        text=f"<b>{html.escape(title_text)}</b>",
        x=0,
        xref="x domain",
        y=1.18,
        yref=axis_domain_ref("y", subplot_row),
        showarrow=False,
        align="left",
        row=subplot_row,
        col=1,
    )


def build_percentage_figure_from_hours(
    hours_fig: go.Figure,
    total_period_hours: Optional[float] = None,
    weekly_reference_hours: Optional[float] = None,
    weekly_reference_note: Optional[str] = None,
    show_weekly_reference_note_in_title: bool = False,
    weekly_progress_guidance: Optional[Dict[str, Any]] = None,
) -> go.Figure:
    fig = go.Figure(hours_fig.to_dict())

    totals_by_axis: Dict[str, float] = {}
    xaxis_by_yaxis: Dict[str, str] = {}
    for trace in fig.data:
        if getattr(trace, "type", None) != "bar":
            continue
        axis_id = getattr(trace, "yaxis", None) or "y"
        xaxis_by_yaxis.setdefault(axis_id, getattr(trace, "xaxis", None) or "x")
        y_vals = _to_float_list(getattr(trace, "y", None))
        totals_by_axis[axis_id] = totals_by_axis.get(axis_id, 0.0) + sum(y_vals)

    period_hours = float(total_period_hours) if total_period_hours is not None else None
    weekly_hours_ref = float(weekly_reference_hours) if weekly_reference_hours is not None else None
    denom_by_axis: Dict[str, float] = {}
    axis_role_by_id: Dict[str, str] = {}
    for axis_id, plotted_total in totals_by_axis.items():
        yaxis_layout_key = _layout_axis_key(axis_id, "y")
        axis_obj = getattr(fig.layout, yaxis_layout_key, None)
        title_text = _axis_title_text(axis_obj)
        xaxis_id = xaxis_by_yaxis.get(axis_id, "x")
        if _is_date_xaxis(fig, xaxis_id) and title_text.lower() == "hours":
            axis_role_by_id[axis_id] = "reported_weekly"
        if weekly_hours_ref is not None and weekly_hours_ref > 0 and _is_date_xaxis(fig, xaxis_id):
            denom_by_axis[axis_id] = weekly_hours_ref
        elif period_hours is not None and period_hours > 0 and title_text.lower() == "hours":
            denom_by_axis[axis_id] = period_hours
        else:
            denom_by_axis[axis_id] = plotted_total

    for trace in fig.data:
        if getattr(trace, "type", None) != "bar":
            continue
        axis_id = getattr(trace, "yaxis", None) or "y"
        denom = denom_by_axis.get(axis_id, 0.0)
        if denom <= 0:
            continue

        y_orig = _to_float_list(getattr(trace, "y", None))
        trace.customdata = y_orig
        trace.y = [v * 100.0 / denom for v in y_orig]

        ht = getattr(trace, "hovertemplate", None) or ""
        if ht and "<extra></extra>" in ht:
            body, _ = ht.split("<extra></extra>", 1)
        else:
            body = ht

        if "%{y" in body:
            body = (
                body.replace("Total hours=%{y:.2f}", "Total hours=%{customdata:.2f}")
                .replace("Hours=%{y:.2f}", "Hours=%{customdata:.2f}")
                .replace("Weight=%{y:.2f}", "Weight=%{customdata:.2f}")
            )
        if "Percent=%{y" not in body:
            body = body + "<br>Percent=%{y:.1f}%"
        trace.hovertemplate = body + "<extra></extra>"

    for axis_id in totals_by_axis.keys():
        yaxis_layout_key = _layout_axis_key(axis_id, "y")
        axis_obj = getattr(fig.layout, yaxis_layout_key, None)
        if axis_obj is None:
            continue
        axis_obj.title = dict(text="Percent")
        axis_obj.ticksuffix = "%"
        axis_obj.tickformat = ".0f"
        max_stack = _max_stacked_y_for_axis(fig, axis_id)
        max_target = max_stack
        if weekly_progress_guidance and axis_role_by_id.get(axis_id) == "reported_weekly":
            worked_pct = _to_float(weekly_progress_guidance.get("worked_pct"))
            remaining_pct = _to_float(weekly_progress_guidance.get("remaining_pct"))
            if worked_pct is not None:
                max_target = max(max_target, worked_pct)
            if remaining_pct is not None:
                max_target = max(max_target, remaining_pct)
        axis_obj.range = [0, (max_target * 1.2) if max_target > 0 else 100]

    if weekly_progress_guidance:
        worked_pct = _to_float(weekly_progress_guidance.get("worked_pct"))
        worked_x0 = weekly_progress_guidance.get("worked_x0")
        worked_x1 = weekly_progress_guidance.get("worked_x1")
        remaining_pct = _to_float(weekly_progress_guidance.get("remaining_pct"))
        remaining_x0 = weekly_progress_guidance.get("remaining_x0")
        remaining_x1 = weekly_progress_guidance.get("remaining_x1")
        show_worked_line = bool(weekly_progress_guidance.get("show_worked_line", True))

        for axis_id, role in axis_role_by_id.items():
            if role != "reported_weekly":
                continue
            xaxis_id = xaxis_by_yaxis.get(axis_id, "x")
            if show_worked_line and worked_pct is not None and worked_x0 is not None and worked_x1 is not None:
                fig.add_shape(
                    type="line",
                    x0=pd.Timestamp(worked_x0),
                    x1=pd.Timestamp(worked_x1),
                    y0=worked_pct,
                    y1=worked_pct,
                    xref=xaxis_id,
                    yref=axis_id,
                    line=dict(color=BASE_BLUE, width=2, dash="dot"),
                )
            if (
                remaining_pct is not None
                and remaining_x0 is not None
                and remaining_x1 is not None
                and pd.Timestamp(remaining_x0) <= pd.Timestamp(remaining_x1)
            ):
                fig.add_shape(
                    type="line",
                    x0=pd.Timestamp(remaining_x0),
                    x1=pd.Timestamp(remaining_x1),
                    y0=remaining_pct,
                    y1=remaining_pct,
                    xref=xaxis_id,
                    yref=axis_id,
                    line=dict(color=BASE_RED, width=2, dash="dash"),
                )

    if show_weekly_reference_note_in_title and weekly_reference_note:
        _append_weekly_reference_note_to_titles(fig, weekly_reference_note)
    return fig


def render_plot_sections_html(
    section_payloads: Optional[List[Dict[str, Any]]],
    div_prefix: str,
) -> str:
    if not section_payloads:
        return ""

    blocks: List[str] = []
    for idx, payload in enumerate(section_payloads):
        fig = payload.get("figure")
        if fig is None:
            continue
        title = str(payload.get("title", f"Section {idx + 1}"))
        section_id = str(payload.get("section_id", f"section-{idx + 1}"))
        foldable = bool(payload.get("foldable", True))
        is_open = bool(payload.get("open", False))
        extra_html = str(payload.get("extra_html") or "")

        fig_div_id = f"{div_prefix}-{section_id}"
        fig_html = pio.to_html(fig, include_plotlyjs=False, full_html=False, div_id=fig_div_id)
        body_html = f"<div class='plot-section-body'>{extra_html}{fig_html}</div>"

        if foldable:
            open_attr = " open" if is_open else ""
            blocks.append(
                "<details class='plot-section'" + open_attr + ">"
                f"<summary><span class='plot-section-title'>{html.escape(title)}</span></summary>"
                f"{body_html}"
                "</details>"
            )
        else:
            blocks.append(
                "<section class='plot-section plot-section-fixed'>"
                f"<div class='plot-section-title'>{html.escape(title)}</div>"
                f"{body_html}"
                "</section>"
            )

    if not blocks:
        return ""
    return "<div class='plot-sections'>" + "".join(blocks) + "</div>"


# ----------------------------
# Export
# ----------------------------



WITH_HOURS_SUFFIX = "_with_hours"


def _strip_with_hours_suffix(name: str) -> str:
    return name[:-len(WITH_HOURS_SUFFIX)] if name.endswith(WITH_HOURS_SUFFIX) else name


def _ensure_with_hours_suffix(name: str) -> str:
    return name if name.endswith(WITH_HOURS_SUFFIX) else f"{name}{WITH_HOURS_SUFFIX}"


def export_tabbed_report(
    counts_fig: go.Figure,
    hours_fig: go.Figure,
    percentage_fig: go.Figure,
    output_dir: str,
    output_archive_dir: str,
    base_name: str,
    archive_base_name: str,
    export_date: str,
    header_context: Dict[str, Any],
    tables_html: str,
    hours_metrics_html: str,
    percentage_metrics_html: str,
    percentage_explanation_html: str,
    percentage_section_payloads: Optional[List[Dict[str, Any]]],
    sideways_bar_chart_html: str,
    nn_note: Optional[str],
    nn_summary: Optional[Dict[str, Any]],
) -> Tuple[str, str, str]:
    lite_base_name = _strip_with_hours_suffix(base_name)
    with_hours_base_name = _ensure_with_hours_suffix(lite_base_name)
    lite_archive_base_name = _strip_with_hours_suffix(archive_base_name)
    with_hours_archive_base_name = _ensure_with_hours_suffix(lite_archive_base_name)

    html_path = os.path.join(output_dir, f"{with_hours_base_name}.html")
    png_path = os.path.join(output_dir, f"{with_hours_base_name}.png")
    lite_html_path = os.path.join(output_dir, f"{lite_base_name}.html")

    dated_with_hours_base_name = f"{with_hours_archive_base_name}_generated_{export_date}"
    dated_html_path = os.path.join(output_archive_dir, f"{dated_with_hours_base_name}.html")
    dated_png_path = os.path.join(output_archive_dir, f"{dated_with_hours_base_name}.png")
    dated_lite_html_path = os.path.join(output_archive_dir, f"{lite_archive_base_name}_generated_{export_date}.html")

    write_tabbed_html(
        counts_fig,
        hours_fig,
        percentage_fig,
        html_path,
        header_context,
        tables_html,
        hours_metrics_html,
        percentage_metrics_html,
        percentage_explanation_html,
        percentage_section_payloads,
        sideways_bar_chart_html,
        nn_note,
        nn_summary,
    )
    write_tabbed_html(
        counts_fig,
        hours_fig,
        percentage_fig,
        lite_html_path,
        header_context,
        tables_html,
        hours_metrics_html,
        percentage_metrics_html,
        percentage_explanation_html,
        percentage_section_payloads,
        sideways_bar_chart_html,
        nn_note,
        nn_summary,
        enabled_tabs=("percentage", "projects"),
    )

    counts_fig.write_image(png_path, scale=2)  # requires kaleido

    shutil.copyfile(html_path, dated_html_path)
    shutil.copyfile(png_path, dated_png_path)
    shutil.copyfile(lite_html_path, dated_lite_html_path)

    return html_path, png_path, lite_html_path


def export_multi_period_report(
    period_payloads: Dict[str, Dict[str, Any]],
    output_dir: str,
    output_archive_dir: str,
    base_name: str,
    archive_base_name: str,
    export_date: str,
    header_context: Dict[str, Any],
    tables_html: str,
    projects_filters_html: str = "",
) -> Tuple[str, str]:
    lite_base_name = _strip_with_hours_suffix(base_name)
    with_hours_base_name = _ensure_with_hours_suffix(lite_base_name)
    lite_archive_base_name = _strip_with_hours_suffix(archive_base_name)
    with_hours_archive_base_name = _ensure_with_hours_suffix(lite_archive_base_name)

    html_path = os.path.join(output_dir, f"{with_hours_base_name}.html")
    lite_html_path = os.path.join(output_dir, f"{lite_base_name}.html")
    dated_with_hours_base_name = f"{with_hours_archive_base_name}_generated_{export_date}"
    dated_html_path = os.path.join(output_archive_dir, f"{dated_with_hours_base_name}.html")
    dated_lite_html_path = os.path.join(output_archive_dir, f"{lite_archive_base_name}_generated_{export_date}.html")

    write_multi_period_tabbed_html(
        period_payloads,
        html_path,
        header_context,
        tables_html,
        projects_filters_html=projects_filters_html,
    )
    write_multi_period_tabbed_html(
        period_payloads,
        lite_html_path,
        header_context,
        tables_html,
        projects_filters_html=projects_filters_html,
        enabled_tabs=("percentage", "projects"),
    )

    shutil.copyfile(html_path, dated_html_path)
    shutil.copyfile(lite_html_path, dated_lite_html_path)
    return html_path, lite_html_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate project portfolio reports.")
    parser.add_argument(
        "--report-type",
        choices=REPORT_TYPE_CHOICES,
        default=None,
        help="Report type to generate (overrides runtime.default_report_type in config).",
    )
    parser.add_argument(
        "--asof",
        default=None,
        help="As-of date in YYYY-MM-DD (defaults to today).",
    )
    parser.add_argument(
        "--config",
        default=DEFAULT_CONFIG_PATH,
        help="Path to config JSON (default: projexcellent_config.json).",
    )
    return parser.parse_args()


def parse_asof_date(asof_str: Optional[str]) -> date:
    if not asof_str:
        return date.today()
    try:
        return date.fromisoformat(asof_str)
    except ValueError as exc:
        raise SystemExit(f"Invalid --asof date '{asof_str}' (expected YYYY-MM-DD).") from exc




def main() -> None:
    args = parse_args()
    _apply_runtime_config(args.config)
    asof_date = parse_asof_date(args.asof)
    default_report_type = str(CONFIG.get("runtime", {}).get("default_report_type", "all")).strip().lower()
    report_type = str(args.report_type or default_report_type).strip().lower()
    if report_type not in REPORT_TYPE_CHOICES:
        raise SystemExit(
            f"Invalid report type '{report_type}' in config/runtime.default_report_type "
            f"(allowed: {', '.join(REPORT_TYPE_CHOICES)})."
        )
    generate_reports(report_type, asof_date)

def _verify_and_collect_deliverables(project_folder_path: str, folder_name: str) -> Tuple[str, Dict[str, Any]]:
    """Return (deliverables_dir, deliverables_payload). Raises human-readable error if missing folder."""
    deliverables_dir = os.path.join(project_folder_path, "Deliverables")
    if not os.path.isdir(deliverables_dir):
        raise FileNotFoundError(deliverables_dir)
    payload: Dict[str, Any] = {"texts": [], "images": []}

    # Text files (.txt) excluding readme.txt (case-insensitive)
    try:
        entries = sorted(os.listdir(deliverables_dir))
    except Exception as exc:
        raise RuntimeError(f"Failed to list Deliverables folder for '{folder_name}': {deliverables_dir} ({exc})") from exc

    for fn in entries:
        full = os.path.join(deliverables_dir, fn)
        if not os.path.isfile(full):
            continue
        lower = fn.lower()
        if lower.endswith(".txt") and lower != "readme.txt":
            try:
                with open(full, "r", encoding="utf-8", errors="replace") as f:
                    payload["texts"].append({"filename": fn, "content": f.read()})
            except Exception as exc:
                payload["texts"].append({"filename": fn, "content": f"(Failed to read file: {exc})"})
        elif lower.endswith(".png"):
            try:
                with open(full, "rb") as f:
                    b64 = base64.b64encode(f.read()).decode("ascii")
                payload["images"].append({"filename": fn, "data_uri": f"data:image/png;base64,{b64}"})
            except Exception as exc:
                # Skip unreadable images, but keep an error placeholder via an empty list entry? We'll skip.
                pass

    payload["texts"].sort(key=lambda d: d.get("filename", ""))
    payload["images"].sort(key=lambda d: d.get("filename", ""))
    return deliverables_dir, payload


def load_and_validate_projects(projecten_dir: str) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    """
    Return signature:
      projects_df
      time_entries_df
      project_info_map
      deliverables_map: project_id -> {"texts":[...], "images":[...]}
    """
    project_rows: List[Dict[str, Any]] = []
    all_time_entries: List[pd.DataFrame] = []
    project_info_map: Dict[str, Dict[str, Any]] = {}
    deliverables_map: Dict[str, Dict[str, Any]] = {}

    missing_deliverables: List[Tuple[str, str]] = []

    for folder_path in discover_project_folders(projecten_dir):
        folder_name = os.path.basename(folder_path)
        derived_project_id = derive_project_id_from_folder(folder_name)

        project_info_path = os.path.join(folder_path, "project_info.xlsx")
        time_log_path = os.path.join(folder_path, "time_log.xlsx")

        if not os.path.exists(project_info_path):
            raise FileNotFoundError(f"Missing project_info.xlsx in project folder '{folder_name}'")

        info = read_project_info_kv_from_xlsx(project_info_path)
        project_info_map[derived_project_id] = dict(info)

        info_project_id = str(info.get("project_id", "")).strip()
        if not info_project_id:
            raise ValueError(f"'project_id' missing or empty in project_info.xlsx for '{folder_name}'")
        if info_project_id != derived_project_id:
            raise ValueError(
                f"Project ID mismatch in folder '{folder_name}'. "
                f"Derived from folder: '{derived_project_id}', "
                f"but project_info.xlsx contains: '{info_project_id}'."
            )

        status = str(info.get("status", "")).strip()
        actual_end_date = parse_date(info.get("actual_end_date"))
        if status == "Closed" and actual_end_date is None:
            raise ValueError(
                f"Project '{folder_name}' is status=Closed but actual_end_date is missing in project_info.xlsx."
            )

        # Deliverables verification + collection
        try:
            _, deliverables_payload = _verify_and_collect_deliverables(folder_path, folder_name)
            deliverables_map[derived_project_id] = deliverables_payload
        except FileNotFoundError:
            missing_deliverables.append((folder_name, os.path.join(folder_path, "Deliverables")))
            deliverables_map[derived_project_id] = {"texts": [], "images": []}

        project_row = dict(info)
        project_row["project_id"] = derived_project_id
        project_row["__folder_name"] = folder_name
        project_row["__project_folder"] = os.path.relpath(folder_path, SCRIPT_DIR)
        project_row["__project_info_file"] = os.path.relpath(project_info_path, SCRIPT_DIR)
        project_row["__reporting_role"] = resolve_project_reporting_role(info, folder_name)

        project_row["start_date"] = parse_date(info.get("start_date"))
        project_row["target_end_date"] = parse_date(info.get("target_end_date"))
        project_row["actual_end_date"] = actual_end_date

        project_row['programma(s)'] = project_row.get("programma (if multiple, separate by |)") or project_row.get("programma")
        programma_values = _split_pipe_values(project_row.get("programma (if multiple, separate by |)") or project_row.get("programma"))
        if programma_values:
            project_row["programma"] = programma_values[0]
            for idx, extra in enumerate(programma_values[1:], start=2):
                project_row[f"programma{idx:02d}"] = extra

        project_row['theme(s)'] = project_row.get("theme (if multiple, separate by |)") or project_row.get("theme")
        theme_values = _split_pipe_values(project_row.get("theme (if multiple, separate by |)") or project_row.get("theme"))
        if theme_values:
            project_row["theme"] = theme_values[0]
            for idx, extra in enumerate(theme_values[1:], start=2):
                project_row[f"theme{idx:02d}"] = extra

        project_row['requester(s)'] = project_row.get("requester(s) (if multiple, separate by |)") or project_row.get("requester")
        requester_values = _split_pipe_values(project_row.get("requester(s) (if multiple, separate by |)") or project_row.get("requester"))
        if requester_values:
            project_row["requester"] = requester_values[0]
            for idx, extra in enumerate(requester_values[1:], start=2):
                project_row[f"requester{idx:02d}"] = extra

        if os.path.exists(time_log_path):
            meta = read_time_log_project_metadata(time_log_path)
            if meta.get("project_id") and meta["project_id"] != derived_project_id:
                raise ValueError(
                    f"time_log.xlsx metadata project_id mismatch in '{folder_name}'. "
                    f"Derived folder id: '{derived_project_id}', metadata says: '{meta['project_id']}'."
                )

            time_df = read_time_log_entries(time_log_path)
            if not time_df.empty:
                time_df = time_df.copy()
                time_df["project_id"] = derived_project_id
                time_df["programma"] = str(project_row.get("programma", "Unknown") or "Unknown")
                time_df["project_name"] = str(project_row.get("project_name", derived_project_id) or derived_project_id)
                time_df["__project_folder"] = project_row["__project_folder"]
                time_df["duration_hours"] = pd.to_numeric(time_df["duration_minutes"], errors="coerce") / 60.0
                all_time_entries.append(time_df)

        project_rows.append(project_row)

    if missing_deliverables:
        lines = ["ERROR: Missing Deliverables folder for the following project(s):"]
        for folder_name, expected_path in missing_deliverables:
            lines.append(f"- {folder_name}  (expected: {expected_path})")
        raise ValueError("\n".join(lines))

    projects_df = pd.DataFrame(project_rows)
    project_names = (
        projects_df["project_name"].fillna("").astype(str).str.strip()
        if "project_name" in projects_df.columns
        else pd.Series([""] * len(projects_df), index=projects_df.index, dtype="object")
    )
    project_names_norm = project_names.str.casefold()
    duplicate_name_mask = (project_names_norm != "") & project_names_norm.duplicated(keep=False)
    if duplicate_name_mask.any():
        duplicate_rows = projects_df.loc[
            duplicate_name_mask, ["project_id", "__folder_name", "__project_info_file"]
        ].copy()
        duplicate_rows["project_name"] = project_names.loc[duplicate_name_mask]
        duplicate_rows = duplicate_rows.sort_values(["project_name", "project_id"], ascending=[True, True])

        error_lines = [
            "ERROR: Duplicate project_name values found in project_info.xlsx.",
            "Each project must use a unique 'project_name' to avoid duplicate or mixed reporting output.",
            "Update the sheet 'ProjectInfo' key 'project_name' in these files:",
        ]

        for _, duplicate_row in duplicate_rows.iterrows():
            error_lines.append(
                (
                    f"- project_name='{duplicate_row.get('project_name', '')}' | "
                    f"project_id='{duplicate_row.get('project_id', '')}' | "
                    f"folder='{duplicate_row.get('__folder_name', '')}' | "
                    f"file='{duplicate_row.get('__project_info_file', '')}'"
                )
            )

        raise ValueError("\n".join(error_lines))
    for col in ["programma", "requester", "status", "project_name", "theme"]:
        if col not in projects_df.columns:
            projects_df[col] = "Unknown"
        projects_df[col] = projects_df[col].fillna("Unknown").replace("", "Unknown")

    time_entries_df = pd.concat(all_time_entries, ignore_index=True) if all_time_entries else pd.DataFrame()
    projects_df = projects_df.sort_values(["created_at"]).reset_index(drop=True)
    if not time_entries_df.empty:
        if "date" in time_entries_df.columns:
            time_entries_df["date"] = pd.to_datetime(time_entries_df["date"], errors="coerce")
            time_entries_df = (
                time_entries_df
                .sort_values(["date", "project_id"], ascending=[True, True], na_position="last")
                .reset_index(drop=True)
            )
        elif "Date*" in time_entries_df.columns:
            time_entries_df = (
                time_entries_df
                .sort_values(["Date*", "project_id"], ascending=[True, True], na_position="last")
                .reset_index(drop=True)
            )

    return projects_df, time_entries_df, project_info_map, deliverables_map


def _projects_filter_controls_html(projects_df: pd.DataFrame) -> str:
    # Build dropdown options (unique, non-empty) for priority, estimated_magnitude, programma tokens
    def uniq_sorted(col: str) -> List[str]:
        if col not in projects_df.columns:
            return []
        vals = (
            projects_df[col]
            .fillna("")
            .astype(str)
            .map(lambda s: s.strip())
            .loc[lambda s: s != ""]
            .unique()
            .tolist()
        )
        return sorted(vals, key=lambda s: s.lower())

    priorities = uniq_sorted("priority")
    magnitudes = uniq_sorted("estimated_magnitude")

    # Programma tokens (supports "programma", "programma02", etc, and pipe-separated)
    programmas: set[str] = set()
    for _, row in projects_df.iterrows():
        vals = extract_group_values(row, "programma")
        for v in vals:
            for tok in _split_pipe_values(v):
                programmas.add(tok)
            if v and v.strip():
                programmas.add(v.strip())
    programma_list = sorted(programmas, key=lambda s: s.lower())

    def opts(options: List[str]) -> str:
        out = ["<option value=''>All</option>"]
        for v in options:
            out.append(f"<option value='{html.escape(v)}'>{html.escape(v)}</option>")
        return "".join(out)

    status_buttons = [
        ("", "All"),
        ("Proposed", "Proposed"),
        ("Active", "Active"),
        ("On-hold", "On-hold"),
        ("Closed", "Closed"),
        ("Cancelled", "Cancelled"),
    ]
    btns = []
    for val, label in status_buttons:
        css = "tab-btn status-btn" + (" active" if val == "" else "")
        btns.append(
            f"<button class='{css}' data-status='{html.escape(val)}' onclick=\"setStatusFilter('{html.escape(val)}')\">{html.escape(label)}</button>"
        )

    sort_options = [
        ("last_updated", "Last log date"),
        ("hours_reported", "Hours reported"),
        ("hours_cap", "Hours cap"),
        ("pct_completed", "% completed"),
        ("estimated_magnitude", "Estimated magnitude"),
        ("start_date", "Start date"),
        ("priority", "Priority"),
    ]
    sort_btns: List[str] = []
    for val, label in sort_options:
        css = "tab-btn sort-btn" + (" active" if val == "last_updated" else "")
        shown = f"{label} ▼" if val == "last_updated" else label
        sort_btns.append(
            (
                f"<button class='{css}' data-sort='{html.escape(val)}' data-label='{html.escape(label)}' "
                f"onclick=\"setSort('{html.escape(val)}')\">{html.escape(shown)}</button>"
            )
        )

    return (
        "<div class='projects-controls'>"
        "<div class='projects-filters'>"
        "<div class='projects-filter-block'>"
        "<div class='projects-filter-label'>Status</div>"
        "<div class='projects-status-buttons'>" + "".join(btns) + "</div>"
        "</div>"
        "<div class='projects-filter-block'>"
        "<div class='projects-filter-label'>Sort by</div>"
        "<div class='projects-sort-buttons'>" + "".join(sort_btns) + "</div>"
        "</div>"
        "<div class='projects-filter-block'>"
        "<div class='projects-filter-label'>Priority</div>"
        f"<select id='filter-priority' onchange='applyProjectsFilters()'>{opts(priorities)}</select>"
        "</div>"
        "<div class='projects-filter-block'>"
        "<div class='projects-filter-label'>Magnitude</div>"
        f"<select id='filter-magnitude' onchange='applyProjectsFilters()'>{opts(magnitudes)}</select>"
        "</div>"
        "<div class='projects-filter-block'>"
        "<div class='projects-filter-label'>Programma</div>"
        f"<select id='filter-programma' onchange='applyProjectsFilters()'>{opts(programma_list)}</select>"
        "</div>"
        "<div class='projects-filter-block'>"
        "<div class='projects-filter-label'>Deliverables</div>"
        "<label class='projects-checkbox'>"
        "<input type='checkbox' id='filter-has-deliverables' onchange='applyProjectsFilters()'/> "
        "Has deliverables"
        "</label>"
        "</div>"
        "</div>"
        "</div>"
    )


def build_projects_page_html(
    projects_df: pd.DataFrame,
    time_entries_df: pd.DataFrame,
    project_info_map: Dict[str, Dict[str, Any]],
    deliverables_map: Dict[str, Dict[str, Any]],
) -> Tuple[str, str]:
    """Returns (filters_html, projects_list_html)."""
    filters_html = _projects_filter_controls_html(projects_df)

    # Last log + hours reported by project
    last_logged_by_project: Dict[str, Optional[pd.Timestamp]] = {}
    hours_reported_by_project: Dict[str, float] = {}
    if (
        time_entries_df is not None
        and not time_entries_df.empty
        and "project_id" in time_entries_df.columns
    ):
        cols = ["project_id"]
        if "date" in time_entries_df.columns:
            cols.append("date")
        if "duration_hours" in time_entries_df.columns:
            cols.append("duration_hours")
        if "duration_minutes" in time_entries_df.columns:
            cols.append("duration_minutes")
        tmp = time_entries_df[cols].copy()
        tmp["project_id"] = tmp["project_id"].astype(str).str.strip()
        if "date" in tmp.columns:
            tmp["date"] = pd.to_datetime(tmp["date"], errors="coerce")
            tmp_sorted = tmp.dropna(subset=["date"]).sort_values(
                ["project_id", "date"], ascending=[True, True], na_position="last"
            )
            latest_rows = tmp_sorted.groupby("project_id", as_index=False).tail(1)
            last_logged_by_project = {
                str(pid).strip(): ts
                for pid, ts in zip(latest_rows["project_id"], latest_rows["date"])
                if str(pid).strip()
            }
        if "duration_hours" in tmp.columns:
            tmp["__hours"] = pd.to_numeric(tmp["duration_hours"], errors="coerce")
        elif "duration_minutes" in tmp.columns:
            tmp["__hours"] = pd.to_numeric(tmp["duration_minutes"], errors="coerce") / 60.0
        else:
            tmp["__hours"] = 0.0
        hgroup = tmp.groupby("project_id")["__hours"].sum(min_count=1).fillna(0.0)
        hours_reported_by_project = {
            str(pid).strip(): float(val) for pid, val in hgroup.items() if str(pid).strip()
        }

    projects_view = projects_df.copy()
    projects_view["project_id"] = projects_view.get("project_id", pd.Series(dtype="object")).astype(str).str.strip()
    projects_view["__last_log_date"] = projects_view["project_id"].map(last_logged_by_project)
    projects_view["__last_log_date"] = pd.to_datetime(projects_view["__last_log_date"], errors="coerce")
    projects_view["__last_updated"] = projects_view["__last_log_date"]
    sort_cols = ["__last_updated"]
    ascending = [False]
    if "created_at" in projects_view.columns:
        sort_cols.append("created_at")
        ascending.append(False)
    sort_cols.append("project_id")
    ascending.append(True)
    projects_view = projects_view.sort_values(sort_cols, ascending=ascending, na_position="last").reset_index(drop=True)

    what_i_did_col: Optional[str] = None
    for candidate in ("WhatIDid*", "WhatIDid"):
        if time_entries_df is not None and not time_entries_df.empty and candidate in time_entries_df.columns:
            what_i_did_col = candidate
            break

    items: List[str] = []
    for _, row in projects_view.iterrows():
        project_id = str(row.get("project_id", "")).strip()
        project_name = str(row.get("project_name", project_id)).strip() or project_id

        status = str(row.get("status", "")).strip()
        priority = str(row.get("priority", "")).strip()
        magnitude = str(row.get("estimated_magnitude", "")).strip()
        priority_rank = _priority_rank(priority)
        magnitude_hours = float(estimate_magnitude_weight(magnitude))
        status_l = status.strip().lower()
        priority_l = priority.strip().lower()

        programmas = extract_group_values(row, "programma") or []
        programma_attr = ",".join([p.strip() for p in programmas if p and p.strip()])

        info = project_info_map.get(project_id, {})
        hours_reported = float(hours_reported_by_project.get(project_id, 0.0) or 0.0)
        hours_cap = _to_float_relaxed(row.get("hours_cap"))
        if hours_cap is None:
            cap_key = _find_col(list(info.keys()), ["hours", "cap"]) if info else None
            if cap_key:
                hours_cap = _to_float_relaxed(info.get(cap_key))
        hours_cap_attr = float(hours_cap) if hours_cap is not None else -1.0
        pct_completed = (hours_reported / hours_cap * 100.0) if hours_cap is not None and hours_cap > 0 else None
        pct_completed_val = float(pct_completed) if pct_completed is not None else -1.0
        progress_pct = max(0.0, min(100.0, pct_completed if pct_completed is not None else 0.0))
        progress_state = " over" if pct_completed is not None and pct_completed > 100 else ""

        start_date_val = parse_date(row.get("start_date")) or parse_date(info.get("start_date"))
        start_date_epoch = int(datetime.combine(start_date_val, datetime.min.time()).timestamp()) if start_date_val else -1

        last_log_ts = last_logged_by_project.get(project_id)
        last_updated_val = pd.to_datetime(last_log_ts, errors="coerce")
        last_updated_epoch = int(last_updated_val.timestamp()) if not pd.isna(last_updated_val) else -1

        if status_l == "closed":
            summary_color = "#B8B8B8"
        elif status_l == "active":
            summary_color = {
                "low": BASE_BLACK,
                "medium": BASE_BLUE,
                "high": BASE_RED,
                "critical": BASE_ORANGE,
            }.get(priority_l, BASE_BLACK)
        else:
            summary_color = BASE_BLACK

        progress_text_left = f"Recorded hours: {hours_reported:.1f} h"
        if hours_cap is not None and hours_cap > 0:
            progress_text_right = f"{hours_cap:.1f} h cap • {pct_completed_val:.0f}%"
        else:
            progress_text_right = "No hours cap"
        progress_bar_html = (
            "<div class='project-progress'>"
            "<div class='project-progress-header'>"
            f"<div class='project-progress-left'>{html.escape(progress_text_left)}</div>"
            f"<div class='project-progress-right'>{html.escape(progress_text_right)}</div>"
            "</div>"
            "<div class='project-progress-bar'>"
            f"<div class='project-progress-fill{progress_state}' style='width:{progress_pct:.1f}%'></div>"
            "</div>"
            "</div>"
        )

        info_rows: List[str] = []
        for key in sorted(info.keys()):
            val = info.get(key)
            if val is None or (isinstance(val, float) and pd.isna(val)) or str(val).strip() == "":
                continue
            info_rows.append(
                "<tr><td>" + html.escape(str(key)) + "</td><td>" + html.escape(str(val)) + "</td></tr>"
            )
        info_table = (
            "<table class='project-info-table'>"
            + ("".join(info_rows) if info_rows else "<tr><td colspan='2'>No data</td></tr>")
            + "</table>"
        )

        dels = deliverables_map.get(project_id, {"texts": [], "images": []})
        has_deliverables = bool((dels.get("texts") or []) or (dels.get("images") or []))
        txt_blocks: List[str] = []
        for t in dels.get("texts", []) or []:
            fn = html.escape(str(t.get("filename", "")))
            content = html.escape(str(t.get("content", "") or ""))
            txt_blocks.append(
                "<div class='deliverable-text'>"
                f"<div class='deliverable-fn'>{fn}</div>"
                f"<pre>{content}</pre>"
                "</div>"
            )
        img_blocks: List[str] = []
        for im in dels.get("images", []) or []:
            fn = html.escape(str(im.get("filename", "")))
            uri = str(im.get("data_uri", "") or "")
            if not uri:
                continue
            img_blocks.append(
                "<div class='deliverable-image'>"
                f"<div class='deliverable-fn'>{fn}</div>"
                f"<img src='{html.escape(uri)}' alt='{fn}'/>"
                "</div>"
            )

        text_col = "<div class='deliverables-text'>" + "".join(txt_blocks) + "</div>" if txt_blocks else ""
        img_col = "<div class='deliverables-images'>" + "".join(img_blocks) + "</div>" if img_blocks else ""

        # Recent time logs table (most recent entries)
        timelog_html = ""
        activitytype_html = ""
        if (
            time_entries_df is not None
            and not time_entries_df.empty
            and "project_id" in time_entries_df.columns
            and "duration_minutes" in time_entries_df.columns
        ):
            proj_entries_all = time_entries_df.loc[
                time_entries_df["project_id"].astype(str).str.strip() == project_id
            ].copy()
            if not proj_entries_all.empty:
                activitytype_html = build_project_activitytype_pie_html(proj_entries_all)
                proj_entries = proj_entries_all.copy()
                if "date" in proj_entries.columns:
                    proj_entries["date"] = pd.to_datetime(proj_entries["date"], errors="coerce")
                    proj_entries = proj_entries.sort_values(["date"], ascending=False, na_position="last")
                proj_entries = proj_entries.head(8)

                rows: List[str] = []
                for _, entry in proj_entries.iterrows():
                    d = entry.get("date")
                    date_text = ""
                    try:
                        if pd.notna(d):
                            date_text = pd.Timestamp(d).date().isoformat()
                    except Exception:
                        date_text = ""

                    dur_text = _format_minutes_hhmm(entry.get("duration_minutes"))
                    desc_val = entry.get(what_i_did_col) if what_i_did_col else ""
                    desc_html = _escape_html_multiline(desc_val).strip()
                    if not desc_html:
                        desc_html = "<span class='hours-entry-empty'>(no details)</span>"
                    rows.append(
                        "<tr>"
                        f"<td class='timelog-date'>{html.escape(date_text)}</td>"
                        f"<td class='timelog-duration'>{html.escape(dur_text)}</td>"
                        f"<td>{desc_html}</td>"
                        "</tr>"
                    )

                timelog_html = (
                    "<div class='project-timelog'>"
                    "<div class='project-timelog-title'>Recent time logs</div>"
                    "<table class='project-timelog-table'>"
                    "<thead><tr><th>Date</th><th>Duration</th><th>Details</th></tr></thead>"
                    "<tbody>"
                    + "".join(rows)
                    + "</tbody>"
                    "</table>"
                    "</div>"
                )

        expanded = (
            "<div class='project-expanded'>"
            "<div class='project-col project-col-table'>"
            + progress_bar_html
            + info_table
            + activitytype_html
            + timelog_html
            + "</div>"
            + ("<div class='project-col project-col-text'>" + text_col + "</div>" if text_col else "")
            + ("<div class='project-col project-col-images'>" + img_col + "</div>" if img_col else "")
            + "</div>"
        )

        last_log_text = ""
        if last_log_ts is not None and pd.notna(last_log_ts):
            try:
                last_log_text = pd.Timestamp(last_log_ts).date().isoformat()
            except Exception:
                last_log_text = ""
        summary_suffix = f" <span class='project-last-log'>(last log: {html.escape(last_log_text)})</span>" if last_log_text else ""

        items.append(
            "<details class='project-item' "
            f"data-project-id='{html.escape(project_id)}' "
            f"data-status='{html.escape(status)}' "
            f"data-priority='{html.escape(priority)}' "
            f"data-priority-rank='{priority_rank}' "
            f"data-magnitude='{html.escape(magnitude)}' "
            f"data-magnitude-rank='{magnitude_hours:.0f}' "
            f"data-programmas='{html.escape(programma_attr)}' "
            f"data-has-deliverables='{'1' if has_deliverables else '0'}' "
            f"data-last-updated='{last_updated_epoch}' "
            f"data-hours-reported='{hours_reported:.3f}' "
            f"data-hours-cap='{hours_cap_attr:.3f}' "
            f"data-pct-completed='{pct_completed_val:.3f}' "
            f"data-start-date='{start_date_epoch}'>"
            f"<summary><span class='project-summary-title' style='color:{html.escape(summary_color)}'>{html.escape(project_id)} — {html.escape(project_name)}</span>{summary_suffix}</summary>"
            f"{expanded}"
            "</details>"
        )

    list_html = "<div class='projects-list'>" + "".join(items) + "</div>"
    return filters_html, list_html


def write_tabbed_html(
    counts_fig: go.Figure,
    hours_fig: go.Figure,
    percentage_fig: go.Figure,
    out_html_path: str,
    header_context: Dict[str, Any],
    tables_html: str,
    hours_metrics_html: str,
    percentage_metrics_html: str,
    percentage_explanation_html: str,
    percentage_section_payloads: Optional[List[Dict[str, Any]]],
    sideways_bar_chart_html: str,
    nn_note: Optional[str],
    nn_summary: Optional[Dict[str, Any]],
    enabled_tabs: Tuple[str, ...] = ("hours", "percentage", "projects"),
) -> None:
    """Write a single-period HTML report with configurable tabs."""
    enabled_tabs_norm: List[str] = []
    for t in enabled_tabs:
        if t in ("counts", "hours", "percentage", "projects") and t not in enabled_tabs_norm:
            enabled_tabs_norm.append(t)
    if not enabled_tabs_norm:
        enabled_tabs_norm = ["percentage", "projects"]
    default_tab = enabled_tabs_norm[0]

    counts_html = (
        pio.to_html(counts_fig, include_plotlyjs=False, full_html=False, div_id="counts-fig")
        if "counts" in enabled_tabs_norm
        else ""
    )
    hours_html = (
        pio.to_html(hours_fig, include_plotlyjs=False, full_html=False, div_id="hours-fig")
        if "hours" in enabled_tabs_norm
        else ""
    )
    percentage_html = (
        pio.to_html(percentage_fig, include_plotlyjs=False, full_html=False, div_id="percentage-fig")
        if "percentage" in enabled_tabs_norm
        else ""
    )
    percentage_sections_html = render_plot_sections_html(
        percentage_section_payloads,
        "percentage-section",
    )
    plotly_cdn = _plotly_cdn_src()

    title_raw = str(header_context.get("title_text", "Project Portfolio Overview"))
    person_name_raw = str(header_context.get("person_name", "john doe"))
    title_text = html.escape(title_raw)
    person_name_text = html.escape(person_name_raw)
    header_summary_text = html.escape(f"{title_raw} ({person_name_raw})")
    export_date = html.escape(str(header_context.get("export_date", "")))
    period_label = html.escape(str(header_context.get("period_label", "")))
    period_range = html.escape(str(header_context.get("period_range", "")))

    profile_uri = header_context.get("profile_data_uri")
    company_logo_uri = header_context.get("company_logo_data_uri")

    profile_img_html = f"<img class='profile-img' src='{profile_uri}' alt='Profile'/>" if profile_uri else ""
    company_logo_img_html = (
        f"<img class='company-logo-img' src='{company_logo_uri}' alt='{html.escape(_company_label_long())} logo'/>"
        if company_logo_uri
        else ""
    )
    nn_note_html = f"<div class='nn-note'>{html.escape(nn_note)}</div>" if nn_note else ""
    nn_sideways_bar_title_html = build_nn_sideways_bar_title_html(nn_summary)
    sideways_bar_chart_block_html = (
        "<div class='nn-sideways-bar-block'>"
        f"{nn_sideways_bar_title_html}"
        f"{sideways_bar_chart_html}"
        "</div>"
        if sideways_bar_chart_html
        else ""
    )
    header_nn_html = f"<div class='header-nn'>{sideways_bar_chart_block_html}</div>" if sideways_bar_chart_block_html else ""
    header_branding_html = ""
    if company_logo_img_html or profile_img_html:
        header_branding_html = (
            "<div class='header-branding'>"
            f"{company_logo_img_html}"
            f"{profile_img_html}"
            "</div>"
        )

    projects_html = tables_html or ""

    tab_labels = {"counts": "Counts", "hours": "Hours", "percentage": "Percentage", "projects": "Projects"}
    tab_buttons_html = "".join(
        [
            (
                f"<button class=\"tab-btn{' active' if t == default_tab else ''}\" "
                f"id=\"btn-{t}\" onclick=\"showTab('{t}')\">{tab_labels[t]}</button>"
            )
            for t in enabled_tabs_norm
        ]
    )

    tab_panels: List[str] = []
    for t in enabled_tabs_norm:
        panel_css = "tab-panel active" if t == default_tab else "tab-panel"
        if t == "counts":
            tab_panels.append(f"<div class=\"{panel_css}\" id=\"tab-counts\">{counts_html}</div>")
        elif t == "hours":
            tab_panels.append(
                f"<div class=\"{panel_css}\" id=\"tab-hours\">"
                f"<div class='hours-metrics'>{hours_metrics_html}</div>"
                f"{hours_html}"
                f"</div>"
            )
        elif t == "percentage":
            percentage_panel_html = (
                percentage_sections_html
                if percentage_sections_html
                else f"{percentage_html}{percentage_explanation_html}"
            )
            tab_panels.append(
                f"<div class=\"{panel_css}\" id=\"tab-percentage\">"
                f"<div class='hours-metrics'>{percentage_metrics_html}</div>"
                f"{percentage_panel_html}"
                f"</div>"
            )
        elif t == "projects":
            tab_panels.append(f"<div class=\"{panel_css}\" id=\"tab-projects\">{projects_html}</div>")

    tab_panels_html = "".join(tab_panels)

    html_content = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>{title_text}</title>
  <script src="{plotly_cdn}"></script>
  <style>
    body {{
      font-family: "Segoe UI", Tahoma, sans-serif;
      margin: 0;
      background: #FAFAFA;
      color: #111;
    }}
    .page {{ padding: 24px 28px 40px; }}
    .sticky-header {{
      position: sticky;
      top: 0;
      z-index: 50;
      background: #FAFAFA;
      padding-top: 8px;
      box-shadow: 0 2px 6px rgba(0,0,0,0.08);
    }}
    .header-collapsible {{
      border: 1px solid #DDD;
      border-radius: 10px;
      background: #FFF;
      box-shadow: 0 1px 2px rgba(0,0,0,0.05);
      margin-bottom: 8px;
    }}
    .header-collapsible summary {{
      cursor: pointer;
      list-style: none;
      outline: none;
      padding: 8px 10px;
      font-size: 11px;
      font-weight: 600;
      color: #333;
      display: flex;
      align-items: center;
      justify-content: center;
      gap: 6px;
      background: #F5F7FA;
      border-bottom: 1px solid #E5E8ED;
      user-select: none;
    }}
    .header-collapsible summary::-webkit-details-marker {{ display: none; }}
    .header-collapsible summary::before {{
      content: "▾";
      color: #555;
      font-weight: 700;
    }}
    .header-collapsible:not([open]) summary::before {{ content: "▸"; }}
    .header-collapsible:not([open]) summary {{ border-bottom: 0; }}
    .header-collapsible-body {{ padding: 0 10px 8px; }}
    .report-header {{
      display: flex;
      gap: 6px;
      align-items: center;
      justify-content: center;
      flex-wrap: nowrap;
      padding: 10px 0 2px;
      text-align: center;
    }}
    .header-left {{
      flex: 0 1 340px;
      min-width: 260px;
      display: flex;
      flex-direction: column;
      align-items: flex-start;
      justify-content: center;
      text-align: left;
    }}
    .header-left h1 {{ margin: 0 0 6px 0; font-size: 26px; }}
    .header-left .subtitle {{ margin: 0 0 6px 0; font-size: 14px; color: #555; font-weight: 600; }}
    .header-left .meta {{ font-size: 14px; color: #444; text-align: left; }}
    .header-left .nn-note {{ text-align: left; }}
    .header-nn {{
      display: flex;
      flex: 0 1 520px;
      min-width: 380px;
      justify-content: flex-start;
      align-items: center;
    }}
    .header-branding {{
      display: flex;
      flex: 0 0 auto;
      min-width: 220px;
      justify-content: center;
      align-items: center;
      gap: 8px;
      flex-wrap: nowrap;
    }}
    .nn-sideways-bar-block {{
      display: flex;
      flex-direction: column;
      align-items: center;
      gap: 2px;
      min-width: 420px;
      max-width: 540px;
      width: min(540px, 52vw);
    }}
    .nn-sideways-bar-title-row {{
      display: flex;
      align-items: center;
      justify-content: center;
      gap: 4px;
    }}
    .nn-sideways-bar-title {{
      writing-mode: horizontal-tb;
      transform: none;
      font-size: 11px;
      color: #111;
      line-height: 1.15;
    }}
    .nn-sideways-bar-subtitle {{
      color: #444;
      font-weight: 500;
    }}
    .nn-help-icon {{
      position: relative;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      width: 16px;
      height: 16px;
      border: 1px solid {BASE_BLUE};
      border-radius: 999px;
      color: {BASE_BLUE};
      font-size: 11px;
      font-weight: 700;
      cursor: help;
      user-select: none;
    }}
    .nn-help-icon:focus {{
      outline: 2px solid rgba(1,55,138,0.35);
      outline-offset: 2px;
    }}
    .nn-help-tooltip {{
      position: absolute;
      top: calc(100% + 8px);
      right: 0;
      width: min(430px, 80vw);
      background: #0F2F66;
      color: #FFF;
      border-radius: 8px;
      padding: 10px 12px;
      font-size: 12px;
      font-weight: 500;
      line-height: 1.4;
      text-align: left;
      box-shadow: 0 10px 24px rgba(0,0,0,0.25);
      opacity: 0;
      transform: translateY(-4px);
      transition: opacity 120ms ease, transform 120ms ease;
      pointer-events: none;
      z-index: 200;
    }}
    .nn-help-icon:hover .nn-help-tooltip,
    .nn-help-icon:focus .nn-help-tooltip,
    .nn-help-icon:focus-visible .nn-help-tooltip {{
      opacity: 1;
      transform: translateY(0);
    }}
    .profile-img {{
      width: 96px;
      height: 96px;
      object-fit: cover;
      border-radius: 10px;
      border: 2px solid #EEE;
      background: #FFF;
    }}
    .company-logo-img {{ height: 52px; max-width: 180px; object-fit: contain; }}
    @media (max-width: 1200px) {{
      .report-header {{
        flex-wrap: wrap;
        gap: 12px;
      }}
      .header-left {{
        flex: 1 1 100%;
      }}
      .header-nn {{
        flex: 1 1 100%;
        min-width: 300px;
      }}
      .nn-sideways-bar-block {{
        min-width: 300px;
        width: min(540px, 94vw);
      }}
      .header-branding {{
        min-width: 220px;
        justify-content: flex-start;
      }}
    }}
    .tabs {{
      display: flex;
      gap: 8px;
      margin: 4px 0 12px;
      padding-bottom: 12px;
      flex-wrap: wrap;
      justify-content: center;
      align-items: center;
    }}
    .tab-btn {{
      padding: 8px 16px;
      border: 1px solid #CCC;
      border-radius: 6px;
      background: #FFF;
      cursor: pointer;
      font-weight: 600;
    }}
    .tab-btn.active {{
      background: dodgerblue;
      border-color: dodgerblue;
      color: #FFF;
    }}
    .tabs-row1 .tab-btn.active {{ background: royalblue; border-color: royalblue; }}
    .tabs-row2 .tab-btn.active {{ background: seagreen; border-color: seagreen; }}
    .tabs-row3 .tab-btn.active {{ background: darkorange; border-color: darkorange; }}
    .tab-panel {{ display: none; }}
    .tab-panel.active {{ display: block; }}
    .hours-metrics {{ margin: 6px 0 16px; }}
	    .hours-breakdown {{
	      background: #FFF;
	      border: 1px solid #DDD;
	      border-radius: 10px;
	      padding: 12px;
	      box-shadow: 0 1px 2px rgba(0,0,0,0.05);
	    }}
	    details.hours-breakdown summary {{
	      cursor: pointer;
	      font-weight: 700;
	      list-style: none;
	      outline: none;
	    }}
	    details.hours-breakdown summary::-webkit-details-marker {{
	      display: none;
	    }}
	    details.hours-breakdown summary::before {{
	      content: "▸";
	      display: inline-block;
	      width: 1em;
	      color: {BASE_BLUE};
	    }}
	    details.hours-breakdown[open] summary::before {{
	      content: "▾";
	    }}
	    .hours-breakdown-header {{
	      display: flex;
	      align-items: baseline;
	      justify-content: space-between;
	      gap: 10px;
	      flex-wrap: wrap;
	      margin-bottom: 8px;
	    }}
	    .hours-breakdown h3 {{ margin: 0; font-size: 16px; }}
	    .hours-breakdown-note {{ margin-top: 8px; font-size: 12px; color: #444; }}
    .hours-breakdown-list {{
      display: flex;
      flex-direction: column;
      gap: 6px;
    }}
    .hours-project {{
      background: #FFF;
      border: 1px solid #EEE;
      border-radius: 8px;
      padding: 6px 10px;
    }}
    .hours-project[open] {{
      border-color: #CCC;
      box-shadow: 0 1px 2px rgba(0,0,0,0.05);
    }}
    .hours-project summary {{
      cursor: pointer;
      font-weight: 600;
      list-style: none;
      outline: none;
    }}
    .hours-project summary::-webkit-details-marker {{
      display: none;
    }}
    .hours-project summary::before {{
      content: "▸";
      display: inline-block;
      width: 1em;
      color: {BASE_BLUE};
    }}
    .hours-project[open] summary::before {{
      content: "▾";
    }}
    .hours-project-total {{
      font-variant-numeric: tabular-nums;
    }}
    .hours-project-percent {{
      color: #444;
      font-weight: 600;
    }}
    .hours-project-entries {{
      margin-top: 8px;
      padding-left: 1.2em;
    }}
    .hours-entry-table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 12px;
    }}
    .hours-entry-table th {{
      text-align: left;
      padding: 6px 6px;
      color: #555;
      border-bottom: 1px solid #EEE;
    }}
    .hours-entry-table td {{
      padding: 6px 6px;
      border-bottom: 1px solid #F3F3F3;
      vertical-align: top;
      word-break: break-word;
    }}
    .hours-entry-duration {{
      width: 110px;
      text-align: right;
      font-variant-numeric: tabular-nums;
      white-space: nowrap;
    }}
    .hours-entry-percent {{
      width: 90px;
      text-align: right;
      font-variant-numeric: tabular-nums;
      white-space: nowrap;
    }}
    .hours-entry-empty {{
      color: #777;
      font-style: italic;
    }}
    .nn-metrics {{
      display: flex;
      flex-wrap: wrap;
      gap: 12px 18px;
      font-size: 14px;
      background: #FFF;
      border: 1px solid #DDD;
      border-radius: 8px;
      padding: 10px 12px;
      box-shadow: 0 1px 2px rgba(0,0,0,0.05);
    }}

    /* Projects page styles */
    .projects-controls {{
      background: #FFF;
      border: 1px solid #DDD;
      border-radius: 10px;
      padding: 10px 12px;
      box-shadow: 0 1px 2px rgba(0,0,0,0.05);
      margin: 8px 0 12px;
    }}
    .projects-filters {{
      display: flex;
      gap: 12px 16px;
      flex-wrap: wrap;
      align-items: flex-end;
    }}
    .projects-filter-block {{ display: flex; flex-direction: column; gap: 4px; }}
    .projects-filter-label {{ font-size: 12px; color: #444; font-weight: 600; }}
    .projects-status-buttons {{ display: flex; gap: 6px; flex-wrap: wrap; }}
    .projects-sort-buttons {{ display: flex; gap: 6px; flex-wrap: wrap; }}
    .tab-btn.status-btn {{ padding: 6px 10px; font-size: 13px; }}
    .tab-btn.sort-btn {{ padding: 6px 10px; font-size: 13px; }}
    .projects-checkbox {{ display: flex; gap: 6px; align-items: center; font-weight: 600; }}
    .project-last-log {{ color: #444; font-weight: 600; font-size: 12px; }}
    .project-summary-title {{ font-weight: 700; }}
    .project-progress {{
      background: #F7F9FC;
      border: 1px solid #E3E7EE;
      border-radius: 8px;
      padding: 8px 10px;
      margin-bottom: 10px;
    }}
    .project-progress-header {{
      display: flex;
      justify-content: space-between;
      gap: 10px;
      align-items: baseline;
      flex-wrap: wrap;
      font-size: 12px;
      font-weight: 700;
    }}
    .project-progress-right {{ color: #555; }}
    .project-progress-bar {{
      margin-top: 6px;
      height: 10px;
      background: #E4E8EE;
      border-radius: 999px;
      overflow: hidden;
    }}
    .project-progress-fill {{
      height: 100%;
      width: 0;
      background: {BASE_BLUE};
      border-radius: 999px;
    }}
    .project-progress-fill.over {{ background: {BASE_ORANGE}; }}
    .projects-filters select {{
      padding: 6px 10px;
      border: 1px solid #CCC;
      border-radius: 6px;
      background: #FFF;
      font-weight: 600;
    }}
    .projects-list {{ display: flex; flex-direction: column; gap: 10px; margin-top: 10px; }}
    details.project-item {{
      background: #FFF;
      border: 1px solid #DDD;
      border-radius: 10px;
      padding: 8px 10px;
      box-shadow: 0 1px 2px rgba(0,0,0,0.05);
    }}
    details.project-item summary {{
      cursor: pointer;
      font-weight: 700;
      list-style: none;
      outline: none;
    }}
    details.project-item summary::-webkit-details-marker {{ display: none; }}
    details.project-item summary::before {{
      content: "▸";
      display: inline-block;
      width: 1em;
      color: {BASE_BLUE};
    }}
    details.project-item[open] summary::before {{ content: "▾"; }}
    .project-expanded {{
      margin-top: 10px;
      display: flex;
      gap: 14px;
      align-items: flex-start;
      flex-wrap: wrap;
    }}
    .project-col {{ min-width: 280px; }}
    .project-col-table {{ flex: 1 1 360px; }}
    .project-col-text {{ flex: 1 1 320px; }}
    .project-col-images {{ flex: 1 1 320px; }}
    .project-info-table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 12px;
      background: #FFF;
    }}
    .project-info-table td {{
      padding: 4px 6px;
      border-bottom: 1px solid #EEE;
      vertical-align: top;
      word-break: break-word;
    }}
    .project-info-table td:first-child {{ width: 42%; color: #555; }}
    .project-timelog {{ margin-top: 10px; }}
    .project-timelog-title {{ font-weight: 700; margin: 10px 0 6px; }}
    .project-timelog-table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
    .project-timelog-table th {{ text-align: left; padding: 6px 6px; color: #555; border-bottom: 1px solid #EEE; }}
    .project-timelog-table td {{ padding: 6px 6px; border-bottom: 1px solid #F3F3F3; vertical-align: top; word-break: break-word; }}
    .timelog-date {{ width: 110px; white-space: nowrap; font-variant-numeric: tabular-nums; }}
    .timelog-duration {{ width: 100px; text-align: right; white-space: nowrap; font-variant-numeric: tabular-nums; }}
    .project-activitytype {{ margin-top: 10px; }}
    .project-activitytype-pie {{
      width: 140px;
      height: 140px;
      border-radius: 50%;
      border: 1px solid #DDD;
      margin: 8px 0 10px;
      box-shadow: inset 0 0 0 1px rgba(255,255,255,0.7);
    }}
    .project-activitytype-legend {{ display: flex; flex-direction: column; gap: 4px; font-size: 12px; }}
    .project-activitytype-row {{ display: grid; grid-template-columns: 12px 1fr; gap: 8px; align-items: center; }}
    .project-activitytype-dot {{ width: 10px; height: 10px; border-radius: 999px; display: inline-block; }}
    .project-activitytype-label {{ color: #333; }}
    .deliverable-fn {{ font-weight: 700; margin: 8px 0 6px; }}
    .deliverables-text pre {{
      white-space: pre-wrap;
      word-break: break-word;
      max-height: 420px;
      overflow: auto;
      background: #FAFAFA;
      border: 1px solid #EEE;
      border-radius: 8px;
      padding: 8px;
      margin: 0;
      font-size: 12px;
    }}
    .deliverables-images img {{
      max-width: 420px;
      height: auto;
      border-radius: 8px;
      border: 1px solid #EEE;
      background: #FFF;
    }}
    .nn-note {{ margin-top: 6px; font-size: 13px; color: #8A3B3B; }}
    .plot-sections {{ display: flex; flex-direction: column; gap: 10px; }}
    .plot-section {{
      background: #FFF;
      border: 1px solid #DDD;
      border-radius: 10px;
      box-shadow: 0 1px 2px rgba(0,0,0,0.05);
    }}
    .plot-section-title {{ font-size: 16px; font-weight: 700; }}
    .plot-section-fixed {{ padding: 10px 12px 12px; }}
    .plot-section-fixed .plot-section-title {{ margin-bottom: 8px; }}
    .plot-section-body {{ min-height: 20px; }}
    details.plot-section summary {{
      cursor: pointer;
      list-style: none;
      outline: none;
      padding: 10px 12px;
      display: flex;
      align-items: center;
      gap: 6px;
    }}
    details.plot-section summary::-webkit-details-marker {{ display: none; }}
    details.plot-section summary::before {{
      content: "▸";
      display: inline-block;
      width: 1em;
      color: {BASE_BLUE};
      font-weight: 700;
    }}
    details.plot-section[open] summary::before {{ content: "▾"; }}
    details.plot-section .plot-section-body {{ padding: 0 12px 12px; }}
    .weekly-guidance {{
      margin-top: 10px;
      padding: 10px 12px;
      background: #FFF;
      border: 1px solid #DDD;
      border-radius: 8px;
      font-size: 13px;
      color: #333;
      line-height: 1.45;
      display: flex;
      flex-direction: column;
      gap: 6px;
    }}
    .weekly-guidance-formula {{
      font-family: Consolas, "Courier New", monospace;
      background: #F7F9FC;
      border: 1px solid #E6EBF2;
      border-radius: 6px;
      padding: 6px 8px;
      color: #223;
    }}
  </style>
</head>
<body>
  <div class="page">
    <div class="sticky-header">
      <details class="header-collapsible" id="header-collapsible" open>
        <summary>{header_summary_text}</summary>
        <div class="header-collapsible-body">
          <div class="report-header">
            <div class="header-left">
              <h1>{title_text}</h1>
              <div class="subtitle">{person_name_text}</div>
              <div class="meta"><b>{period_label}</b> — {period_range}</div>
              <div class="meta">Generated: {export_date}</div>
              {nn_note_html}
            </div>
            {header_nn_html}
            {header_branding_html}
          </div>
        </div>
      </details>

      <div class="tabs tabs-row1">
        {tab_buttons_html}
      </div>
    </div>

    {tab_panels_html}
  </div>

  <script>
    var projectsStatusFilter = "";
    var projectsSortKey = "last_updated";
    var projectsSortDesc = true;
    var enabledTabs = {enabled_tabs_norm!r};

    function resizePlotlyIn(rootEl) {{
      if (!rootEl || !window.Plotly) {{
        return;
      }}
      rootEl.querySelectorAll(".plotly-graph-div").forEach(function(plotEl) {{
        Plotly.Plots.resize(plotEl);
      }});
    }}

    function bindPlotSectionToggles() {{
      document.querySelectorAll("details.plot-section").forEach(function(secEl) {{
        secEl.addEventListener("toggle", function() {{
          if (secEl.open) {{
            resizePlotlyIn(secEl);
          }}
        }});
      }});
    }}

    function bindHeaderToggle() {{
      var headerToggle = document.getElementById("header-collapsible");
      if (!headerToggle) {{
        return;
      }}
      headerToggle.addEventListener("toggle", function() {{
        if (!headerToggle.open) {{
          return;
        }}
        resizePlotlyIn(headerToggle);
        var activeTab = document.querySelector(".tab-panel.active");
        if (activeTab) {{
          resizePlotlyIn(activeTab);
        }}
      }});
    }}

    function showTab(name) {{
      enabledTabs.forEach(function(t) {{
        document.getElementById("tab-" + t).classList.toggle("active", t === name);
        document.getElementById("btn-" + t).classList.toggle("active", t === name);
      }});
      resizePlotlyIn(document.getElementById("tab-" + name));
      if (name === "projects") {{
        applyProjectsFilters();
      }}
    }}

    function setStatusFilter(statusVal) {{
      projectsStatusFilter = statusVal || "";
      document.querySelectorAll(".status-btn").forEach(function(btn) {{
        btn.classList.remove("active");
      }});
      document.querySelectorAll(".status-btn").forEach(function(btn) {{
        if ((btn.getAttribute("data-status") || "") === projectsStatusFilter) {{
          btn.classList.add("active");
        }}
      }});
      applyProjectsFilters();
    }}

    function setSort(sortKey) {{
      var key = sortKey || "last_updated";
      if (projectsSortKey === key) {{
        projectsSortDesc = !projectsSortDesc;
      }} else {{
        projectsSortKey = key;
        projectsSortDesc = true;
      }}
      applyProjectsFilters();
    }}

    function refreshSortButtons() {{
      document.querySelectorAll(".sort-btn").forEach(function(btn) {{
        var key = (btn.getAttribute("data-sort") || "");
        var label = (btn.getAttribute("data-label") || btn.textContent || "").replace(/[\\s]*[▲▼]$/, "");
        var active = (key === projectsSortKey);
        btn.classList.toggle("active", active);
        btn.textContent = active ? (label + (projectsSortDesc ? " ▼" : " ▲")) : label;
      }});
    }}

    function applyProjectsFilters() {{
      var prio = document.getElementById("filter-priority");
      var mag = document.getElementById("filter-magnitude");
      var prog = document.getElementById("filter-programma");
      var hasDelivEl = document.getElementById("filter-has-deliverables");
      var prioVal = prio ? (prio.value || "") : "";
      var magVal = mag ? (mag.value || "") : "";
      var progVal = prog ? (prog.value || "") : "";
      var hasDeliv = hasDelivEl ? !!hasDelivEl.checked : false;
      var sortVal = projectsSortKey || "last_updated";
      var sortDesc = (projectsSortDesc !== false);
      var listEl = document.querySelector(".projects-list");
      var items = Array.prototype.slice.call(document.querySelectorAll(".project-item"));

      items.forEach(function(item) {{
        var ok = true;
        var s = (item.getAttribute("data-status") || "");
        var p = (item.getAttribute("data-priority") || "");
        var m = (item.getAttribute("data-magnitude") || "");
        var progs = (item.getAttribute("data-programmas") || "");
        var hd = (item.getAttribute("data-has-deliverables") || "0");

        if (projectsStatusFilter) {{
          ok = ok && (s.toLowerCase() === projectsStatusFilter.toLowerCase());
        }}
        if (prioVal) {{
          ok = ok && (p.toLowerCase() === prioVal.toLowerCase());
        }}
        if (magVal) {{
          ok = ok && (m.toLowerCase() === magVal.toLowerCase());
        }}
        if (progVal) {{
          var tokens = progs.split(",").map(function(x) {{ return x.trim().toLowerCase(); }});
          ok = ok && (tokens.indexOf(progVal.toLowerCase()) >= 0);
        }}
        if (hasDeliv) {{
          ok = ok && (hd === "1");
        }}
        item.style.display = ok ? "" : "none";
      }});

      var cmpNum = function(a, b, attr, desc) {{
        var av = parseFloat(a.getAttribute(attr) || "");
        var bv = parseFloat(b.getAttribute(attr) || "");
        var aMissing = !isFinite(av) || av < 0;
        var bMissing = !isFinite(bv) || bv < 0;
        if (aMissing && bMissing) return 0;
        if (aMissing) return 1;
        if (bMissing) return -1;
        return desc ? (bv - av) : (av - bv);
      }};

      var cmpProjectId = function(a, b) {{
        var aid = (a.getAttribute("data-project-id") || "").toLowerCase();
        var bid = (b.getAttribute("data-project-id") || "").toLowerCase();
        if (aid < bid) return -1;
        if (aid > bid) return 1;
        return 0;
      }};

      items.sort(function(a, b) {{
        var diff = 0;
        if (sortVal === "hours_reported") {{
          diff = cmpNum(a, b, "data-hours-reported", sortDesc);
        }} else if (sortVal === "hours_cap") {{
          diff = cmpNum(a, b, "data-hours-cap", sortDesc);
        }} else if (sortVal === "pct_completed") {{
          diff = cmpNum(a, b, "data-pct-completed", sortDesc);
        }} else if (sortVal === "estimated_magnitude") {{
          diff = cmpNum(a, b, "data-magnitude-rank", sortDesc);
        }} else if (sortVal === "start_date") {{
          diff = cmpNum(a, b, "data-start-date", sortDesc);
        }} else if (sortVal === "priority") {{
          diff = cmpNum(a, b, "data-priority-rank", sortDesc);
        }} else {{
          diff = cmpNum(a, b, "data-last-updated", sortDesc);
        }}
        if (diff !== 0) return diff;
        return cmpProjectId(a, b);
      }});

      if (listEl) {{
        items.forEach(function(item) {{
          listEl.appendChild(item);
        }});
      }}
      refreshSortButtons();
    }}

    window.addEventListener("load", function() {{
      bindPlotSectionToggles();
      bindHeaderToggle();
      var activeTab = document.querySelector(".tab-panel.active");
      if (activeTab) {{
        resizePlotlyIn(activeTab);
      }}
    }});
  </script>
</body>
</html>
"""

    with open(out_html_path, "w", encoding="utf-8") as f:
        f.write(html_content)


def write_multi_period_tabbed_html(
    period_payloads: Dict[str, Dict[str, Any]],
    out_html_path: str,
    header_context: Dict[str, Any],
    tables_html: str,
    projects_filters_html: str = "",
    enabled_tabs: Tuple[str, ...] = ("hours", "percentage", "projects"),
) -> None:
    """Write a combined multi-period HTML report with configurable tabs."""
    plotly_cdn = _plotly_cdn_src()
    title_raw = str(header_context.get("title_text", "Project Portfolio Overview"))
    person_name_raw = str(header_context.get("person_name", "john doe"))
    title_text = html.escape(title_raw)
    person_name_text = html.escape(person_name_raw)
    header_summary_text = html.escape(f"{title_raw} ({person_name_raw})")
    export_date = html.escape(str(header_context.get("export_date", "")))

    profile_uri = header_context.get("profile_data_uri")
    company_logo_uri = header_context.get("company_logo_data_uri")

    profile_img_html = f"<img class='profile-img' src='{profile_uri}' alt='Profile'/>" if profile_uri else ""
    company_logo_img_html = (
        f"<img class='company-logo-img' src='{company_logo_uri}' alt='{html.escape(_company_label_long())} logo'/>"
        if company_logo_uri
        else ""
    )
    header_branding_html = ""
    if company_logo_img_html or profile_img_html:
        header_branding_html = (
            "<div class='header-branding'>"
            f"{company_logo_img_html}"
            f"{profile_img_html}"
            "</div>"
        )

    enabled_tabs_norm: List[str] = []
    for t in enabled_tabs:
        if t in ("counts", "hours", "percentage", "projects") and t not in enabled_tabs_norm:
            enabled_tabs_norm.append(t)
    if not enabled_tabs_norm:
        enabled_tabs_norm = ["percentage", "projects"]
    default_tab = enabled_tabs_norm[0]

    projects_html = tables_html or ""
    projects_filters_html = projects_filters_html or ""

    month_period_ids = sorted(
        [p for p in period_payloads.keys() if str(p).startswith("monthly-")],
        key=lambda p: str(p)[len("monthly-"):],
        reverse=False,
    )
    daily_period_ids = sorted(
        [p for p in period_payloads.keys() if str(p).startswith("daily-")],
        key=lambda p: str(p)[len("daily-"):],
        reverse=True,
    )

    period_groups: List[str] = []
    if daily_period_ids:
        period_groups.append("daily")
    if "weekly" in period_payloads:
        period_groups.append("weekly")
    if "biweekly" in period_payloads:
        period_groups.append("biweekly")
    if month_period_ids:
        period_groups.append("monthly")
    if "yearly" in period_payloads:
        period_groups.append("yearly")
    if not period_groups:
        raise ValueError("No period payloads provided.")

    default_group = "weekly" if "weekly" in period_groups else period_groups[0]
    if default_group == "daily" and not daily_period_ids and "weekly" in period_groups:
        default_group = "weekly"
    default_day_id = daily_period_ids[0] if daily_period_ids else ""
    default_month_id = month_period_ids[-1] if month_period_ids else ""
    if default_group == "daily":
        default_period_id = default_day_id
    elif default_group == "monthly":
        default_period_id = default_month_id
    else:
        default_period_id = default_group

    period_group_buttons_html_parts: List[str] = []
    month_buttons_html_parts: List[str] = []
    period_meta_html_parts: List[str] = []
    period_note_html_parts: List[str] = []
    period_sideways_bar_chart_block_parts: List[str] = []
    period_counts_panels_parts: List[str] = []
    period_hours_panels_parts: List[str] = []
    period_percentage_panels_parts: List[str] = []

    group_labels: Dict[str, str] = {}
    if daily_period_ids:
        group_labels["daily"] = "1-day"
    if "weekly" in period_payloads:
        group_labels["weekly"] = html.escape(str(period_payloads["weekly"].get("label", "1-week")))
    if "biweekly" in period_payloads:
        group_labels["biweekly"] = html.escape(str(period_payloads["biweekly"].get("label", "2-weeks")))
    if month_period_ids:
        group_labels["monthly"] = "Month"
    if "yearly" in period_payloads:
        group_labels["yearly"] = html.escape(str(period_payloads["yearly"].get("label", "Year")))

    for group_key in ("daily", "weekly", "biweekly", "monthly", "yearly"):
        if group_key not in period_groups:
            continue
        label = group_labels.get(group_key, html.escape(group_key))
        is_default = group_key == default_group
        period_group_buttons_html_parts.append(
            (
                f"<button class=\"tab-btn period-btn{' active' if is_default else ''}\" "
                f"id=\"btn-period-{group_key}\" onclick=\"showPeriodGroup('{group_key}')\">{label}</button>"
            )
        )

    day_options_html_parts: List[str] = []
    for day_id in daily_period_ids:
        payload = period_payloads[day_id]
        option_label = html.escape(str(payload.get("label", day_id)))
        selected_attr = " selected" if day_id == default_day_id else ""
        day_options_html_parts.append(
            f"<option value=\"{html.escape(day_id)}\"{selected_attr}>{option_label}</option>"
        )

    for month_id in month_period_ids:
        payload = period_payloads[month_id]
        label = html.escape(str(payload.get("label", month_id)))
        is_default = month_id == default_month_id
        month_buttons_html_parts.append(
            (
                f"<button class=\"tab-btn month-btn{' active' if is_default else ''}\" "
                f"id=\"btn-month-{month_id}\" onclick=\"showMonth('{month_id}')\">{label}</button>"
            )
        )

    period_ids: List[str] = []
    period_ids.extend(daily_period_ids)
    for key in ("weekly", "biweekly"):
        if key in period_payloads:
            period_ids.append(key)
    period_ids.extend(month_period_ids)
    if "yearly" in period_payloads:
        period_ids.append("yearly")

    for period_id in period_ids:
        payload = period_payloads[period_id]
        label = html.escape(str(payload.get("label", period_id)))
        period_range = html.escape(str(payload.get("period_range", "")))
        is_default = period_id == default_period_id

        period_meta_html_parts.append(
            (
                f"<span class=\"period-meta{' active' if is_default else ''}\" "
                f"id=\"meta-{period_id}\"><b>{label}</b> — {period_range}</span>"
            )
        )

        nn_note = payload.get("nn_note") or ""
        period_note_html_parts.append(
            (
                f"<div class=\"nn-note period-note{' active' if is_default else ''}\" "
                f"id=\"nn-note-{period_id}\">{html.escape(str(nn_note))}</div>"
            )
        )

        sideways_bar_chart_html = payload.get("sideways_bar_chart_html") or ""
        if sideways_bar_chart_html:
            nn_sideways_bar_title_html = build_nn_sideways_bar_title_html(payload.get("nn_summary"))
            period_sideways_bar_chart_block_parts.append(
                (
                    f"<div class=\"nn-sideways-bar-block period-nn{' active' if is_default else ''}\" "
                    f"id=\"nn-sideways-bar-block-{period_id}\">"
                    f"{nn_sideways_bar_title_html}"
                    f"{sideways_bar_chart_html}"
                    "</div>"
                )
            )

        show_plots = bool(payload.get("show_plots", True))
        hours_metrics_html = payload.get("hours_metrics_html") or ""
        percentage_metrics_html = payload.get("percentage_metrics_html") or hours_metrics_html
        percentage_explanation_html = payload.get("percentage_explanation_html") or ""
        percentage_section_payloads = payload.get("percentage_section_payloads")
        table_only_html = payload.get("table_only_html") or percentage_metrics_html or hours_metrics_html

        if show_plots:
            counts_fig = payload["counts_fig"]
            hours_fig = payload["hours_fig"]
            percentage_fig = payload["percentage_fig"]
            counts_div_id = f"counts-fig-{period_id}"
            hours_div_id = f"hours-fig-{period_id}"
            percentage_div_id = f"percentage-fig-{period_id}"
            counts_html = pio.to_html(counts_fig, include_plotlyjs=False, full_html=False, div_id=counts_div_id)
            hours_html = pio.to_html(hours_fig, include_plotlyjs=False, full_html=False, div_id=hours_div_id)
            percentage_html = pio.to_html(percentage_fig, include_plotlyjs=False, full_html=False, div_id=percentage_div_id)
            percentage_sections_html = render_plot_sections_html(
                percentage_section_payloads,
                f"percentage-section-{period_id}",
            )
        else:
            counts_html = f"<div class=\"hours-metrics\">{table_only_html}</div>"
            hours_html = ""
            percentage_html = ""
            percentage_sections_html = ""

        period_counts_panels_parts.append(
            (
                f"<div class=\"period-panel{' active' if is_default else ''}\" "
                f"id=\"period-counts-{period_id}\">{counts_html}</div>"
            )
        )
        period_hours_panels_parts.append(
            (
                f"<div class=\"period-panel{' active' if is_default else ''}\" "
                f"id=\"period-hours-{period_id}\">"
                f"<div class=\"hours-metrics\">{hours_metrics_html}</div>"
                f"{hours_html}"
                "</div>"
            )
        )
        period_percentage_panels_parts.append(
            (
                f"<div class=\"period-panel{' active' if is_default else ''}\" "
                f"id=\"period-percentage-{period_id}\">"
                f"<div class=\"hours-metrics\">{percentage_metrics_html}</div>"
                f"{percentage_sections_html if percentage_sections_html else (percentage_html + percentage_explanation_html)}"
                "</div>"
            )
        )

    period_group_buttons_html = "\n".join(period_group_buttons_html_parts)
    day_select_html = "\n".join(day_options_html_parts)
    month_buttons_html = "\n".join(month_buttons_html_parts)
    period_meta_html = "\n".join(period_meta_html_parts)
    period_note_html = "\n".join(period_note_html_parts)
    sideways_bar_chart_blocks_html = "\n".join(period_sideways_bar_chart_block_parts)
    counts_panels_html = "\n".join(period_counts_panels_parts)
    hours_panels_html = "\n".join(period_hours_panels_parts)
    percentage_panels_html = "\n".join(period_percentage_panels_parts)

    tab_labels = {"counts": "Counts", "hours": "Hours", "percentage": "Percentage", "projects": "Projects"}
    tab_buttons_html = "".join(
        [
            (
                f"<button class=\"tab-btn{' active' if t == default_tab else ''}\" "
                f"id=\"btn-{t}\" onclick=\"showTab('{t}')\">{tab_labels[t]}</button>"
            )
            for t in enabled_tabs_norm
        ]
    )

    tab_panels: List[str] = []
    for t in enabled_tabs_norm:
        panel_css = "tab-panel active" if t == default_tab else "tab-panel"
        if t == "counts":
            tab_panels.append(f"<div class=\"{panel_css}\" id=\"tab-counts\">{counts_panels_html}</div>")
        elif t == "hours":
            tab_panels.append(f"<div class=\"{panel_css}\" id=\"tab-hours\">{hours_panels_html}</div>")
        elif t == "percentage":
            tab_panels.append(f"<div class=\"{panel_css}\" id=\"tab-percentage\">{percentage_panels_html}</div>")
        elif t == "projects":
            tab_panels.append(f"<div class=\"{panel_css}\" id=\"tab-projects\">{projects_html}</div>")
    tab_panels_html = "".join(tab_panels)

    html_content = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8"/>
  <meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>{title_text}</title>
  <script src="{plotly_cdn}"></script>
  <style>
    body {{ font-family: "Segoe UI", Tahoma, sans-serif; margin: 0; background: #FAFAFA; color: #111; }}
    .page {{ padding: 24px 28px 40px; }}
    .sticky-header {{ position: sticky; top: 0; z-index: 50; background: #FAFAFA; padding-top: 8px; box-shadow: 0 2px 6px rgba(0,0,0,0.08); }}
    .header-collapsible {{
      border: 1px solid #DDD;
      border-radius: 10px;
      background: #FFF;
      box-shadow: 0 1px 2px rgba(0,0,0,0.05);
      margin-bottom: 8px;
    }}
    .header-collapsible summary {{
      cursor: pointer;
      list-style: none;
      outline: none;
      padding: 8px 10px;
      font-size: 11px;
      font-weight: 600;
      color: #333;
      display: flex;
      align-items: center;
      justify-content: center;
      gap: 6px;
      background: #F5F7FA;
      border-bottom: 1px solid #E5E8ED;
      user-select: none;
    }}
    .header-collapsible summary::-webkit-details-marker {{ display: none; }}
    .header-collapsible summary::before {{ content: "▾"; color: #555; font-weight: 700; }}
    .header-collapsible:not([open]) summary::before {{ content: "▸"; }}
    .header-collapsible:not([open]) summary {{ border-bottom: 0; }}
    .header-collapsible-body {{ padding: 0 10px 8px; }}
    .report-header {{ display: flex; gap: 6px; align-items: center; justify-content: center; flex-wrap: nowrap; padding: 10px 0 2px; text-align: center; }}
    .header-left {{ flex: 0 1 340px; min-width: 260px; display: flex; flex-direction: column; align-items: flex-start; justify-content: center; text-align: left; }}
    .header-left h1 {{ margin: 0 0 6px 0; font-size: 26px; }}
    .header-left .subtitle {{ margin: 0 0 6px 0; font-size: 14px; color: #555; font-weight: 600; }}
    .header-left .meta {{ font-size: 14px; color: #444; text-align: left; }}
    .header-left .nn-note {{ text-align: left; }}
    .header-nn {{ display: flex; flex: 0 1 520px; min-width: 380px; justify-content: flex-start; align-items: center; }}
    .header-branding {{ display: flex; flex: 0 0 auto; min-width: 220px; justify-content: center; align-items: center; gap: 8px; flex-wrap: nowrap; }}
    .nn-sideways-bar-block {{
      display: flex;
      flex-direction: column;
      align-items: center;
      gap: 2px;
      min-width: 420px;
      max-width: 540px;
      width: min(540px, 52vw);
    }}
    .nn-sideways-bar-title-row {{
      display: flex;
      align-items: center;
      gap: 4px;
      justify-content: center;
    }}
    .nn-sideways-bar-title {{
      writing-mode: horizontal-tb;
      transform: none;
      font-size: 11px;
      color: #111;
      line-height: 1.15;
    }}
    .nn-sideways-bar-subtitle {{
      color: #444;
      font-weight: 500;
    }}
    .nn-help-icon {{
      position: relative;
      display: inline-flex;
      align-items: center;
      justify-content: center;
      width: 16px;
      height: 16px;
      border: 1px solid {BASE_BLUE};
      border-radius: 999px;
      color: {BASE_BLUE};
      font-size: 11px;
      font-weight: 700;
      cursor: help;
      user-select: none;
    }}
    .nn-help-icon:focus {{
      outline: 2px solid rgba(1,55,138,0.35);
      outline-offset: 2px;
    }}
    .nn-help-tooltip {{
      position: absolute;
      top: calc(100% + 8px);
      right: 0;
      width: min(430px, 80vw);
      background: #0F2F66;
      color: #FFF;
      border-radius: 8px;
      padding: 10px 12px;
      font-size: 12px;
      font-weight: 500;
      line-height: 1.4;
      text-align: left;
      box-shadow: 0 10px 24px rgba(0,0,0,0.25);
      opacity: 0;
      transform: translateY(-4px);
      transition: opacity 120ms ease, transform 120ms ease;
      pointer-events: none;
      z-index: 200;
    }}
    .nn-help-icon:hover .nn-help-tooltip,
    .nn-help-icon:focus .nn-help-tooltip,
    .nn-help-icon:focus-visible .nn-help-tooltip {{
      opacity: 1;
      transform: translateY(0);
    }}
    .profile-img {{ width: 96px; height: 96px; object-fit: cover; border-radius: 10px; border: 2px solid #EEE; background: #FFF; }}
    .company-logo-img {{ height: 52px; max-width: 180px; object-fit: contain; }}
    @media (max-width: 1200px) {{
      .report-header {{ flex-wrap: wrap; gap: 12px; }}
      .header-left {{ flex: 1 1 100%; }}
      .header-nn {{ flex: 1 1 100%; min-width: 300px; }}
      .nn-sideways-bar-block {{ min-width: 300px; width: min(540px, 94vw); }}
      .header-branding {{ min-width: 220px; justify-content: flex-start; }}
    }}
    .tabs {{ display: flex; gap: 8px; margin: 2px 0 8px; padding-bottom: 8px; flex-wrap: wrap; justify-content: center; align-items: center; }}
    .day-tabs {{ display: none; align-items: center; gap: 8px; }}
    .day-tabs.active {{ display: flex; }}
    .day-tabs label {{ font-size: 13px; color: #444; font-weight: 600; }}
    .day-tabs select {{ padding: 6px 10px; border: 1px solid #CCC; border-radius: 6px; background: #FFF; font-weight: 600; }}
    .month-tabs {{ display: none; }}
    .month-tabs.active {{ display: flex; }}
    .tab-btn {{ padding: 8px 16px; border: 1px solid #CCC; border-radius: 6px; background: #FFF; cursor: pointer; font-weight: 600; }}
    .tab-btn.month-btn {{ padding: 6px 12px; font-size: 13px; }}
    .tab-btn.active {{ background: dodgerblue; border-color: dodgerblue; color: #FFF; }}
    .tabs-row1 .tab-btn.active {{ background: royalblue; border-color: royalblue; }}
    .tabs-row2 .tab-btn.active {{ background: seagreen; border-color: seagreen; }}
    .tabs-row3 .tab-btn.active {{ background: darkorange; border-color: darkorange; }}
    .tab-panel {{ display: none; }}
    .tab-panel.active {{ display: block; }}
    .period-panel {{ display: none; }}
    .period-panel.active {{ display: block; }}
    .period-meta {{ display: none; }}
    .period-meta.active {{ display: inline; }}
    .period-note {{ display: none; }}
    .period-note.active {{ display: block; }}
    .period-note.active:empty {{ display: none; }}
    .period-nn {{ display: none; }}
    .period-nn.active {{ display: block; }}
    .hours-metrics {{ margin: 6px 0 16px; }}
    .hours-metrics:empty {{ display: none; margin: 0; }}
    .nn-note {{ margin-top: 6px; font-size: 13px; color: #8A3B3B; }}
    .hours-breakdown {{
      background: #FFF;
      border: 1px solid #DDD;
      border-radius: 10px;
      padding: 12px;
      box-shadow: 0 1px 2px rgba(0,0,0,0.05);
    }}
    details.hours-breakdown summary {{
      cursor: pointer;
      font-weight: 700;
      list-style: none;
      outline: none;
    }}
    details.hours-breakdown summary::-webkit-details-marker {{ display: none; }}
    details.hours-breakdown summary::before {{ content: "▸"; display: inline-block; width: 1em; color: {BASE_BLUE}; }}
    details.hours-breakdown[open] summary::before {{ content: "▾"; }}
    .hours-breakdown-header {{
      display: flex;
      align-items: baseline;
      justify-content: space-between;
      gap: 10px;
      flex-wrap: wrap;
      margin-bottom: 8px;
    }}
    .hours-breakdown h3 {{ margin: 0; font-size: 16px; }}
    .hours-breakdown-note {{ margin-top: 8px; font-size: 12px; color: #444; }}
    .hours-breakdown-list {{ display: flex; flex-direction: column; gap: 6px; }}
    .hours-project {{
      background: #FFF;
      border: 1px solid #EEE;
      border-radius: 8px;
      padding: 6px 10px;
    }}
    .hours-project[open] {{
      border-color: #CCC;
      box-shadow: 0 1px 2px rgba(0,0,0,0.05);
    }}
    .hours-project summary {{
      cursor: pointer;
      font-weight: 600;
      list-style: none;
      outline: none;
    }}
    .hours-project summary::-webkit-details-marker {{ display: none; }}
    .hours-project summary::before {{ content: "▸"; display: inline-block; width: 1em; color: {BASE_BLUE}; }}
    .hours-project[open] summary::before {{ content: "▾"; }}
    .hours-project-total {{ font-variant-numeric: tabular-nums; }}
    .hours-project-percent {{ color: #444; font-weight: 600; }}
    .hours-project-entries {{ margin-top: 8px; padding-left: 1.2em; }}
    .hours-entry-table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
    .hours-entry-table th {{ text-align: left; padding: 6px 6px; color: #555; border-bottom: 1px solid #EEE; }}
    .hours-entry-table td {{ padding: 6px 6px; border-bottom: 1px solid #F3F3F3; vertical-align: top; word-break: break-word; }}
    .hours-entry-duration {{ width: 110px; text-align: right; font-variant-numeric: tabular-nums; white-space: nowrap; }}
    .hours-entry-percent {{ width: 90px; text-align: right; font-variant-numeric: tabular-nums; white-space: nowrap; }}
    .hours-entry-empty {{ color: #777; font-style: italic; }}
    .nn-metrics {{
      display: flex;
      flex-wrap: wrap;
      gap: 12px 18px;
      font-size: 14px;
      background: #FFF;
      border: 1px solid #DDD;
      border-radius: 8px;
      padding: 10px 12px;
      box-shadow: 0 1px 2px rgba(0,0,0,0.05);
    }}

    /* Projects page styles (same as single report) */
    .projects-controls {{
      background: #FFF;
      border: 1px solid #DDD;
      border-radius: 10px;
      padding: 10px 12px;
      box-shadow: 0 1px 2px rgba(0,0,0,0.05);
      margin: 8px 0 12px;
    }}
    .projects-filters {{ display: flex; gap: 12px 16px; flex-wrap: wrap; align-items: flex-end; }}
    .projects-filter-block {{ display: flex; flex-direction: column; gap: 4px; }}
    .projects-filter-label {{ font-size: 12px; color: #444; font-weight: 600; }}
    .projects-status-buttons {{ display: flex; gap: 6px; flex-wrap: wrap; }}
    .projects-sort-buttons {{ display: flex; gap: 6px; flex-wrap: wrap; }}
    .tab-btn.status-btn {{ padding: 6px 10px; font-size: 13px; }}
    .tab-btn.sort-btn {{ padding: 6px 10px; font-size: 13px; }}
    .projects-checkbox {{ display: flex; gap: 6px; align-items: center; font-weight: 600; }}
    .project-last-log {{ color: #444; font-weight: 600; font-size: 12px; }}
    .project-summary-title {{ font-weight: 700; }}
    .project-progress {{
      background: #F7F9FC;
      border: 1px solid #E3E7EE;
      border-radius: 8px;
      padding: 8px 10px;
      margin-bottom: 10px;
    }}
    .project-progress-header {{
      display: flex;
      justify-content: space-between;
      gap: 10px;
      align-items: baseline;
      flex-wrap: wrap;
      font-size: 12px;
      font-weight: 700;
    }}
    .project-progress-right {{ color: #555; }}
    .project-progress-bar {{
      margin-top: 6px;
      height: 10px;
      background: #E4E8EE;
      border-radius: 999px;
      overflow: hidden;
    }}
    .project-progress-fill {{
      height: 100%;
      width: 0;
      background: {BASE_BLUE};
      border-radius: 999px;
    }}
    .project-progress-fill.over {{ background: {BASE_ORANGE}; }}
    .projects-filters select {{ padding: 6px 10px; border: 1px solid #CCC; border-radius: 6px; background: #FFF; font-weight: 600; }}
    .projects-list {{ display: flex; flex-direction: column; gap: 10px; margin-top: 10px; }}
    details.project-item {{ background: #FFF; border: 1px solid #DDD; border-radius: 10px; padding: 8px 10px; box-shadow: 0 1px 2px rgba(0,0,0,0.05); }}
    details.project-item summary {{ cursor: pointer; font-weight: 700; list-style: none; outline: none; }}
    details.project-item summary::-webkit-details-marker {{ display: none; }}
    details.project-item summary::before {{ content: "▸"; display: inline-block; width: 1em; color: {BASE_BLUE}; }}
    details.project-item[open] summary::before {{ content: "▾"; }}
    .project-expanded {{ margin-top: 10px; display: flex; gap: 14px; align-items: flex-start; flex-wrap: wrap; }}
    .project-col {{ min-width: 280px; }}
    .project-col-table {{ flex: 1 1 360px; }}
    .project-col-text {{ flex: 1 1 320px; }}
    .project-col-images {{ flex: 1 1 320px; }}
    .project-info-table {{ width: 100%; border-collapse: collapse; font-size: 12px; background: #FFF; }}
    .project-info-table td {{ padding: 4px 6px; border-bottom: 1px solid #EEE; vertical-align: top; word-break: break-word; }}
    .project-info-table td:first-child {{ width: 42%; color: #555; }}
    .project-timelog {{ margin-top: 10px; }}
    .project-timelog-title {{ font-weight: 700; margin: 10px 0 6px; }}
    .project-timelog-table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
    .project-timelog-table th {{ text-align: left; padding: 6px 6px; color: #555; border-bottom: 1px solid #EEE; }}
    .project-timelog-table td {{ padding: 6px 6px; border-bottom: 1px solid #F3F3F3; vertical-align: top; word-break: break-word; }}
    .timelog-date {{ width: 110px; white-space: nowrap; font-variant-numeric: tabular-nums; }}
    .timelog-duration {{ width: 100px; text-align: right; white-space: nowrap; font-variant-numeric: tabular-nums; }}
    .project-activitytype {{ margin-top: 10px; }}
    .project-activitytype-pie {{
      width: 140px;
      height: 140px;
      border-radius: 50%;
      border: 1px solid #DDD;
      margin: 8px 0 10px;
      box-shadow: inset 0 0 0 1px rgba(255,255,255,0.7);
    }}
    .project-activitytype-legend {{ display: flex; flex-direction: column; gap: 4px; font-size: 12px; }}
    .project-activitytype-row {{ display: grid; grid-template-columns: 12px 1fr; gap: 8px; align-items: center; }}
    .project-activitytype-dot {{ width: 10px; height: 10px; border-radius: 999px; display: inline-block; }}
    .project-activitytype-label {{ color: #333; }}
    .deliverable-fn {{ font-weight: 700; margin: 8px 0 6px; }}
    .deliverables-text pre {{ white-space: pre-wrap; word-break: break-word; max-height: 420px; overflow: auto; background: #FAFAFA; border: 1px solid #EEE; border-radius: 8px; padding: 8px; margin: 0; font-size: 12px; }}
    .deliverables-images img {{ max-width: 420px; height: auto; border-radius: 8px; border: 1px solid #EEE; background: #FFF; }}
    .plot-sections {{ display: flex; flex-direction: column; gap: 10px; }}
    .plot-section {{ background: #FFF; border: 1px solid #DDD; border-radius: 10px; box-shadow: 0 1px 2px rgba(0,0,0,0.05); }}
    .plot-section-title {{ font-size: 16px; font-weight: 700; }}
    .plot-section-fixed {{ padding: 10px 12px 12px; }}
    .plot-section-fixed .plot-section-title {{ margin-bottom: 8px; }}
    .plot-section-body {{ min-height: 20px; }}
    details.plot-section summary {{ cursor: pointer; list-style: none; outline: none; padding: 10px 12px; display: flex; align-items: center; gap: 6px; }}
    details.plot-section summary::-webkit-details-marker {{ display: none; }}
    details.plot-section summary::before {{ content: "▸"; display: inline-block; width: 1em; color: {BASE_BLUE}; font-weight: 700; }}
    details.plot-section[open] summary::before {{ content: "▾"; }}
    details.plot-section .plot-section-body {{ padding: 0 12px 12px; }}
    .weekly-guidance {{ margin-top: 10px; padding: 10px 12px; background: #FFF; border: 1px solid #DDD; border-radius: 8px; font-size: 13px; color: #333; line-height: 1.45; display: flex; flex-direction: column; gap: 6px; }}
    .weekly-guidance-formula {{ font-family: Consolas, "Courier New", monospace; background: #F7F9FC; border: 1px solid #E6EBF2; border-radius: 6px; padding: 6px 8px; color: #223; }}
  </style>
</head>
<body>
  <div class="page">
    <div class="sticky-header">
      <details class="header-collapsible" id="header-collapsible" open>
        <summary>{header_summary_text}</summary>
        <div class="header-collapsible-body">
          <div class="report-header">
            <div class="header-left">
              <h1>{title_text}</h1>
              <div class="subtitle">{person_name_text}</div>
              <div class="meta">{period_meta_html}</div>
              <div class="meta">Generated: {export_date}</div>
              {period_note_html}
            </div>
            <div class="header-nn">
              {sideways_bar_chart_blocks_html}
            </div>
            {header_branding_html}
          </div>
        </div>
      </details>

      <div class="tabs tabs-row1">
        {tab_buttons_html}
      </div>

      <div id="period-controls">
        <div id="period-controls-inner">
          <div class="tabs tabs-row2">
            {period_group_buttons_html}
          </div>
          <div class="tabs tabs-row3 day-tabs{' active' if default_group == 'daily' else ''}" id="day-tabs">
            <label for="day-select">Day</label>
            <select id="day-select" onchange="showDay(this.value)">
              {day_select_html}
            </select>
          </div>
          <div class="tabs tabs-row3 month-tabs{' active' if default_group == 'monthly' else ''}" id="month-tabs">
            {month_buttons_html}
          </div>
        </div>
        <div id="projects-controls-top" style="display:none">
          {projects_filters_html}
        </div>
      </div>
    </div>

    {tab_panels_html}
  </div>

  <script>
    var currentTab = "{default_tab}";
    var enabledTabs = {enabled_tabs_norm!r};
    var currentPeriodId = "{default_period_id}";
		    var currentDailyId = "{default_day_id}";
		    var dailyPeriodIds = {daily_period_ids!r};
		    var currentMonthlyId = "{default_month_id}";
		    var projectsStatusFilter = "";
    var projectsSortKey = "last_updated";
    var projectsSortDesc = true;

    function resizePlotlyIn(rootEl) {{
      if (!rootEl || !window.Plotly) {{
        return;
      }}
      rootEl.querySelectorAll(".plotly-graph-div").forEach(function(plotEl) {{
        Plotly.Plots.resize(plotEl);
      }});
    }}

    function bindPlotSectionToggles() {{
      document.querySelectorAll("details.plot-section").forEach(function(secEl) {{
        secEl.addEventListener("toggle", function() {{
          if (secEl.open) {{
            resizePlotlyIn(secEl);
          }}
        }});
      }});
    }}

    function bindHeaderToggle() {{
      var headerToggle = document.getElementById("header-collapsible");
      if (!headerToggle) {{
        return;
      }}
      headerToggle.addEventListener("toggle", function() {{
        if (!headerToggle.open) {{
          return;
        }}
        resizePlotlyIn(headerToggle);
        updateView();
      }});
    }}

    function showTab(name) {{
      currentTab = name;
      updateView();
    }}

	    function showPeriodGroup(group) {{
	      if (group === "daily") {{
	        if (!currentDailyId && dailyPeriodIds.length > 0) {{
	          currentDailyId = dailyPeriodIds[0];
	        }}
	        if (currentDailyId) {{
	          currentPeriodId = currentDailyId;
	        }}
	      }} else if (group === "monthly") {{
	        if (currentMonthlyId) {{
	          currentPeriodId = currentMonthlyId;
	        }}
	      }} else {{
	        currentPeriodId = group;
	      }}
	      updateView();
	    }}

	    function showDay(dayId) {{
	      if (!dayId) {{
	        return;
	      }}
	      currentDailyId = dayId;
	      currentPeriodId = dayId;
	      updateView();
	    }}

	    function showMonth(monthId) {{
	      currentMonthlyId = monthId;
	      currentPeriodId = monthId;
	      updateView();
	    }}

    function setStatusFilter(statusVal) {{
      projectsStatusFilter = statusVal || "";
      document.querySelectorAll(".status-btn").forEach(function(btn) {{
        btn.classList.remove("active");
      }});
      document.querySelectorAll(".status-btn").forEach(function(btn) {{
        if ((btn.getAttribute("data-status") || "") === projectsStatusFilter) {{
          btn.classList.add("active");
        }}
      }});
      applyProjectsFilters();
    }}

    function setSort(sortKey) {{
      var key = sortKey || "last_updated";
      if (projectsSortKey === key) {{
        projectsSortDesc = !projectsSortDesc;
      }} else {{
        projectsSortKey = key;
        projectsSortDesc = true;
      }}
      applyProjectsFilters();
    }}

    function refreshSortButtons() {{
      document.querySelectorAll(".sort-btn").forEach(function(btn) {{
        var key = (btn.getAttribute("data-sort") || "");
        var label = (btn.getAttribute("data-label") || btn.textContent || "").replace(/[\\s]*[▲▼]$/, "");
        var active = (key === projectsSortKey);
        btn.classList.toggle("active", active);
        btn.textContent = active ? (label + (projectsSortDesc ? " ▼" : " ▲")) : label;
      }});
    }}

    function applyProjectsFilters() {{
      var prio = document.getElementById("filter-priority");
      var mag = document.getElementById("filter-magnitude");
      var prog = document.getElementById("filter-programma");
      var hasDelivEl = document.getElementById("filter-has-deliverables");
      var prioVal = prio ? (prio.value || "") : "";
      var magVal = mag ? (mag.value || "") : "";
      var progVal = prog ? (prog.value || "") : "";
      var hasDeliv = hasDelivEl ? !!hasDelivEl.checked : false;
      var sortVal = projectsSortKey || "last_updated";
      var sortDesc = (projectsSortDesc !== false);
      var listEl = document.querySelector(".projects-list");
      var items = Array.prototype.slice.call(document.querySelectorAll(".project-item"));

      items.forEach(function(item) {{
        var ok = true;
        var s = (item.getAttribute("data-status") || "");
        var p = (item.getAttribute("data-priority") || "");
        var m = (item.getAttribute("data-magnitude") || "");
        var progs = (item.getAttribute("data-programmas") || "");
        var hd = (item.getAttribute("data-has-deliverables") || "0");

        if (projectsStatusFilter) {{
          ok = ok && (s.toLowerCase() === projectsStatusFilter.toLowerCase());
        }}
        if (prioVal) {{
          ok = ok && (p.toLowerCase() === prioVal.toLowerCase());
        }}
        if (magVal) {{
          ok = ok && (m.toLowerCase() === magVal.toLowerCase());
        }}
        if (progVal) {{
          var tokens = progs.split(",").map(function(x) {{ return x.trim().toLowerCase(); }});
          ok = ok && (tokens.indexOf(progVal.toLowerCase()) >= 0);
        }}
        if (hasDeliv) {{
          ok = ok && (hd === "1");
        }}
        item.style.display = ok ? "" : "none";
      }});

      var cmpNum = function(a, b, attr, desc) {{
        var av = parseFloat(a.getAttribute(attr) || "");
        var bv = parseFloat(b.getAttribute(attr) || "");
        var aMissing = !isFinite(av) || av < 0;
        var bMissing = !isFinite(bv) || bv < 0;
        if (aMissing && bMissing) return 0;
        if (aMissing) return 1;
        if (bMissing) return -1;
        return desc ? (bv - av) : (av - bv);
      }};

      var cmpProjectId = function(a, b) {{
        var aid = (a.getAttribute("data-project-id") || "").toLowerCase();
        var bid = (b.getAttribute("data-project-id") || "").toLowerCase();
        if (aid < bid) return -1;
        if (aid > bid) return 1;
        return 0;
      }};

      items.sort(function(a, b) {{
        var diff = 0;
        if (sortVal === "hours_reported") {{
          diff = cmpNum(a, b, "data-hours-reported", sortDesc);
        }} else if (sortVal === "hours_cap") {{
          diff = cmpNum(a, b, "data-hours-cap", sortDesc);
        }} else if (sortVal === "pct_completed") {{
          diff = cmpNum(a, b, "data-pct-completed", sortDesc);
        }} else if (sortVal === "estimated_magnitude") {{
          diff = cmpNum(a, b, "data-magnitude-rank", sortDesc);
        }} else if (sortVal === "start_date") {{
          diff = cmpNum(a, b, "data-start-date", sortDesc);
        }} else if (sortVal === "priority") {{
          diff = cmpNum(a, b, "data-priority-rank", sortDesc);
        }} else {{
          diff = cmpNum(a, b, "data-last-updated", sortDesc);
        }}
        if (diff !== 0) return diff;
        return cmpProjectId(a, b);
      }});

      if (listEl) {{
        items.forEach(function(item) {{
          listEl.appendChild(item);
        }});
      }}
      refreshSortButtons();
    }}

    function updateView() {{
      enabledTabs.forEach(function(t) {{
        var tab = document.getElementById("tab-" + t);
        var btn = document.getElementById("btn-" + t);
        if (tab) tab.classList.toggle("active", t === currentTab);
        if (btn) btn.classList.toggle("active", t === currentTab);
      }});

      var periodInner = document.getElementById("period-controls-inner");
      var projectsControlsTop = document.getElementById("projects-controls-top");
      if (periodInner && projectsControlsTop) {{
        periodInner.style.display = (currentTab === "projects") ? "none" : "";
        projectsControlsTop.style.display = (currentTab === "projects") ? "" : "none";
      }} else {{
        var periodControls = document.getElementById("period-controls");
        if (periodControls) {{
          periodControls.style.display = (currentTab === "projects") ? "none" : "";
        }}
      }}

	      if (currentTab !== "projects") {{
	        if (currentPeriodId && currentPeriodId.startsWith("daily-")) {{
	          currentDailyId = currentPeriodId;
	        }}
	        var isMonthly = currentPeriodId && currentPeriodId.startsWith("monthly-");
	        var isDaily = currentPeriodId && currentPeriodId.startsWith("daily-");
	        var activeGroup = isMonthly ? "monthly" : (isDaily ? "daily" : currentPeriodId);

	        document.querySelectorAll(".period-btn").forEach(function(btn) {{
	          btn.classList.remove("active");
	        }});
        var activeBtn = document.getElementById("btn-period-" + activeGroup);
	        if (activeBtn) {{
	          activeBtn.classList.add("active");
	        }}

	        var dayTabs = document.getElementById("day-tabs");
	        if (dayTabs) {{
	          dayTabs.classList.toggle("active", isDaily);
	        }}
	        var daySelect = document.getElementById("day-select");
	        if (daySelect && currentDailyId) {{
	          daySelect.value = currentDailyId;
	        }}

	        var monthTabs = document.getElementById("month-tabs");
	        if (monthTabs) {{
	          monthTabs.classList.toggle("active", isMonthly);
	        }}
        document.querySelectorAll(".month-btn").forEach(function(btn) {{
          btn.classList.remove("active");
        }});
        var activeMonthBtn = document.getElementById("btn-month-" + currentMonthlyId);
        if (activeMonthBtn) {{
          activeMonthBtn.classList.add("active");
        }}

        document.querySelectorAll(".period-panel").forEach(function(panel) {{
          panel.classList.remove("active");
        }});
        var countsPanel = document.getElementById("period-counts-" + currentPeriodId);
        var hoursPanel = document.getElementById("period-hours-" + currentPeriodId);
        var percentagePanel = document.getElementById("period-percentage-" + currentPeriodId);
        if (countsPanel) countsPanel.classList.add("active");
        if (hoursPanel) hoursPanel.classList.add("active");
        if (percentagePanel) percentagePanel.classList.add("active");

        document.querySelectorAll(".period-meta").forEach(function(el) {{
          el.classList.remove("active");
        }});
        var metaEl = document.getElementById("meta-" + currentPeriodId);
        if (metaEl) metaEl.classList.add("active");

        document.querySelectorAll(".period-note").forEach(function(el) {{
          el.classList.remove("active");
        }});
        var noteEl = document.getElementById("nn-note-" + currentPeriodId);
        if (noteEl) noteEl.classList.add("active");

        document.querySelectorAll(".period-nn").forEach(function(el) {{
          el.classList.remove("active");
        }});
        var sidewaysBarEl = document.getElementById("nn-sideways-bar-block-" + currentPeriodId);
        if (sidewaysBarEl) sidewaysBarEl.classList.add("active");

        var activePanel = document.getElementById("period-" + currentTab + "-" + currentPeriodId);
        if (activePanel) {{
          resizePlotlyIn(activePanel);
        }}

        var sidewaysBarFigEl = document.getElementById("nn-sideways-bar-" + currentPeriodId);
        if (sidewaysBarFigEl && window.Plotly) {{
          Plotly.Plots.resize(sidewaysBarFigEl);
        }}
      }} else {{
        applyProjectsFilters();
      }}
    }}

    window.addEventListener("load", function() {{
      bindPlotSectionToggles();
      bindHeaderToggle();
      updateView();
    }});
  </script>
</body>
</html>
"""

    with open(out_html_path, "w", encoding="utf-8") as f:
        f.write(html_content)


def _set_projectinfo_kv(ws: Any, key: str, value: Any) -> None:
    for row_idx in range(1, ws.max_row + 1):
        key_cell = ws.cell(row=row_idx, column=1).value
        if str(key_cell).strip() == key:
            ws.cell(row=row_idx, column=2).value = value
            return
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=1).value = key
    ws.cell(row=new_row, column=2).value = value


def _autofill_dummy_projects_if_needed(asof_date: date) -> None:
    """
    When dummy fallback is active, ensure a richer demo dataset:
    - at least 9 dummy projects (adds 5 beyond the original 4)
    - generated time logs up to `asof_date` with ~32h/week per project
    """
    if not USING_DUMMY_FALLBACK:
        return

    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        print(f"WARNING: openpyxl unavailable; skipping dummy-project auto-fill ({exc}).")
        return

    templates_dir = resolve_path(CONFIG, CONFIG.get("paths", {}).get("templates_dir", "Templates"))
    project_info_template = os.path.join(templates_dir, "project_info_template.xlsx")
    time_log_template = os.path.join(templates_dir, "time_log_template.xlsx")
    if not os.path.exists(project_info_template) or not os.path.exists(time_log_template):
        print("WARNING: Templates not found; skipping dummy-project auto-fill.")
        return

    os.makedirs(DUMMY_PROJECTEN_DIR, exist_ok=True)

    specs: List[Dict[str, Any]] = [
        dict(counter=9999, slug="dummy_project_1", project_name="Dummy Project 1", programma="Athlete Support", theme="Monitoring", requester="Performance Team", owner="John Doe", priority="High"),
        dict(counter=9998, slug="dummy_project_2", project_name="Dummy Project 2", programma="Research", theme="Data", requester="Innovation Team", owner="John Doe", priority="Medium"),
        dict(counter=9997, slug="dummy_project_3", project_name="Dummy Project 3", programma="Operations", theme="Planning", requester="Program Office", owner="John Doe", priority="Medium"),
        dict(counter=9996, slug="dummy_project_4", project_name="Dummy Project 4", programma="Education", theme="Knowledge", requester="Coaching Team", owner="John Doe", priority="Low"),
        dict(counter=9995, slug="dummy_project_5", project_name="Dummy Project 5", programma="Athlete Support", theme="Health", requester="Medical Team", owner="John Doe", priority="High"),
        dict(counter=9994, slug="dummy_project_6", project_name="Dummy Project 6", programma="Research", theme="Experiment", requester="Science Team", owner="John Doe", priority="Critical"),
        dict(counter=9993, slug="dummy_project_7", project_name="Dummy Project 7", programma="Operations", theme="Automation", requester="Program Office", owner="John Doe", priority="Medium"),
        dict(counter=9992, slug="dummy_project_8", project_name="Dummy Project 8", programma="Education", theme="Workshops", requester="Coaching Team", owner="John Doe", priority="Low"),
        dict(counter=9991, slug="dummy_project_9", project_name="Dummy Project 9", programma="Athlete Support", theme="Performance", requester="Performance Team", owner="John Doe", priority="High"),
    ]

    base_start = date(asof_date.year, 1, 6)  # first Monday-ish of the year
    activity_types = ["Analysis", "Coordination", "Documentation", "Implementation", "Review"]

    for idx, spec in enumerate(specs):
        project_id = f"{asof_date.year:04d}_{int(spec['counter']):04d}"
        folder_name = f"{project_id}_{spec['slug']}"
        project_dir = os.path.join(DUMMY_PROJECTEN_DIR, folder_name)
        deliverables_dir = os.path.join(project_dir, "Deliverables")
        project_info_path = os.path.join(project_dir, "project_info.xlsx")
        time_log_path = os.path.join(project_dir, "time_log.xlsx")

        os.makedirs(project_dir, exist_ok=True)
        os.makedirs(deliverables_dir, exist_ok=True)

        if not os.path.exists(project_info_path):
            shutil.copy2(project_info_template, project_info_path)
        if not os.path.exists(time_log_path):
            shutil.copy2(time_log_template, time_log_path)

        start_day = base_start + timedelta(days=idx * 7)
        if start_day > asof_date:
            start_day = asof_date - timedelta(days=7)

        # Keep project_info aligned with folder metadata.
        info_wb = load_workbook(project_info_path)
        if "ProjectInfo" in info_wb.sheetnames:
            info_ws = info_wb["ProjectInfo"]
            _set_projectinfo_kv(info_ws, "project_id", project_id)
            _set_projectinfo_kv(info_ws, "project_name", spec["project_name"])
            _set_projectinfo_kv(info_ws, "programma (if multiple, separate by |)", spec["programma"])
            _set_projectinfo_kv(info_ws, "theme (if multiple, separate by |)", spec["theme"])
            _set_projectinfo_kv(info_ws, "owner", spec["owner"])
            _set_projectinfo_kv(info_ws, "requester", spec["requester"])
            _set_projectinfo_kv(info_ws, "status", "Active")
            _set_projectinfo_kv(info_ws, "priority", spec["priority"])
            _set_projectinfo_kv(info_ws, "start_date", start_day)
            _set_projectinfo_kv(info_ws, "target_end_date", "")
            _set_projectinfo_kv(info_ws, "actual_end_date", "")
            _set_projectinfo_kv(info_ws, "last_updated", asof_date)
            _set_projectinfo_kv(info_ws, "notes", "Auto-generated demo data for onboarding.")
            info_wb.save(project_info_path)

        # Ensure a basic deliverables note exists.
        milestones_path = os.path.join(deliverables_dir, "milestones.txt")
        if not os.path.exists(milestones_path):
            with open(milestones_path, "w", encoding="utf-8") as f:
                f.write(
                    "\n".join(
                        [
                            f"Project: {project_id} - {spec['project_name']}",
                            "",
                            "Milestones",
                            "1. Kickoff completed.",
                            "2. Draft deliverable reviewed.",
                            "3. Final deliverable shared.",
                            "",
                            "Notes",
                            "- Demo content generated automatically.",
                        ]
                    )
                    + "\n"
                )

        # Rebuild timelog with deterministic 32h/week pattern up to as-of date.
        tl_wb = load_workbook(time_log_path)
        if "TimeLog" not in tl_wb.sheetnames:
            continue
        tl_ws = tl_wb["TimeLog"]
        tl_ws["B1"] = project_id
        tl_ws["B2"] = spec["project_name"]
        tl_ws["B3"] = spec["programma"]

        if tl_ws.max_row >= 7:
            for row_cells in tl_ws.iter_rows(min_row=7, max_row=tl_ws.max_row, min_col=1, max_col=10):
                for cell in row_cells:
                    cell.value = None

        current = start_day - timedelta(days=start_day.weekday())
        week_idx = 0
        while current <= asof_date:
            for day_offset in (0, 1, 2, 3):  # 4 workdays * 8h = 32h/week
                entry_date = current + timedelta(days=day_offset)
                if entry_date < start_day or entry_date > asof_date:
                    continue
                activity = activity_types[(week_idx + day_offset + idx) % len(activity_types)]
                tl_ws.append(
                    [
                        entry_date,
                        "09:00",
                        "17:00",
                        480,
                        activity,
                        f"{activity} work package for {spec['project_name']}",
                        "",
                        "Continue next iteration",
                        "dummy|autogen",
                        "Office",
                    ]
                )
            current += timedelta(days=7)
            week_idx += 1

        tl_wb.save(time_log_path)

    print(f"Dummy fallback active: ensured {len(specs)} demo projects and regenerated time logs up to {asof_date.isoformat()}.")


def generate_reports(report_type: str, asof_date: date) -> None:
    """Load project data and generate configured report outputs."""
    export_date = date.today().isoformat()
    print(f"As-of date used: {asof_date.isoformat()}")

    _autofill_dummy_projects_if_needed(asof_date)
    projects_df, time_entries_df, project_info_map, deliverables_map = load_and_validate_projects(PROJECTEN_DIR)
    if projects_df.empty:
        raise SystemExit(f"No project folders found under: {PROJECTEN_DIR}")

    complete_missing_hours_project_ids = collect_project_ids_by_role(
        projects_df,
        PROJECT_ROLE_COMPLETE_MISSING_HOURS,
    )
    projects_overview_df = filter_projects_excluding_project_ids(
        projects_df,
        complete_missing_hours_project_ids,
    )
    time_entries_for_overview = filter_time_entries_excluding_project_ids(
        time_entries_df,
        complete_missing_hours_project_ids,
    )
    project_info_map_for_overview = {
        pid: info
        for pid, info in project_info_map.items()
        if str(pid).strip() not in complete_missing_hours_project_ids
    }
    deliverables_map_for_overview = {
        pid: payload
        for pid, payload in deliverables_map.items()
        if str(pid).strip() not in complete_missing_hours_project_ids
    }

    os.makedirs(REPORT_DIR, exist_ok=True)
    os.makedirs(REPORTS_ARCHIVE_DIR, exist_ok=True)

    projects_overview_df.to_csv(os.path.join(REPORT_DIR, "projects_overview.csv"), index=False)
    time_entries_df.to_csv(os.path.join(REPORT_DIR, "time_entries_df.csv"), index=False)

    periods = compute_report_periods(asof_date)
    header_assets = build_header_assets()

    filters_html, projects_list_html = build_projects_page_html(
        projects_overview_df,
        time_entries_for_overview,
        project_info_map_for_overview,
        deliverables_map_for_overview,
    )
    projects_page_html = filters_html + projects_list_html

    nn_df, nn_path, nn_status = load_nn_maandelijks_df(asof_date=asof_date, all_time_entries_df=time_entries_df)
    print(nn_status)
    _, project_color_map = build_color_maps(projects_df)
    timeline_year = asof_date.year

    if report_type in ("combined", "all"):
        period_payloads: Dict[str, Dict[str, Any]] = {}

        for day_info in list_available_day_periods(asof_date, time_entries_df):
            period_start = day_info["start"]
            period_end = day_info["end"]
            day_key = day_info["key"]
            period_id = f"daily-{day_key}"
            period_label = day_info["label"]

            time_entries_filtered = filter_time_entries_by_period(time_entries_df, period_start, period_end)

            nn_note = None
            if nn_df is None:
                nn_note = nn_status
            table_only_html = build_logged_hours_breakdown_html(
                time_entries_filtered,
                show_percentage=True,
                include_total_in_note=True,
                foldable=False,
            )

            period_payloads[period_id] = dict(
                label=period_label,
                period_range=format_period_range_compact(period_start, period_end),
                counts_fig=go.Figure(),
                hours_fig=go.Figure(),
                percentage_fig=go.Figure(),
                show_plots=False,
                table_only_html=table_only_html,
                hours_metrics_html=table_only_html,
                percentage_metrics_html=table_only_html,
                percentage_explanation_html="",
                percentage_section_payloads=[],
                sideways_bar_chart_html="",
                nn_note=nn_note,
                nn_summary=None,
            )

        for rtype in ("weekly", "biweekly", "yearly"):
            period_info = periods[rtype]
            period_start = period_info["start"]
            period_end = period_info["end"]
            period_label = period_info["label"]

            time_entries_filtered = filter_time_entries_by_period(time_entries_df, period_start, period_end)

            nn_summary = None
            nn_note = None
            sideways_bar_chart_html = ""
            hours_metrics_html = ""
            percentage_metrics_html = ""
            if nn_df is None:
                nn_note = nn_status
            else:
                nn_summary, nn_note = compute_nn_summary(
                    nn_df,
                    "yearly",
                    period_end,
                    time_entries_filtered,
                    all_time_entries_df=time_entries_df,
                    asof_date=asof_date,
                )
                if nn_note:
                    nn_note = f"Hours-remaining source: {nn_note}"
            sideways_bar_chart_html = build_nn_sideways_bar_chart_html(nn_summary, div_id=f"nn-sideways-bar-{rtype}")

            if rtype == "yearly":
                hours_metrics_html = build_nn_metrics_html(nn_summary, nn_note)
                percentage_metrics_html = hours_metrics_html
            else:
                hours_metrics_html = build_logged_hours_breakdown_html(time_entries_filtered)
                percentage_metrics_html = build_logged_hours_breakdown_html(time_entries_filtered, show_percentage=True)

            projects_for_counts = projects_df
            if rtype in ("weekly", "biweekly"):
                projects_for_counts = filter_projects_with_hours(projects_df, time_entries_filtered)

            counts_fig = build_counts_figure(
                projects_for_counts,
                export_date,
                period_start,
                period_end,
                period_label,
                project_color_map=project_color_map,
                timeline_projects_df=projects_df,
                timeline_year=timeline_year,
            )
            hours_fig = build_hours_figure(
                projects_df,
                time_entries_filtered,
                export_date,
                period_start,
                period_end,
                period_label,
                report_type=rtype,
                exclude_project_ids=complete_missing_hours_project_ids,
            )
            total_period_hours = (
                float(pd.to_numeric(time_entries_filtered["duration_hours"], errors="coerce").fillna(0.0).sum())
                if not time_entries_filtered.empty and "duration_hours" in time_entries_filtered.columns
                else 0.0
            )
            weekly_ref, weekly_ref_source, weekly_ref_note = compute_weekly_reference_hours(
                time_entries_df, period_end, nn_summary
            )
            weekly_progress_guidance: Optional[Dict[str, Any]] = None
            percentage_explanation_html = ""
            if rtype == "yearly":
                weekly_progress_guidance = compute_weekly_progress_guidance(
                    time_entries_df,
                    period_end,
                    nn_summary,
                    weekly_ref,
                )
                percentage_explanation_html = build_weekly_progress_explanation_html(
                    weekly_progress_guidance,
                    weekly_ref_source,
                )
            percentage_fig = build_percentage_figure_from_hours(
                hours_fig,
                total_period_hours=total_period_hours,
                weekly_reference_hours=weekly_ref,
                weekly_reference_note=weekly_ref_note,
                show_weekly_reference_note_in_title=False,
                weekly_progress_guidance=weekly_progress_guidance,
            )
            percentage_section_payloads = build_percentage_section_payloads(
                build_hours_section_figures(
                    projects_df,
                    time_entries_filtered,
                    period_start,
                    period_end,
                    rtype,
                    exclude_project_ids=complete_missing_hours_project_ids,
                ),
                total_period_hours=total_period_hours,
                weekly_reference_hours=weekly_ref,
                weekly_progress_guidance=weekly_progress_guidance,
                timeline_explanation_html=percentage_explanation_html,
            )

            period_payloads[rtype] = dict(
                label=period_label,
                period_range=format_period_range_compact(period_start, period_end),
                counts_fig=counts_fig,
                hours_fig=hours_fig,
                percentage_fig=percentage_fig,
                hours_metrics_html=hours_metrics_html,
                percentage_metrics_html=percentage_metrics_html,
                percentage_explanation_html=percentage_explanation_html,
                percentage_section_payloads=percentage_section_payloads,
                sideways_bar_chart_html=sideways_bar_chart_html,
                nn_note=nn_note,
                nn_summary=nn_summary,
            )

        for month_info in list_completed_month_periods(asof_date, time_entries_df):
            period_start = month_info["start"]
            period_end = month_info["end"]
            month_key = month_info["key"]
            period_id = f"monthly-{month_key}"
            period_label = month_info["label"]

            time_entries_filtered = filter_time_entries_by_period(time_entries_df, period_start, period_end)

            nn_summary = None
            nn_note = None
            sideways_bar_chart_html = ""
            hours_metrics_html = ""
            percentage_metrics_html = ""
            if nn_df is None:
                nn_note = nn_status
            else:
                nn_summary, nn_note = compute_nn_summary(
                    nn_df,
                    "monthly",
                    period_end,
                    time_entries_filtered,
                    all_time_entries_df=time_entries_df,
                    asof_date=asof_date,
                )
                if nn_note:
                    nn_note = f"Hours-remaining source: {nn_note}"
            sideways_bar_chart_html = build_nn_sideways_bar_chart_html(nn_summary, div_id=f"nn-sideways-bar-{period_id}")
            hours_metrics_html = build_nn_metrics_html(nn_summary, nn_note)
            percentage_metrics_html = hours_metrics_html

            projects_for_counts = filter_projects_with_hours(projects_df, time_entries_filtered)
            counts_fig = build_counts_figure(
                projects_for_counts,
                export_date,
                period_start,
                period_end,
                period_label,
                project_color_map=project_color_map,
                timeline_projects_df=projects_df,
                timeline_year=timeline_year,
            )
            hours_fig = build_hours_figure(
                projects_df,
                time_entries_filtered,
                export_date,
                period_start,
                period_end,
                period_label,
                report_type="monthly",
                exclude_project_ids=complete_missing_hours_project_ids,
            )
            total_period_hours = (
                float(pd.to_numeric(time_entries_filtered["duration_hours"], errors="coerce").fillna(0.0).sum())
                if not time_entries_filtered.empty and "duration_hours" in time_entries_filtered.columns
                else 0.0
            )
            weekly_ref, weekly_ref_source, weekly_ref_note = compute_weekly_reference_hours(
                time_entries_df, period_end, nn_summary
            )
            weekly_progress_guidance = compute_monthly_average_guidance(
                time_entries_filtered,
                period_start,
                period_end,
                weekly_ref,
                nn_summary=nn_summary,
            )
            percentage_explanation_html = build_monthly_average_explanation_html(
                weekly_progress_guidance,
            )
            percentage_fig = build_percentage_figure_from_hours(
                hours_fig,
                total_period_hours=total_period_hours,
                weekly_reference_hours=weekly_ref,
                weekly_reference_note=weekly_ref_note,
                show_weekly_reference_note_in_title=False,
                weekly_progress_guidance=weekly_progress_guidance,
            )
            percentage_section_payloads = build_percentage_section_payloads(
                build_hours_section_figures(
                    projects_df,
                    time_entries_filtered,
                    period_start,
                    period_end,
                    "monthly",
                    exclude_project_ids=complete_missing_hours_project_ids,
                ),
                total_period_hours=total_period_hours,
                weekly_reference_hours=weekly_ref,
                weekly_progress_guidance=weekly_progress_guidance,
                timeline_explanation_html=percentage_explanation_html,
            )

            period_payloads[period_id] = dict(
                label=period_label,
                period_range=format_period_range_compact(period_start, period_end),
                counts_fig=counts_fig,
                hours_fig=hours_fig,
                percentage_fig=percentage_fig,
                hours_metrics_html=hours_metrics_html,
                percentage_metrics_html=percentage_metrics_html,
                percentage_explanation_html=percentage_explanation_html,
                percentage_section_payloads=percentage_section_payloads,
                sideways_bar_chart_html=sideways_bar_chart_html,
                nn_note=nn_note,
                nn_summary=nn_summary,
            )

        header_context = dict(
            title_text=REPORT_TITLE,
            person_name=PERSON_NAME,
            export_date=export_date,
            **header_assets,
        )

        base_name = "project_report"
        archive_base_name = f"project_report_asof_{asof_date.isoformat()}"
        html_path, lite_html_path = export_multi_period_report(
            period_payloads,
            REPORT_DIR,
            REPORTS_ARCHIVE_DIR,
            base_name,
            archive_base_name,
            export_date,
            header_context,
            projects_list_html,
            projects_filters_html=filters_html,
        )
        print(f"Generated combined report -> {html_path}")
        print(f"Generated lite report -> {lite_html_path}")
        return

    rtype = report_type
    periods_local = compute_report_periods(asof_date)
    if rtype not in periods_local:
        raise SystemExit(f"Unknown report type: {rtype}")

    period_info = periods_local[rtype]
    period_start = period_info["start"]
    period_end = period_info["end"]
    period_label = period_info["label"]
    period_key = period_info["key"]

    time_entries_filtered = filter_time_entries_by_period(time_entries_df, period_start, period_end)

    nn_summary = None
    nn_note = None
    sideways_bar_chart_html = ""
    hours_metrics_html = ""
    percentage_metrics_html = ""
    percentage_explanation_html = ""
    percentage_section_payloads: List[Dict[str, Any]] = []
    if nn_df is None:
        nn_note = nn_status
    else:
        nn_summary, nn_note = compute_nn_summary(
            nn_df,
            "monthly" if rtype == "monthly" else "yearly",
            period_end,
            time_entries_filtered,
            all_time_entries_df=time_entries_df,
            asof_date=asof_date,
        )
        if nn_note:
            nn_note = f"Hours-remaining source: {nn_note}"
    sideways_bar_chart_html = build_nn_sideways_bar_chart_html(nn_summary)

    if rtype in ("monthly", "yearly"):
        hours_metrics_html = build_nn_metrics_html(nn_summary, nn_note)
        percentage_metrics_html = hours_metrics_html
    if rtype == "daily":
        table_only_html = build_logged_hours_breakdown_html(
            time_entries_filtered,
            show_percentage=True,
            include_total_in_note=True,
            foldable=False,
        )
        hours_metrics_html = table_only_html
        percentage_metrics_html = table_only_html
    elif rtype in ("weekly", "biweekly"):
        hours_metrics_html = build_logged_hours_breakdown_html(time_entries_filtered)
        percentage_metrics_html = build_logged_hours_breakdown_html(time_entries_filtered, show_percentage=True)

    projects_for_counts = projects_df
    if rtype in ("daily", "weekly", "biweekly", "monthly"):
        projects_for_counts = filter_projects_with_hours(projects_df, time_entries_filtered)

    if rtype == "daily":
        counts_fig = go.Figure()
        hours_fig = go.Figure()
        percentage_fig = go.Figure()
        sideways_bar_chart_html = ""
        percentage_explanation_html = ""
        percentage_section_payloads = []
    else:
        counts_fig = build_counts_figure(
            projects_for_counts,
            export_date,
            period_start,
            period_end,
            period_label,
            project_color_map=project_color_map,
            timeline_projects_df=projects_df,
            timeline_year=timeline_year,
        )
        hours_fig = build_hours_figure(
            projects_df,
            time_entries_filtered,
            export_date,
            period_start,
            period_end,
            period_label,
            report_type=rtype,
            exclude_project_ids=complete_missing_hours_project_ids,
        )
        total_period_hours = (
            float(pd.to_numeric(time_entries_filtered["duration_hours"], errors="coerce").fillna(0.0).sum())
            if not time_entries_filtered.empty and "duration_hours" in time_entries_filtered.columns
            else 0.0
        )
        weekly_ref, weekly_ref_source, weekly_ref_note = compute_weekly_reference_hours(
            time_entries_df, period_end, nn_summary
        )
        weekly_progress_guidance: Optional[Dict[str, Any]] = None
        percentage_explanation_html = ""
        if rtype == "yearly":
            weekly_progress_guidance = compute_weekly_progress_guidance(
                time_entries_df,
                period_end,
                nn_summary,
                weekly_ref,
            )
            percentage_explanation_html = build_weekly_progress_explanation_html(
                weekly_progress_guidance,
                weekly_ref_source,
            )
        elif rtype == "monthly":
            weekly_progress_guidance = compute_monthly_average_guidance(
                time_entries_filtered,
                period_start,
                period_end,
                weekly_ref,
                nn_summary=nn_summary,
            )
            percentage_explanation_html = build_monthly_average_explanation_html(
                weekly_progress_guidance,
            )
        percentage_fig = build_percentage_figure_from_hours(
            hours_fig,
            total_period_hours=total_period_hours,
            weekly_reference_hours=weekly_ref,
            weekly_reference_note=weekly_ref_note,
            show_weekly_reference_note_in_title=False,
            weekly_progress_guidance=weekly_progress_guidance,
        )
        percentage_section_payloads = build_percentage_section_payloads(
            build_hours_section_figures(
                projects_df,
                time_entries_filtered,
                period_start,
                period_end,
                rtype,
                exclude_project_ids=complete_missing_hours_project_ids,
            ),
            total_period_hours=total_period_hours,
            weekly_reference_hours=weekly_ref,
            weekly_progress_guidance=weekly_progress_guidance,
            timeline_explanation_html=percentage_explanation_html,
        )

    period_range = format_period_range_compact(period_start, period_end)
    header_context = dict(
        title_text=REPORT_TITLE,
        person_name=PERSON_NAME,
        export_date=export_date,
        period_label=period_label,
        period_range=period_range,
        **header_assets,
    )

    if rtype == "yearly":
        base_name = "project_report_yearly"
        archive_base_name = f"project_report_yearly_{period_key}"
    elif rtype == "monthly":
        base_name = f"project_report_monthly_{period_key}"
        archive_base_name = base_name
    elif rtype == "daily":
        base_name = f"project_report_daily_{period_key}"
        archive_base_name = base_name
    elif rtype == "biweekly":
        base_name = f"project_report_biweekly_{period_key}"
        archive_base_name = base_name
    else:
        base_name = f"project_report_weekly_{period_key}"
        archive_base_name = base_name

    html_path, png_path, lite_html_path = export_tabbed_report(
        counts_fig,
        hours_fig,
        percentage_fig,
        REPORT_DIR,
        REPORTS_ARCHIVE_DIR,
        base_name,
        archive_base_name,
        export_date,
        header_context,
        projects_page_html,
        hours_metrics_html,
        percentage_metrics_html,
        percentage_explanation_html,
        percentage_section_payloads,
        sideways_bar_chart_html,
        nn_note,
        nn_summary,
    )

    print(f"Generated {rtype} report: {period_range} -> {html_path}")
    print(f"Generated lite report -> {lite_html_path}")
    print(f"PNG exported: {png_path}")


if __name__ == "__main__":
    main()
