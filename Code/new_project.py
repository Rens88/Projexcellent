#!/usr/bin/env python3
"""Create a new Projexcellent project folder from templates."""

from __future__ import annotations

import argparse
import re
import shutil
from datetime import date
from pathlib import Path
from typing import Optional

from projexcellent_config import DEFAULT_CONFIG_PATH, load_config, resolve_path


STATUS_CHOICES = ["Proposed", "Active", "On-hold", "Closed", "Cancelled"]
PRIORITY_CHOICES = ["Low", "Medium", "High", "Critical"]


def sanitize_slug(value: str) -> str:
    slug = value.strip().lower()
    slug = re.sub(r"[^a-z0-9]+", "_", slug)
    slug = re.sub(r"_+", "_", slug).strip("_")
    if not slug:
        raise SystemExit("Slug is empty after sanitization. Use letters and numbers.")
    return slug


def set_kv(ws, key: str, value) -> None:
    for row in range(1, ws.max_row + 1):
        key_cell = ws.cell(row=row, column=1).value
        if str(key_cell).strip() == key:
            ws.cell(row=row, column=2).value = value
            return
    new_row = ws.max_row + 1
    ws.cell(row=new_row, column=1).value = key
    ws.cell(row=new_row, column=2).value = value


def parse_year(value: str) -> int:
    raw = str(value).strip()
    if not re.fullmatch(r"\d{4}", raw):
        raise argparse.ArgumentTypeError("Year must use YYYY format (e.g. 2026).")
    year = int(raw)
    if year < 1900 or year > 9999:
        raise argparse.ArgumentTypeError("Year must be between 1900 and 9999.")
    return year


def _normalize_choice(raw: str, choices: list[str], aliases: Optional[dict[str, str]] = None) -> str:
    text = str(raw).strip()
    if not text:
        raise SystemExit("Input value is empty.")
    aliases = aliases or {}
    key = text.casefold()
    for choice in choices:
        if key == choice.casefold():
            return choice
    if key in aliases:
        return aliases[key]
    valid = ", ".join(choices)
    raise SystemExit(f"Invalid value '{raw}'. Valid options (case-insensitive): {valid}.")


def normalize_status(raw: str) -> str:
    aliases = {
        "on hold": "On-hold",
        "on_hold": "On-hold",
        "onhold": "On-hold",
        "canceled": "Cancelled",
    }
    return _normalize_choice(raw, STATUS_CHOICES, aliases=aliases)


def normalize_priority(raw: str) -> str:
    return _normalize_choice(raw, PRIORITY_CHOICES)


def get_next_counter(projects_dir: Path, year: int) -> int:
    if not projects_dir.exists():
        return 1
    max_counter = 0
    pattern = re.compile(rf"^{year:04d}_(\d{{4,}})_")
    for child in projects_dir.iterdir():
        if not child.is_dir():
            continue
        match = pattern.match(child.name)
        if not match:
            continue
        try:
            max_counter = max(max_counter, int(match.group(1)))
        except ValueError:
            continue
    return max_counter + 1


def parse_args() -> argparse.Namespace:
    this_year = date.today().year
    parser = argparse.ArgumentParser(description="Create a new Projexcellent project.")
    parser.add_argument("--year", type=parse_year, default=this_year, help=f"Project year YYYY (default: {this_year}).")
    parser.add_argument(
        "--counter",
        type=int,
        help="Optional counter override. If omitted, the next available counter for the year is used.",
    )
    parser.add_argument("--slug", required=True, help="Short slug for folder name.")
    parser.add_argument("--project-name", required=True, help="Human-readable project name.")
    parser.add_argument("--programma", default="Other", help="Programma value.")
    parser.add_argument("--theme", default="General", help="Theme value.")
    parser.add_argument("--owner", default="", help="Owner name.")
    parser.add_argument("--requester", default="Unknown", help="Requester name.")
    parser.add_argument(
        "--status",
        default="Proposed",
        help="Initial project status (case-insensitive): Proposed, Active, On-hold, Closed, Cancelled.",
    )
    parser.add_argument(
        "--priority",
        default="Medium",
        help="Initial project priority (case-insensitive): Low, Medium, High, Critical.",
    )
    parser.add_argument("--force", action="store_true", help="Overwrite target folder if it already exists.")
    parser.add_argument(
        "--config",
        default=DEFAULT_CONFIG_PATH,
        help="Path to config JSON (default: projexcellent_config.json).",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    config = load_config(args.config)
    paths_cfg = config.get("paths", {})
    projects_dir = Path(resolve_path(config, paths_cfg.get("projects_dir", "Projecten"))).resolve()
    templates_dir = Path(resolve_path(config, paths_cfg.get("templates_dir", "Templates"))).resolve()
    project_info_template = templates_dir / "project_info_template.xlsx"
    time_log_template = templates_dir / "time_log_template.xlsx"

    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:  # pragma: no cover
        raise SystemExit(
            "openpyxl is required. Run `make_report.bat` (Windows) or `./make_report.command` (macOS/Linux) once first to install dependencies."
        ) from exc

    if not project_info_template.exists():
        raise SystemExit(f"Missing template: {project_info_template}")
    if not time_log_template.exists():
        raise SystemExit(f"Missing template: {time_log_template}")

    counter = args.counter if args.counter is not None else get_next_counter(projects_dir, args.year)
    if counter <= 0:
        raise SystemExit("Counter must be a positive integer.")

    slug = sanitize_slug(args.slug)
    status = normalize_status(args.status)
    priority = normalize_priority(args.priority)
    project_id = f"{args.year:04d}_{counter:04d}"
    folder_name = f"{project_id}_{slug}"
    project_dir = projects_dir / folder_name
    deliverables_dir = project_dir / "Deliverables"

    projects_dir.mkdir(parents=True, exist_ok=True)

    if project_dir.exists():
        if not args.force:
            raise SystemExit(f"Project folder already exists: {project_dir}")
        shutil.rmtree(project_dir)

    project_dir.mkdir(parents=True, exist_ok=True)
    deliverables_dir.mkdir(parents=True, exist_ok=True)

    project_info_path = project_dir / "project_info.xlsx"
    time_log_path = project_dir / "time_log.xlsx"
    shutil.copy2(project_info_template, project_info_path)
    shutil.copy2(time_log_template, time_log_path)

    today = date.today()
    info_wb = load_workbook(project_info_path)
    if "ProjectInfo" not in info_wb.sheetnames:
        raise SystemExit(f"Template missing sheet 'ProjectInfo': {project_info_path}")
    info_ws = info_wb["ProjectInfo"]
    set_kv(info_ws, "project_id", project_id)
    set_kv(info_ws, "project_name", args.project_name.strip())
    set_kv(info_ws, "programma (if multiple, separate by |)", args.programma.strip())
    set_kv(info_ws, "theme (if multiple, separate by |)", args.theme.strip())
    set_kv(info_ws, "owner", args.owner.strip())
    set_kv(info_ws, "requester", args.requester.strip())
    set_kv(info_ws, "status", status)
    set_kv(info_ws, "priority", priority)
    set_kv(info_ws, "start_date", today)
    set_kv(info_ws, "target_end_date", "")
    set_kv(info_ws, "actual_end_date", "")
    set_kv(info_ws, "created_at", today)
    set_kv(info_ws, "last_updated", today)
    set_kv(info_ws, "notes", "Project initialized with new_project.py")
    info_wb.save(project_info_path)

    timelog_wb = load_workbook(time_log_path)
    if "TimeLog" not in timelog_wb.sheetnames:
        raise SystemExit(f"Template missing sheet 'TimeLog': {time_log_path}")
    timelog_ws = timelog_wb["TimeLog"]
    timelog_ws["B1"] = project_id
    timelog_ws["B2"] = args.project_name.strip()
    timelog_ws["B3"] = args.programma.strip()
    timelog_wb.save(time_log_path)

    milestones_path = deliverables_dir / "milestones.txt"
    milestones_path.write_text(
        "\n".join(
            [
                f"Project: {project_id} - {args.project_name.strip()}",
                "",
                "Milestones",
                "1. Kickoff and scope alignment completed.",
                "2. First tangible deliverable drafted.",
                "3. Review feedback incorporated.",
                "4. Final deliverable published.",
                "",
                "Notes",
                "- Keep deliverables (text/images) in this folder.",
                "- Time tracking can be rough; consistency matters more than precision.",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    print(f"Created project: {project_dir}")
    print(f"- {project_info_path.name}")
    print(f"- {time_log_path.name}")
    print(f"- Deliverables/{milestones_path.name}")


if __name__ == "__main__":
    main()
